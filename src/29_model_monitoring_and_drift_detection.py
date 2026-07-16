from __future__ import annotations

"""
HeveMind Script 29
Model Monitoring and Drift Detection
Part 1 of 6: Foundation, configuration and data loading
"""

import hashlib
import importlib.util
import json
import math
import os
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
import sqlite3

import numpy as np
import pandas as pd


# ============================================================
# PROJECT PATHS
# ============================================================
ROOT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT_DIR / "src"
DATA_DIR = ROOT_DIR / "data"
ARTIFACTS_DIR = ROOT_DIR / "artifacts"
REPORTS_DIR = ROOT_DIR / "reports"
LOG_DIR = ROOT_DIR / "logs"
CONFIG_DIR = ROOT_DIR / "config"

MONITORING_DIR = REPORTS_DIR / "model_monitoring"
TABLES_DIR = MONITORING_DIR / "tables"
FIGURES_DIR = MONITORING_DIR / "figures"
HISTORY_DIR = MONITORING_DIR / "history"
EXPORTS_DIR = MONITORING_DIR / "exports"

CONFIG_PATH = CONFIG_DIR / "model_monitoring_config.json"
HISTORY_DATABASE_PATH = HISTORY_DIR / "model_monitoring_history.db"
HISTORY_CSV_PATH = HISTORY_DIR / "monitoring_history.csv"

FEATURE_DRIFT_PATH = TABLES_DIR / "feature_drift.csv"
PREDICTION_DRIFT_PATH = TABLES_DIR / "prediction_drift.csv"
DATA_QUALITY_PATH = TABLES_DIR / "data_quality_monitoring.csv"
ALERTS_PATH = TABLES_DIR / "monitoring_alerts.csv"
HEALTH_COMPONENTS_PATH = TABLES_DIR / "health_components.csv"

DRIFT_SUMMARY_PATH = MONITORING_DIR / "drift_summary.json"
DEPLOYMENT_HEALTH_PATH = MONITORING_DIR / "deployment_health.json"
EXCEL_REPORT_PATH = EXPORTS_DIR / "hevemind_model_monitoring.xlsx"
PDF_REPORT_PATH = EXPORTS_DIR / "hevemind_model_monitoring.pdf"

AUDIT_ENGINE_PATH = SRC_DIR / "24_audit_activity_tracking.py"

APP_NAME = "HeveMind Model Monitoring and Drift Detection"
APP_VERSION = os.getenv("HEVEMIND_MONITORING_VERSION", "1.0.0").strip()

DEFAULT_STANDALONE_ADMIN_MODE = False

DEFAULT_REFERENCE_CANDIDATES = [
    DATA_DIR / "processed" / "development.csv",
    DATA_DIR / "processed" / "development.parquet",
    DATA_DIR / "processed" / "development_dataset.csv",
    DATA_DIR / "processed" / "development_dataset.parquet",
    DATA_DIR / "processed" / "train.csv",
    DATA_DIR / "processed" / "train.parquet",
    ARTIFACTS_DIR / "development_dataset.csv",
    ARTIFACTS_DIR / "development_dataset.parquet",
    REPORTS_DIR / "uncertainty_engine" / "tables" / "development_uncertainty_scores.csv",
]

DEFAULT_CURRENT_CANDIDATES = [
    DATA_DIR / "processed" / "test.csv",
    DATA_DIR / "processed" / "test.parquet",
    DATA_DIR / "processed" / "test_dataset.csv",
    DATA_DIR / "processed" / "test_dataset.parquet",
    ARTIFACTS_DIR / "test_dataset.csv",
    ARTIFACTS_DIR / "test_dataset.parquet",
    REPORTS_DIR / "uncertainty_engine" / "tables" / "test_uncertainty_scores.csv",
    REPORTS_DIR / "deployment_backend" / "wafer_report_index.csv",
]

SUPPORTED_SUFFIXES = {
    ".csv",
    ".parquet",
    ".pq",
    ".feather",
    ".xlsx",
    ".xls",
    ".json",
}

COLUMN_CANDIDATES = {
    "id": ["wafer_id", "record_id", "sample_id", "id"],
    "timestamp": ["timestamp", "datetime", "event_time", "recorded_at"],
    "target": ["target", "label", "failure", "actual_failure", "y_true"],
    "probability": [
        "calibrated_failure_probability",
        "failure_probability",
        "predicted_probability",
        "probability",
        "y_score",
    ],
    "decision": [
        "uncertainty_adjusted_decision",
        "operational_decision",
        "decision",
    ],
    "confidence": [
        "prediction_confidence",
        "confidence",
        "model_confidence",
    ],
    "uncertainty": [
        "combined_uncertainty",
        "uncertainty",
        "prediction_uncertainty",
    ],
    "data_confidence": ["data_confidence", "input_data_confidence"],
    "familiarity": [
        "data_familiarity_band",
        "data_familiarity",
        "familiarity_band",
    ],
}

DEFAULT_CONFIG: dict[str, Any] = {
    "schema_version": 1,
    "application_version": APP_VERSION,
    "number_of_bins": 10,
    "minimum_reference_rows": 100,
    "minimum_current_rows": 30,
    "maximum_features": 590,
    "epsilon": 1e-8,
    "feature_prefixes": ["sensor_"],
    "exclude_columns": [],
    "reference_data_path": None,
    "current_data_path": None,
    "column_candidates": COLUMN_CANDIDATES,
    "feature_drift_thresholds": {
        "psi": {"watch": 0.10, "warning": 0.20, "critical": 0.35},
        "ks_statistic": {"watch": 0.10, "warning": 0.20, "critical": 0.30},
        "standardised_mean_shift": {
            "watch": 0.25,
            "warning": 0.50,
            "critical": 1.00,
        },
        "missing_rate_difference": {
            "watch": 0.03,
            "warning": 0.08,
            "critical": 0.15,
        },
    },
    "prediction_drift_thresholds": {
        "probability_mean_shift": {
            "watch": 0.02,
            "warning": 0.05,
            "critical": 0.10,
        },
        "confidence_mean_shift": {
            "watch": 0.05,
            "warning": 0.10,
            "critical": 0.20,
        },
        "uncertainty_mean_shift": {
            "watch": 0.05,
            "warning": 0.10,
            "critical": 0.20,
        },
        "decision_total_variation": {
            "watch": 0.08,
            "warning": 0.15,
            "critical": 0.25,
        },
    },
    "data_quality_thresholds": {
        "missing_rate": {"watch": 0.05, "warning": 0.15, "critical": 0.30},
        "duplicate_rate": {"watch": 0.01, "warning": 0.03, "critical": 0.10},
        "constant_feature_rate": {
            "watch": 0.02,
            "warning": 0.05,
            "critical": 0.10,
        },
        "unmatched_feature_rate": {
            "watch": 0.02,
            "warning": 0.05,
            "critical": 0.10,
        },
    },
    "health_weights": {
        "data_quality": 0.25,
        "feature_drift": 0.35,
        "prediction_drift": 0.25,
        "confidence_and_uncertainty": 0.15,
    },
    "created_utc": None,
    "updated_utc": None,
}


# ============================================================
# DATA MODELS
# ============================================================
@dataclass(frozen=True)
class ColumnMapping:
    id_column: str | None = None
    timestamp_column: str | None = None
    target_column: str | None = None
    probability_column: str | None = None
    decision_column: str | None = None
    confidence_column: str | None = None
    uncertainty_column: str | None = None
    data_confidence_column: str | None = None
    familiarity_column: str | None = None


@dataclass(frozen=True)
class DatasetDescriptor:
    dataset_name: str
    source_path: Path
    row_count: int
    column_count: int
    duplicate_rows: int
    total_missing_values: int
    numeric_columns: tuple[str, ...]
    categorical_columns: tuple[str, ...]
    feature_columns: tuple[str, ...]
    column_mapping: ColumnMapping
    sha256: str


@dataclass
class MonitoringDataset:
    name: str
    dataframe: pd.DataFrame
    descriptor: DatasetDescriptor


@dataclass(frozen=True)
class MonitoringRunIdentity:
    run_id: str
    generated_utc: str
    application_version: str
    reference_sha256: str
    current_sha256: str
    reference_path: str
    current_path: str


@dataclass
class MonitoringContext:
    config: dict[str, Any]
    run_identity: MonitoringRunIdentity
    reference: MonitoringDataset
    current: MonitoringDataset
    shared_features: list[str]
    reference_only_features: list[str]
    current_only_features: list[str]
    warnings: list[str] = field(default_factory=list)


# ============================================================
# GENERIC UTILITIES
# ============================================================
def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def ensure_directories() -> None:
    for path in [
        CONFIG_DIR,
        MONITORING_DIR,
        TABLES_DIR,
        FIGURES_DIR,
        HISTORY_DIR,
        EXPORTS_DIR,
    ]:
        path.mkdir(parents=True, exist_ok=True)


def safe_text(
    value: Any,
    fallback: str = "Unavailable",
) -> str:
    if value is None:
        return fallback

    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    return text or fallback


def safe_int(
    value: Any,
    fallback: int = 0,
) -> int:
    """
    Safely convert a value to an integer.

    Returns the fallback value when the input is missing,
    non-finite or cannot be converted.
    """
    try:
        if value is None:
            return fallback

        if isinstance(value, bool):
            return int(value)

        if isinstance(value, (int, np.integer)):
            return int(value)

        numeric = float(value)

        if not math.isfinite(numeric):
            return fallback

        return int(numeric)

    except (TypeError, ValueError, OverflowError):
        return fallback


def safe_float(
    value: Any,
    fallback: float | None = None,
) -> float | None:
    try:
        numeric = float(value)
    except (TypeError, ValueError, OverflowError):
        return fallback

    return numeric if math.isfinite(numeric) else fallback


def normalise_column_name(value: Any) -> str:
    text = safe_text(value, "")
    result: list[str] = []
    separator = False

    for character in text:
        if character.isalnum():
            result.append(character.lower())
            separator = False
        elif not separator:
            result.append("_")
            separator = True

    return "".join(result).strip("_")


def json_safe(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, (str, int, bool)):
        return value

    if isinstance(value, float):
        return value if math.isfinite(value) else None

    if isinstance(value, Path):
        return str(value)

    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat()

    if isinstance(value, Mapping):
        return {
            str(key): json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, (list, tuple, set)):
        return [
            json_safe(item)
            for item in value
        ]

    if hasattr(value, "item"):
        try:
            return json_safe(value.item())
        except Exception:
            pass

    return str(value)


def atomic_write_json(
    path: Path,
    payload: Mapping[str, Any],
) -> None:
    ensure_directories()

    temporary = path.with_suffix(
        path.suffix + ".tmp"
    )

    temporary.write_text(
        json.dumps(
            json_safe(dict(payload)),
            indent=4,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    temporary.replace(path)


def load_json_object(path: Path) -> dict[str, Any]:
    payload = json.loads(
        path.read_text(
            encoding="utf-8"
        )
    )

    if not isinstance(payload, dict):
        raise RuntimeError(
            f"Expected a JSON object in: {path}"
        )

    return payload


def deep_merge(
    base: Mapping[str, Any],
    override: Mapping[str, Any],
) -> dict[str, Any]:
    output = dict(base)

    for key, value in override.items():
        if (
            isinstance(output.get(key), dict)
            and isinstance(value, Mapping)
        ):
            output[key] = deep_merge(
                output[key],
                value,
            )
        else:
            output[key] = value

    return output


def calculate_file_sha256(path: Path) -> str:
    digest = hashlib.sha256()

    with path.open("rb") as file:
        for chunk in iter(
            lambda: file.read(1024 * 1024),
            b"",
        ):
            digest.update(chunk)

    return digest.hexdigest()


def ordered_unique(
    values: Iterable[str],
) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []

    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)

    return result


# ============================================================
# CONFIGURATION
# ============================================================
def initialise_monitoring_config() -> dict[str, Any]:
    ensure_directories()
    if CONFIG_PATH.exists():
        return deep_merge(DEFAULT_CONFIG, load_json_object(CONFIG_PATH))

    payload = deep_merge(
        DEFAULT_CONFIG,
        {
            "created_utc": utc_now_iso(),
            "updated_utc": utc_now_iso(),
        },
    )
    atomic_write_json(CONFIG_PATH, payload)
    return payload


def validate_monitoring_config(config: Mapping[str, Any]) -> None:
    if int(config.get("number_of_bins", 10)) < 3:
        raise ValueError("number_of_bins must be at least 3.")

    weights = config.get("health_weights", {})
    if not isinstance(weights, Mapping):
        raise ValueError("health_weights must be a JSON object.")

    weight_sum = sum(float(value) for value in weights.values())
    if not math.isclose(weight_sum, 1.0, rel_tol=1e-6, abs_tol=1e-6):
        raise ValueError("health_weights must sum to 1.0.")


# ============================================================
# TABULAR DATA LOADING
# ============================================================
def load_tabular_file(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Data file was not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"Unsupported tabular file type: {suffix}")

    if suffix == ".csv":
        dataframe = pd.read_csv(path, low_memory=False)
    elif suffix in {".parquet", ".pq"}:
        dataframe = pd.read_parquet(path)
    elif suffix == ".feather":
        dataframe = pd.read_feather(path)
    elif suffix in {".xlsx", ".xls"}:
        dataframe = pd.read_excel(path)
    else:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            dataframe = pd.DataFrame(payload)
        elif isinstance(payload, dict):
            rows = next(
                (
                    payload[key]
                    for key in ["items", "records", "data", "rows"]
                    if isinstance(payload.get(key), list)
                ),
                [payload],
            )
            dataframe = pd.DataFrame(rows)
        else:
            raise RuntimeError("JSON data must contain an object or list.")

    if dataframe.empty:
        raise RuntimeError(f"The loaded dataset is empty: {path}")

    dataframe = dataframe.copy()
    dataframe.columns = [safe_text(column, f"column_{index}") for index, column in enumerate(dataframe.columns)]
    return dataframe


def resolve_existing_path(
    explicit_path: Path | None,
    configured_path: Any,
    candidates: Sequence[Path],
    description: str,
) -> Path:
    if explicit_path is not None:
        if explicit_path.exists():
            return explicit_path.resolve()
        raise FileNotFoundError(f"{description} was not found: {explicit_path}")

    configured_text = safe_text(configured_path, "")
    if configured_text:
        configured = Path(configured_text).expanduser()
        if not configured.is_absolute():
            configured = ROOT_DIR / configured
        if configured.exists():
            return configured.resolve()

    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()

    checked = "\n".join(f"  - {path}" for path in candidates)
    raise FileNotFoundError(
        f"Unable to resolve {description}. Checked:\n{checked}"
    )


# ============================================================
# COLUMN DISCOVERY
# ============================================================
def find_first_column(dataframe: pd.DataFrame, candidates: Sequence[str]) -> str | None:
    lookup = {
        normalise_column_name(column): str(column)
        for column in dataframe.columns
    }
    for candidate in candidates:
        match = lookup.get(normalise_column_name(candidate))
        if match is not None:
            return match
    return None


def discover_column_mapping(
    dataframe: pd.DataFrame,
    config: Mapping[str, Any],
) -> ColumnMapping:
    configured = config.get("column_candidates", COLUMN_CANDIDATES)
    return ColumnMapping(
        id_column=find_first_column(dataframe, configured.get("id", COLUMN_CANDIDATES["id"])),
        timestamp_column=find_first_column(dataframe, configured.get("timestamp", COLUMN_CANDIDATES["timestamp"])),
        target_column=find_first_column(dataframe, configured.get("target", COLUMN_CANDIDATES["target"])),
        probability_column=find_first_column(dataframe, configured.get("probability", COLUMN_CANDIDATES["probability"])),
        decision_column=find_first_column(dataframe, configured.get("decision", COLUMN_CANDIDATES["decision"])),
        confidence_column=find_first_column(dataframe, configured.get("confidence", COLUMN_CANDIDATES["confidence"])),
        uncertainty_column=find_first_column(dataframe, configured.get("uncertainty", COLUMN_CANDIDATES["uncertainty"])),
        data_confidence_column=find_first_column(dataframe, configured.get("data_confidence", COLUMN_CANDIDATES["data_confidence"])),
        familiarity_column=find_first_column(dataframe, configured.get("familiarity", COLUMN_CANDIDATES["familiarity"])),
    )


def discover_feature_columns(
    dataframe: pd.DataFrame,
    mapping: ColumnMapping,
    config: Mapping[str, Any],
) -> list[str]:
    excluded = {
        value
        for value in asdict(mapping).values()
        if value is not None
    }
    excluded.update(str(value) for value in config.get("exclude_columns", []))

    numeric_columns = dataframe.select_dtypes(
        include=[np.number, "bool"]
    ).columns.tolist()

    prefixes = [str(value).lower() for value in config.get("feature_prefixes", ["sensor_"])]
    prefixed = [
        column
        for column in numeric_columns
        if any(str(column).lower().startswith(prefix) for prefix in prefixes)
    ]

    candidates = prefixed or numeric_columns
    features = [str(column) for column in candidates if column not in excluded]
    maximum = int(config.get("maximum_features", 590))
    return ordered_unique(features[:maximum] if maximum > 0 else features)


# ============================================================
# DATASET PREPARATION
# ============================================================
def load_monitoring_dataset(
    name: str,
    source_path: Path,
    config: Mapping[str, Any],
) -> MonitoringDataset:
    dataframe = load_tabular_file(source_path)
    mapping = discover_column_mapping(dataframe, config)
    features = discover_feature_columns(dataframe, mapping, config)

    if not features:
        raise RuntimeError(f"No numerical monitoring features found in {source_path}")

    for feature in features:
        dataframe[feature] = pd.to_numeric(dataframe[feature], errors="coerce")

    for column in [
        mapping.target_column,
        mapping.probability_column,
        mapping.confidence_column,
        mapping.uncertainty_column,
        mapping.data_confidence_column,
    ]:
        if column is not None:
            dataframe[column] = pd.to_numeric(dataframe[column], errors="coerce")

    if mapping.timestamp_column is not None:
        dataframe[mapping.timestamp_column] = pd.to_datetime(
            dataframe[mapping.timestamp_column],
            errors="coerce",
            utc=True,
        )

    numeric = tuple(
        str(column)
        for column in dataframe.select_dtypes(include=[np.number, "bool"]).columns
    )
    categorical = tuple(
        str(column)
        for column in dataframe.columns
        if str(column) not in numeric
    )

    descriptor = DatasetDescriptor(
        dataset_name=name,
        source_path=source_path,
        row_count=len(dataframe),
        column_count=len(dataframe.columns),
        duplicate_rows=int(dataframe.duplicated().sum()),
        total_missing_values=int(dataframe.isna().sum().sum()),
        numeric_columns=numeric,
        categorical_columns=categorical,
        feature_columns=tuple(features),
        column_mapping=mapping,
        sha256=calculate_file_sha256(source_path),
    )
    return MonitoringDataset(name=name, dataframe=dataframe, descriptor=descriptor)


def build_monitoring_context(
    reference_path: Path | None = None,
    current_path: Path | None = None,
    config_path: Path | None = None,
) -> MonitoringContext:
    ensure_directories()

    config = (
        deep_merge(DEFAULT_CONFIG, load_json_object(config_path))
        if config_path is not None
        else initialise_monitoring_config()
    )
    validate_monitoring_config(config)

    resolved_reference = resolve_existing_path(
        reference_path,
        config.get("reference_data_path"),
        DEFAULT_REFERENCE_CANDIDATES,
        "reference dataset",
    )
    resolved_current = resolve_existing_path(
        current_path,
        config.get("current_data_path"),
        DEFAULT_CURRENT_CANDIDATES,
        "current deployment dataset",
    )

    reference = load_monitoring_dataset("Reference", resolved_reference, config)
    current = load_monitoring_dataset("Current", resolved_current, config)

    minimum_reference = int(config.get("minimum_reference_rows", 100))
    minimum_current = int(config.get("minimum_current_rows", 30))

    if reference.descriptor.row_count < minimum_reference:
        raise RuntimeError(
            f"Reference dataset has {reference.descriptor.row_count} rows; "
            f"at least {minimum_reference} are required."
        )
    if current.descriptor.row_count < minimum_current:
        raise RuntimeError(
            f"Current dataset has {current.descriptor.row_count} rows; "
            f"at least {minimum_current} are required."
        )

    reference_features = list(reference.descriptor.feature_columns)
    current_features = list(current.descriptor.feature_columns)
    reference_set = set(reference_features)
    current_set = set(current_features)

    shared = [feature for feature in reference_features if feature in current_set]
    reference_only = [feature for feature in reference_features if feature not in current_set]
    current_only = [feature for feature in current_features if feature not in reference_set]

    if not shared:
        raise RuntimeError(
            "Reference and current datasets do not share monitoring features."
        )

    generated_utc = utc_now_iso()
    run_id = hashlib.sha256(
        (
            reference.descriptor.sha256
            + current.descriptor.sha256
            + generated_utc
        ).encode("utf-8")
    ).hexdigest()[:16]

    warnings_list: list[str] = []
    if reference_only:
        warnings_list.append(
            f"{len(reference_only)} reference features are absent from current data."
        )
    if current_only:
        warnings_list.append(
            f"{len(current_only)} current features are absent from reference data."
        )
    if current.descriptor.duplicate_rows:
        warnings_list.append("Current data contains duplicated rows.")

    return MonitoringContext(
        config=config,
        run_identity=MonitoringRunIdentity(
            run_id=run_id,
            generated_utc=generated_utc,
            application_version=APP_VERSION,
            reference_sha256=reference.descriptor.sha256,
            current_sha256=current.descriptor.sha256,
            reference_path=str(resolved_reference),
            current_path=str(resolved_current),
        ),
        reference=reference,
        current=current,
        shared_features=shared,
        reference_only_features=reference_only,
        current_only_features=current_only,
        warnings=warnings_list,
    )


# ============================================================
# OPTIONAL AUDIT INTEGRATION
# ============================================================
def load_audit_engine() -> Any | None:
    if not AUDIT_ENGINE_PATH.exists():
        return None

    specification = importlib.util.spec_from_file_location(
        "hevemind_audit_activity_tracking",
        AUDIT_ENGINE_PATH,
    )
    if specification is None or specification.loader is None:
        return None

    module = importlib.util.module_from_spec(specification)
    sys.modules["hevemind_audit_activity_tracking"] = module

    try:
        specification.loader.exec_module(module)
    except Exception:
        return None
    return module


def record_monitoring_event(
    *,
    event_type: str,
    outcome: str,
    action: str,
    resource_id: str | None,
    details: Mapping[str, Any] | None = None,
    actor: Mapping[str, Any] | None = None,
) -> None:
    audit_engine = load_audit_engine()
    if audit_engine is None:
        return

    actor_payload = dict(actor or {})
    try:
        audit_engine.record_event(
            source_component="29_model_monitoring_and_drift_detection",
            event_type=event_type,
            category="governance",
            outcome=outcome,
            user_email=actor_payload.get("email"),
            display_name=actor_payload.get("display_name"),
            role=actor_payload.get("role"),
            auth_source=actor_payload.get("auth_source"),
            session_id=actor_payload.get("session_id"),
            resource_type="model_monitoring_run",
            resource_id=resource_id,
            action=action,
            message="HeveMind model monitoring activity recorded.",
            details=json_safe(dict(details or {})),
        )
    except Exception:
        pass


def build_context_summary(context: MonitoringContext) -> dict[str, Any]:
    return {
        "run_identity": asdict(context.run_identity),
        "reference": json_safe(asdict(context.reference.descriptor)),
        "current": json_safe(asdict(context.current.descriptor)),
        "shared_feature_count": len(context.shared_features),
        "reference_only_feature_count": len(context.reference_only_features),
        "current_only_feature_count": len(context.current_only_features),
        "warnings": context.warnings,
    }


# End of Part 1.
# Append Part 2 immediately below this line.


# ============================================================
# PART 2 — CORE STATISTICAL FEATURE DRIFT ENGINE
# ============================================================

SEVERITY_ORDER = {
    "Normal": 0,
    "Watch": 1,
    "Warning": 2,
    "Critical": 3,
}

SEVERITY_SCORE = {
    "Normal": 100.0,
    "Watch": 82.0,
    "Warning": 58.0,
    "Critical": 25.0,
}


@dataclass(frozen=True)
class NumericDriftMetrics:
    feature: str
    reference_count: int
    current_count: int
    reference_missing_rate: float
    current_missing_rate: float
    missing_rate_difference: float
    reference_mean: float | None
    current_mean: float | None
    absolute_mean_shift: float | None
    standardised_mean_shift: float | None
    reference_std: float | None
    current_std: float | None
    standard_deviation_ratio: float | None
    reference_median: float | None
    current_median: float | None
    absolute_median_shift: float | None
    psi: float | None
    ks_statistic: float | None
    ks_p_value: float | None
    wasserstein_distance: float | None
    reference_constant: bool
    current_constant: bool
    metric_severities: dict[str, str]
    overall_severity: str
    feature_health_score: float
    drift_detected: bool
    notes: str


@dataclass(frozen=True)
class FeatureDriftSummary:
    monitored_features: int
    normal_features: int
    watch_features: int
    warning_features: int
    critical_features: int
    drifted_features: int
    drift_rate: float
    mean_feature_health_score: float
    maximum_psi: float | None
    maximum_ks_statistic: float | None
    maximum_standardised_mean_shift: float | None
    maximum_missing_rate_difference: float | None
    dominant_severity: str


def finite_numeric_array(values: Any) -> np.ndarray:
    series = pd.to_numeric(
        pd.Series(values),
        errors="coerce",
    )

    array = series.to_numpy(
        dtype=float,
        copy=False,
    )

    return array[
        np.isfinite(array)
    ]


def numeric_series(values: Any) -> pd.Series:
    return pd.to_numeric(
        pd.Series(values),
        errors="coerce",
    )


def safe_mean(values: np.ndarray) -> float | None:
    if values.size == 0:
        return None

    result = float(
        np.mean(values)
    )

    return result if math.isfinite(result) else None


def safe_median(values: np.ndarray) -> float | None:
    if values.size == 0:
        return None

    result = float(
        np.median(values)
    )

    return result if math.isfinite(result) else None


def safe_standard_deviation(
    values: np.ndarray,
    ddof: int = 1,
) -> float | None:
    if values.size <= ddof:
        return None

    result = float(
        np.std(
            values,
            ddof=ddof,
        )
    )

    return result if math.isfinite(result) else None


def absolute_difference(
    first: float | None,
    second: float | None,
) -> float | None:
    if first is None or second is None:
        return None

    result = abs(
        second - first
    )

    return result if math.isfinite(result) else None


def safe_ratio(
    numerator: float | None,
    denominator: float | None,
    epsilon: float,
) -> float | None:
    if numerator is None or denominator is None:
        return None

    denominator_value = abs(
        denominator
    )

    if denominator_value < epsilon:
        if abs(numerator) < epsilon:
            return 1.0

        return None

    result = numerator / denominator

    return result if math.isfinite(result) else None


def standardised_mean_shift(
    reference_mean: float | None,
    current_mean: float | None,
    reference_std: float | None,
    epsilon: float,
) -> float | None:
    if reference_mean is None or current_mean is None:
        return None

    absolute_shift = abs(
        current_mean
        - reference_mean
    )

    if reference_std is None or reference_std < epsilon:
        if absolute_shift < epsilon:
            return 0.0

        return None

    result = absolute_shift / reference_std

    return result if math.isfinite(result) else None


def is_effectively_constant(
    values: np.ndarray,
    epsilon: float,
) -> bool:
    if values.size == 0:
        return False

    if values.size == 1:
        return True

    minimum = float(
        np.min(values)
    )

    maximum = float(
        np.max(values)
    )

    return abs(
        maximum
        - minimum
    ) <= epsilon


def severity_from_thresholds(
    value: float | None,
    thresholds: Mapping[str, Any],
    *,
    use_absolute_value: bool = True,
) -> str:
    if value is None:
        return "Normal"

    numeric = abs(
        value
    ) if use_absolute_value else value

    critical = safe_float(
        thresholds.get(
            "critical"
        )
    )

    warning = safe_float(
        thresholds.get(
            "warning"
        )
    )

    watch = safe_float(
        thresholds.get(
            "watch"
        )
    )

    if critical is not None and numeric >= critical:
        return "Critical"

    if warning is not None and numeric >= warning:
        return "Warning"

    if watch is not None and numeric >= watch:
        return "Watch"

    return "Normal"


def maximum_severity(
    severities: Iterable[str],
) -> str:
    severity_list = [
        severity
        for severity in severities
        if severity in SEVERITY_ORDER
    ]

    if not severity_list:
        return "Normal"

    return max(
        severity_list,
        key=lambda severity: SEVERITY_ORDER[
            severity
        ],
    )


def weighted_health_from_severities(
    metric_severities: Mapping[str, str],
    metric_weights: Mapping[str, float] | None = None,
) -> float:
    if not metric_severities:
        return 100.0

    weights = dict(
        metric_weights
        or {}
    )

    weighted_sum = 0.0
    total_weight = 0.0

    for metric_name, severity in metric_severities.items():
        weight = safe_float(
            weights.get(
                metric_name,
                1.0,
            ),
            1.0,
        )

        if weight is None or weight <= 0:
            continue

        weighted_sum += (
            SEVERITY_SCORE.get(
                severity,
                100.0,
            )
            * weight
        )

        total_weight += weight

    if total_weight <= 0:
        return 100.0

    return round(
        weighted_sum / total_weight,
        2,
    )


def generate_quantile_bin_edges(
    reference_values: np.ndarray,
    number_of_bins: int,
    epsilon: float,
) -> np.ndarray:
    if reference_values.size == 0:
        return np.asarray(
            [
                -np.inf,
                np.inf,
            ],
            dtype=float,
        )

    finite_values = reference_values[
        np.isfinite(
            reference_values
        )
    ]

    if finite_values.size == 0:
        return np.asarray(
            [
                -np.inf,
                np.inf,
            ],
            dtype=float,
        )

    minimum = float(
        np.min(
            finite_values
        )
    )

    maximum = float(
        np.max(
            finite_values
        )
    )

    if abs(
        maximum
        - minimum
    ) <= epsilon:
        return np.asarray(
            [
                -np.inf,
                np.inf,
            ],
            dtype=float,
        )

    quantiles = np.linspace(
        0.0,
        1.0,
        number_of_bins
        + 1,
    )

    edges = np.quantile(
        finite_values,
        quantiles,
    )

    edges = np.unique(
        edges.astype(
            float
        )
    )

    if edges.size < 2:
        return np.asarray(
            [
                -np.inf,
                np.inf,
            ],
            dtype=float,
        )

    edges[
        0
    ] = -np.inf

    edges[
        -1
    ] = np.inf

    return edges


def histogram_proportions(
    values: np.ndarray,
    bin_edges: np.ndarray,
    epsilon: float,
) -> np.ndarray:
    if values.size == 0:
        bin_count = max(
            len(
                bin_edges
            )
            - 1,
            1,
        )

        return np.full(
            bin_count,
            1.0
            / bin_count,
            dtype=float,
        )

    counts, _ = np.histogram(
        values,
        bins=bin_edges,
    )

    proportions = counts.astype(
        float
    )

    total = float(
        proportions.sum()
    )

    if total <= 0:
        proportions = np.full(
            proportions.shape,
            1.0
            / max(
                len(
                    proportions
                ),
                1,
            ),
            dtype=float,
        )
    else:
        proportions /= total

    proportions = np.clip(
        proportions,
        epsilon,
        None,
    )

    proportions /= proportions.sum()

    return proportions


def population_stability_index(
    reference_values: np.ndarray,
    current_values: np.ndarray,
    number_of_bins: int,
    epsilon: float,
) -> float | None:
    if reference_values.size == 0 or current_values.size == 0:
        return None

    bin_edges = generate_quantile_bin_edges(
        reference_values,
        number_of_bins,
        epsilon,
    )

    reference_proportions = histogram_proportions(
        reference_values,
        bin_edges,
        epsilon,
    )

    current_proportions = histogram_proportions(
        current_values,
        bin_edges,
        epsilon,
    )

    psi_components = (
        current_proportions
        - reference_proportions
    ) * np.log(
        current_proportions
        / reference_proportions
    )

    result = float(
        np.sum(
            psi_components
        )
    )

    if not math.isfinite(
        result
    ):
        return None

    return max(
        result,
        0.0,
    )


def empirical_cdf_values(
    sorted_values: np.ndarray,
    evaluation_points: np.ndarray,
) -> np.ndarray:
    if sorted_values.size == 0:
        return np.zeros_like(
            evaluation_points,
            dtype=float,
        )

    return (
        np.searchsorted(
            sorted_values,
            evaluation_points,
            side="right",
        )
        / sorted_values.size
    )


def kolmogorov_smirnov_two_sample(
    reference_values: np.ndarray,
    current_values: np.ndarray,
) -> tuple[
    float | None,
    float | None,
]:
    if reference_values.size == 0 or current_values.size == 0:
        return None, None

    reference_sorted = np.sort(
        reference_values
    )

    current_sorted = np.sort(
        current_values
    )

    evaluation_points = np.unique(
        np.concatenate(
            [
                reference_sorted,
                current_sorted,
            ]
        )
    )

    reference_cdf = empirical_cdf_values(
        reference_sorted,
        evaluation_points,
    )

    current_cdf = empirical_cdf_values(
        current_sorted,
        evaluation_points,
    )

    statistic = float(
        np.max(
            np.abs(
                reference_cdf
                - current_cdf
            )
        )
    )

    reference_count = reference_sorted.size
    current_count = current_sorted.size

    effective_sample_size = (
        reference_count
        * current_count
        / (
            reference_count
            + current_count
        )
    )

    if effective_sample_size <= 0:
        return statistic, None

    scaled_statistic = (
        math.sqrt(
            effective_sample_size
        )
        + 0.12
        + 0.11
        / math.sqrt(
            effective_sample_size
        )
    ) * statistic

    if scaled_statistic <= 0:
        p_value = 1.0
    else:
        alternating_sum = 0.0

        for index in range(
            1,
            101,
        ):
            term = (
                (-1.0)
                ** (
                    index
                    - 1
                )
                * math.exp(
                    -2.0
                    * (
                        index
                        ** 2
                    )
                    * (
                        scaled_statistic
                        ** 2
                    )
                )
            )

            alternating_sum += term

            if abs(
                term
            ) < 1e-12:
                break

        p_value = min(
            max(
                2.0
                * alternating_sum,
                0.0,
            ),
            1.0,
        )

    return statistic, p_value


def one_dimensional_wasserstein_distance(
    reference_values: np.ndarray,
    current_values: np.ndarray,
) -> float | None:
    if reference_values.size == 0 or current_values.size == 0:
        return None

    reference_sorted = np.sort(
        reference_values
    )

    current_sorted = np.sort(
        current_values
    )

    quantile_count = max(
        min(
            max(
                reference_sorted.size,
                current_sorted.size,
            ),
            1000,
        ),
        20,
    )

    quantiles = np.linspace(
        0.0,
        1.0,
        quantile_count,
    )

    reference_quantiles = np.quantile(
        reference_sorted,
        quantiles,
    )

    current_quantiles = np.quantile(
        current_sorted,
        quantiles,
    )

    result = float(
        np.trapezoid(
            np.abs(
                reference_quantiles
                - current_quantiles
            ),
            quantiles,
        )
    )

    return result if math.isfinite(
        result
    ) else None


def build_feature_notes(
    *,
    reference_count: int,
    current_count: int,
    reference_constant: bool,
    current_constant: bool,
    standardised_shift: float | None,
    psi: float | None,
    missing_difference: float,
) -> str:
    notes: list[str] = []

    if reference_count == 0:
        notes.append(
            "No finite reference measurements."
        )

    if current_count == 0:
        notes.append(
            "No finite current measurements."
        )

    if reference_constant:
        notes.append(
            "Reference sensor is constant."
        )

    if current_constant:
        notes.append(
            "Current sensor is constant."
        )

    if reference_constant and not current_constant:
        notes.append(
            "A previously constant sensor now varies."
        )

    if not reference_constant and current_constant:
        notes.append(
            "Sensor became constant in current data."
        )

    if standardised_shift is None and reference_count > 0 and current_count > 0:
        notes.append(
            "Standardised mean shift is unavailable because reference variance is negligible."
        )

    if psi is None:
        notes.append(
            "PSI is unavailable."
        )

    if missing_difference > 0:
        notes.append(
            "Current missingness exceeds reference missingness."
        )

    return " ".join(
        notes
    ) or "No additional data-quality note."


def evaluate_numeric_feature_drift(
    feature: str,
    reference_series: pd.Series,
    current_series: pd.Series,
    config: Mapping[str, Any],
) -> NumericDriftMetrics:
    epsilon = safe_float(
        config.get(
            "epsilon"
        ),
        1e-8,
    ) or 1e-8

    number_of_bins = int(
        config.get(
            "number_of_bins",
            10,
        )
    )

    reference_numeric = numeric_series(
        reference_series
    )

    current_numeric = numeric_series(
        current_series
    )

    reference_missing_rate = float(
        reference_numeric.isna().mean()
    )

    current_missing_rate = float(
        current_numeric.isna().mean()
    )

    missing_difference = abs(
        current_missing_rate
        - reference_missing_rate
    )

    reference_values = finite_numeric_array(
        reference_numeric
    )

    current_values = finite_numeric_array(
        current_numeric
    )

    reference_mean = safe_mean(
        reference_values
    )

    current_mean = safe_mean(
        current_values
    )

    reference_median = safe_median(
        reference_values
    )

    current_median = safe_median(
        current_values
    )

    reference_std = safe_standard_deviation(
        reference_values
    )

    current_std = safe_standard_deviation(
        current_values
    )

    absolute_mean_shift = absolute_difference(
        reference_mean,
        current_mean,
    )

    standardised_shift = standardised_mean_shift(
        reference_mean,
        current_mean,
        reference_std,
        epsilon,
    )

    std_ratio = safe_ratio(
        current_std,
        reference_std,
        epsilon,
    )

    median_shift = absolute_difference(
        reference_median,
        current_median,
    )

    psi_value = population_stability_index(
        reference_values,
        current_values,
        number_of_bins,
        epsilon,
    )

    ks_statistic, ks_p_value = kolmogorov_smirnov_two_sample(
        reference_values,
        current_values,
    )

    wasserstein = one_dimensional_wasserstein_distance(
        reference_values,
        current_values,
    )

    reference_constant = is_effectively_constant(
        reference_values,
        epsilon,
    )

    current_constant = is_effectively_constant(
        current_values,
        epsilon,
    )

    threshold_config = config.get(
        "feature_drift_thresholds",
        {},
    )

    psi_severity = severity_from_thresholds(
        psi_value,
        threshold_config.get(
            "psi",
            {},
        ),
    )

    ks_severity = severity_from_thresholds(
        ks_statistic,
        threshold_config.get(
            "ks_statistic",
            {},
        ),
    )

    mean_severity = severity_from_thresholds(
        standardised_shift,
        threshold_config.get(
            "standardised_mean_shift",
            {},
        ),
    )

    missing_severity = severity_from_thresholds(
        missing_difference,
        threshold_config.get(
            "missing_rate_difference",
            {},
        ),
    )

    constant_severity = "Normal"

    if (
        not reference_constant
        and current_constant
        and current_values.size
        > 1
    ):
        constant_severity = "Critical"

    elif (
        reference_constant
        != current_constant
    ):
        constant_severity = "Warning"

    availability_severity = "Normal"

    if reference_values.size == 0:
        availability_severity = "Critical"

    elif current_values.size == 0:
        availability_severity = "Critical"

    metric_severities = {
        "psi": psi_severity,
        "ks_statistic": ks_severity,
        "standardised_mean_shift": mean_severity,
        "missing_rate_difference": missing_severity,
        "constant_sensor_change": constant_severity,
        "measurement_availability": availability_severity,
    }

    overall_severity = maximum_severity(
        metric_severities.values()
    )

    feature_health = weighted_health_from_severities(
        metric_severities,
        {
            "psi": 0.30,
            "ks_statistic": 0.25,
            "standardised_mean_shift": 0.20,
            "missing_rate_difference": 0.15,
            "constant_sensor_change": 0.07,
            "measurement_availability": 0.03,
        },
    )

    notes = build_feature_notes(
        reference_count=int(
            reference_values.size
        ),
        current_count=int(
            current_values.size
        ),
        reference_constant=reference_constant,
        current_constant=current_constant,
        standardised_shift=standardised_shift,
        psi=psi_value,
        missing_difference=(
            current_missing_rate
            - reference_missing_rate
        ),
    )

    return NumericDriftMetrics(
        feature=feature,
        reference_count=int(
            reference_values.size
        ),
        current_count=int(
            current_values.size
        ),
        reference_missing_rate=reference_missing_rate,
        current_missing_rate=current_missing_rate,
        missing_rate_difference=missing_difference,
        reference_mean=reference_mean,
        current_mean=current_mean,
        absolute_mean_shift=absolute_mean_shift,
        standardised_mean_shift=standardised_shift,
        reference_std=reference_std,
        current_std=current_std,
        standard_deviation_ratio=std_ratio,
        reference_median=reference_median,
        current_median=current_median,
        absolute_median_shift=median_shift,
        psi=psi_value,
        ks_statistic=ks_statistic,
        ks_p_value=ks_p_value,
        wasserstein_distance=wasserstein,
        reference_constant=reference_constant,
        current_constant=current_constant,
        metric_severities=metric_severities,
        overall_severity=overall_severity,
        feature_health_score=feature_health,
        drift_detected=(
            overall_severity
            != "Normal"
        ),
        notes=notes,
    )


def numeric_drift_metrics_to_record(
    metrics: NumericDriftMetrics,
) -> dict[str, Any]:
    record = asdict(
        metrics
    )

    metric_severities = record.pop(
        "metric_severities"
    )

    for metric_name, severity in metric_severities.items():
        record[
            f"{metric_name}_severity"
        ] = severity

    return json_safe(
        record
    )


def build_feature_drift_table(
    context: MonitoringContext,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []

    reference_dataframe = context.reference.dataframe
    current_dataframe = context.current.dataframe

    for feature in context.shared_features:
        metrics = evaluate_numeric_feature_drift(
            feature=feature,
            reference_series=reference_dataframe[
                feature
            ],
            current_series=current_dataframe[
                feature
            ],
            config=context.config,
        )

        records.append(
            numeric_drift_metrics_to_record(
                metrics
            )
        )

    feature_drift = pd.DataFrame(
        records
    )

    if feature_drift.empty:
        raise RuntimeError(
            "Feature drift evaluation produced no records."
        )

    feature_drift[
        "severity_rank"
    ] = feature_drift[
        "overall_severity"
    ].map(
        SEVERITY_ORDER
    ).fillna(
        0
    ).astype(
        int
    )

    feature_drift = feature_drift.sort_values(
        by=[
            "severity_rank",
            "feature_health_score",
            "psi",
            "ks_statistic",
        ],
        ascending=[
            False,
            True,
            False,
            False,
        ],
        na_position="last",
    ).reset_index(
        drop=True
    )

    feature_drift.insert(
        0,
        "monitoring_rank",
        np.arange(
            1,
            len(
                feature_drift
            )
            + 1,
        ),
    )

    return feature_drift


def summarise_feature_drift(
    feature_drift: pd.DataFrame,
) -> FeatureDriftSummary:
    severity_counts = (
        feature_drift[
            "overall_severity"
        ]
        .value_counts()
        .to_dict()
    )

    monitored_features = int(
        len(
            feature_drift
        )
    )

    normal_features = int(
        severity_counts.get(
            "Normal",
            0,
        )
    )

    watch_features = int(
        severity_counts.get(
            "Watch",
            0,
        )
    )

    warning_features = int(
        severity_counts.get(
            "Warning",
            0,
        )
    )

    critical_features = int(
        severity_counts.get(
            "Critical",
            0,
        )
    )

    drifted_features = (
        watch_features
        + warning_features
        + critical_features
    )

    drift_rate = (
        drifted_features
        / monitored_features
        if monitored_features
        else 0.0
    )

    health_values = pd.to_numeric(
        feature_drift[
            "feature_health_score"
        ],
        errors="coerce",
    )

    mean_health = float(
        health_values.mean()
    ) if health_values.notna().any() else 100.0

    def maximum_numeric(
        column: str,
    ) -> float | None:
        numeric = pd.to_numeric(
            feature_drift[
                column
            ],
            errors="coerce",
        )

        if not numeric.notna().any():
            return None

        result = float(
            numeric.max()
        )

        return result if math.isfinite(
            result
        ) else None

    dominant = maximum_severity(
        feature_drift[
            "overall_severity"
        ].tolist()
    )

    return FeatureDriftSummary(
        monitored_features=monitored_features,
        normal_features=normal_features,
        watch_features=watch_features,
        warning_features=warning_features,
        critical_features=critical_features,
        drifted_features=drifted_features,
        drift_rate=drift_rate,
        mean_feature_health_score=round(
            mean_health,
            2,
        ),
        maximum_psi=maximum_numeric(
            "psi"
        ),
        maximum_ks_statistic=maximum_numeric(
            "ks_statistic"
        ),
        maximum_standardised_mean_shift=maximum_numeric(
            "standardised_mean_shift"
        ),
        maximum_missing_rate_difference=maximum_numeric(
            "missing_rate_difference"
        ),
        dominant_severity=dominant,
    )


def build_unmatched_feature_records(
    context: MonitoringContext,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []

    for feature in context.reference_only_features:
        records.append(
            {
                "feature": feature,
                "feature_status": "Missing from current data",
                "severity": "Critical",
                "reference_available": True,
                "current_available": False,
            }
        )

    for feature in context.current_only_features:
        records.append(
            {
                "feature": feature,
                "feature_status": "New feature in current data",
                "severity": "Warning",
                "reference_available": False,
                "current_available": True,
            }
        )

    return pd.DataFrame(
        records,
        columns=[
            "feature",
            "feature_status",
            "severity",
            "reference_available",
            "current_available",
        ],
    )


def validate_feature_drift_output(
    feature_drift: pd.DataFrame,
    expected_feature_count: int,
) -> None:
    if feature_drift.empty:
        raise RuntimeError(
            "Feature drift table is empty."
        )

    if len(
        feature_drift
    ) != expected_feature_count:
        raise RuntimeError(
            "Feature drift row count does not match the number of shared features."
        )

    if feature_drift[
        "feature"
    ].duplicated().any():
        duplicated = feature_drift.loc[
            feature_drift[
                "feature"
            ].duplicated(
                keep=False
            ),
            "feature",
        ].tolist()

        raise RuntimeError(
            "Duplicate feature drift records were generated: "
            + ", ".join(
                map(
                    str,
                    duplicated[
                        :20
                    ],
                )
            )
        )

    invalid_severities = set(
        feature_drift[
            "overall_severity"
        ].dropna().unique()
    ) - set(
        SEVERITY_ORDER
    )

    if invalid_severities:
        raise RuntimeError(
            "Unexpected feature severities: "
            + ", ".join(
                sorted(
                    invalid_severities
                )
            )
        )

    bounded_columns = {
        "reference_missing_rate": (
            0.0,
            1.0,
        ),
        "current_missing_rate": (
            0.0,
            1.0,
        ),
        "missing_rate_difference": (
            0.0,
            1.0,
        ),
        "ks_statistic": (
            0.0,
            1.0,
        ),
        "ks_p_value": (
            0.0,
            1.0,
        ),
        "feature_health_score": (
            0.0,
            100.0,
        ),
    }

    for column, (
        lower,
        upper,
    ) in bounded_columns.items():
        numeric = pd.to_numeric(
            feature_drift[
                column
            ],
            errors="coerce",
        ).dropna()

        invalid = numeric.loc[
            (
                numeric
                < lower
            )
            | (
                numeric
                > upper
            )
        ]

        if not invalid.empty:
            raise RuntimeError(
                f"Column {column} contains values outside [{lower}, {upper}]."
            )

    psi_values = pd.to_numeric(
        feature_drift[
            "psi"
        ],
        errors="coerce",
    ).dropna()

    if (
        psi_values
        < 0
    ).any():
        raise RuntimeError(
            "PSI contains negative values."
        )

    wasserstein_values = pd.to_numeric(
        feature_drift[
            "wasserstein_distance"
        ],
        errors="coerce",
    ).dropna()

    if (
        wasserstein_values
        < 0
    ).any():
        raise RuntimeError(
            "Wasserstein distance contains negative values."
        )


def run_feature_drift_engine(
    context: MonitoringContext,
    *,
    save_output: bool = True,
) -> tuple[
    pd.DataFrame,
    FeatureDriftSummary,
    pd.DataFrame,
]:
    feature_drift = build_feature_drift_table(
        context
    )

    validate_feature_drift_output(
        feature_drift,
        expected_feature_count=len(
            context.shared_features
        ),
    )

    summary = summarise_feature_drift(
        feature_drift
    )

    unmatched_features = build_unmatched_feature_records(
        context
    )

    if save_output:
        ensure_directories()

        feature_drift.to_csv(
            FEATURE_DRIFT_PATH,
            index=False,
        )

        unmatched_path = (
            TABLES_DIR
            / "unmatched_features.csv"
        )

        unmatched_features.to_csv(
            unmatched_path,
            index=False,
        )

        atomic_write_json(
            TABLES_DIR
            / "feature_drift_summary.json",
            {
                "run_id": context.run_identity.run_id,
                "generated_utc": context.run_identity.generated_utc,
                "summary": asdict(
                    summary
                ),
                "feature_drift_path": str(
                    FEATURE_DRIFT_PATH
                ),
                "unmatched_features_path": str(
                    unmatched_path
                ),
            },
        )

    return (
        feature_drift,
        summary,
        unmatched_features,
    )


def build_feature_drift_console_summary(
    summary: FeatureDriftSummary,
) -> str:
    lines = [
        "="
        * 118,
        "HEVEMIND FEATURE DRIFT ENGINE",
        "="
        * 118,
        "",
        f"Monitored features:              {summary.monitored_features}",
        f"Normal features:                 {summary.normal_features}",
        f"Watch features:                  {summary.watch_features}",
        f"Warning features:                {summary.warning_features}",
        f"Critical features:               {summary.critical_features}",
        f"Features with detected drift:    {summary.drifted_features}",
        f"Drift rate:                      {summary.drift_rate:.4f}",
        f"Mean feature health score:       {summary.mean_feature_health_score:.2f}",
        f"Dominant severity:               {summary.dominant_severity}",
        f"Maximum PSI:                     {summary.maximum_psi}",
        f"Maximum KS statistic:            {summary.maximum_ks_statistic}",
        (
            "Maximum standardised shift:      "
            f"{summary.maximum_standardised_mean_shift}"
        ),
        (
            "Maximum missing-rate difference: "
            f"{summary.maximum_missing_rate_difference}"
        ),
        "",
        f"Saved feature table:             {FEATURE_DRIFT_PATH}",
    ]

    return "\n".join(
        lines
    )


# End of Part 2.
# Append Part 3 immediately below this line.


# ============================================================
# PART 3 — PREDICTION, DECISION, CONFIDENCE, UNCERTAINTY
#          AND DATA-QUALITY MONITORING
# ============================================================

@dataclass(frozen=True)
class DistributionDriftMetric:
    metric_name: str
    reference_count: int
    current_count: int
    reference_mean: float | None
    current_mean: float | None
    absolute_mean_shift: float | None
    reference_std: float | None
    current_std: float | None
    psi: float | None
    ks_statistic: float | None
    ks_p_value: float | None
    wasserstein_distance: float | None
    severity: str
    health_score: float
    notes: str


@dataclass(frozen=True)
class DecisionDriftMetric:
    reference_count: int
    current_count: int
    reference_distribution: dict[str, float]
    current_distribution: dict[str, float]
    total_variation_distance: float
    largest_category_shift: float
    largest_shift_category: str | None
    unseen_current_categories: tuple[str, ...]
    missing_reference_categories: tuple[str, ...]
    severity: str
    health_score: float
    notes: str


@dataclass(frozen=True)
class DataQualitySummary:
    reference_rows: int
    current_rows: int
    current_duplicate_rows: int
    current_duplicate_rate: float
    reference_total_missing_rate: float
    current_total_missing_rate: float
    current_constant_features: int
    current_constant_feature_rate: float
    reference_only_features: int
    current_only_features: int
    unmatched_feature_rate: float
    id_duplicate_count: int | None
    id_duplicate_rate: float | None
    invalid_probability_count: int | None
    invalid_confidence_count: int | None
    invalid_uncertainty_count: int | None
    severity: str
    health_score: float


@dataclass(frozen=True)
class PredictionMonitoringSummary:
    monitored_numeric_outputs: int
    numeric_output_severity: str
    decision_severity: str
    data_quality_severity: str
    mean_numeric_health_score: float
    decision_health_score: float | None
    data_quality_health_score: float
    dominant_severity: str


def series_missing_rate(
    dataframe: pd.DataFrame,
    column: str | None,
) -> float | None:
    if column is None or column not in dataframe.columns:
        return None

    return float(
        dataframe[
            column
        ].isna().mean()
    )


def validate_probability_like_values(
    dataframe: pd.DataFrame,
    column: str | None,
) -> tuple[
    int | None,
    float | None,
]:
    if column is None or column not in dataframe.columns:
        return None, None

    values = pd.to_numeric(
        dataframe[
            column
        ],
        errors="coerce",
    )

    finite_mask = values.notna()

    if not finite_mask.any():
        return 0, 0.0

    invalid_mask = (
        finite_mask
        & (
            (
                values
                < 0.0
            )
            | (
                values
                > 1.0
            )
        )
    )

    invalid_count = int(
        invalid_mask.sum()
    )

    invalid_rate = float(
        invalid_count
        / finite_mask.sum()
    )

    return (
        invalid_count,
        invalid_rate,
    )


def align_operational_columns(
    reference_mapping: ColumnMapping,
    current_mapping: ColumnMapping,
) -> dict[str, tuple[str | None, str | None]]:
    return {
        "failure_probability": (
            reference_mapping.probability_column,
            current_mapping.probability_column,
        ),
        "prediction_confidence": (
            reference_mapping.confidence_column,
            current_mapping.confidence_column,
        ),
        "combined_uncertainty": (
            reference_mapping.uncertainty_column,
            current_mapping.uncertainty_column,
        ),
        "data_confidence": (
            reference_mapping.data_confidence_column,
            current_mapping.data_confidence_column,
        ),
    }


def operational_threshold_key(
    metric_name: str,
) -> str:
    mapping = {
        "failure_probability": "probability_mean_shift",
        "prediction_confidence": "confidence_mean_shift",
        "combined_uncertainty": "uncertainty_mean_shift",
        "data_confidence": "confidence_mean_shift",
    }

    return mapping.get(
        metric_name,
        "probability_mean_shift",
    )


def build_distribution_notes(
    *,
    metric_name: str,
    reference_count: int,
    current_count: int,
    reference_mean: float | None,
    current_mean: float | None,
) -> str:
    notes: list[str] = []

    if reference_count == 0:
        notes.append(
            "Reference output is unavailable."
        )

    if current_count == 0:
        notes.append(
            "Current output is unavailable."
        )

    if (
        reference_mean is not None
        and current_mean is not None
    ):
        direction = (
            "increased"
            if current_mean
            > reference_mean
            else "decreased"
            if current_mean
            < reference_mean
            else "did not change"
        )

        notes.append(
            f"{metric_name.replace('_', ' ').title()} mean {direction}."
        )

    return " ".join(
        notes
    ) or "No additional monitoring note."


def evaluate_distribution_drift(
    *,
    metric_name: str,
    reference_series: pd.Series,
    current_series: pd.Series,
    config: Mapping[str, Any],
) -> DistributionDriftMetric:
    reference_values = finite_numeric_array(
        reference_series
    )

    current_values = finite_numeric_array(
        current_series
    )

    reference_mean = safe_mean(
        reference_values
    )

    current_mean = safe_mean(
        current_values
    )

    reference_std = safe_standard_deviation(
        reference_values
    )

    current_std = safe_standard_deviation(
        current_values
    )

    mean_shift = absolute_difference(
        reference_mean,
        current_mean,
    )

    epsilon = safe_float(
        config.get(
            "epsilon"
        ),
        1e-8,
    ) or 1e-8

    number_of_bins = int(
        config.get(
            "number_of_bins",
            10,
        )
    )

    psi_value = population_stability_index(
        reference_values,
        current_values,
        number_of_bins,
        epsilon,
    )

    ks_statistic, ks_p_value = kolmogorov_smirnov_two_sample(
        reference_values,
        current_values,
    )

    wasserstein = one_dimensional_wasserstein_distance(
        reference_values,
        current_values,
    )

    threshold_group = config.get(
        "prediction_drift_thresholds",
        {},
    )

    threshold_name = operational_threshold_key(
        metric_name
    )

    mean_severity = severity_from_thresholds(
        mean_shift,
        threshold_group.get(
            threshold_name,
            {},
        ),
    )

    psi_severity = severity_from_thresholds(
        psi_value,
        config.get(
            "feature_drift_thresholds",
            {},
        ).get(
            "psi",
            {},
        ),
    )

    ks_severity = severity_from_thresholds(
        ks_statistic,
        config.get(
            "feature_drift_thresholds",
            {},
        ).get(
            "ks_statistic",
            {},
        ),
    )

    availability_severity = "Normal"

    if reference_values.size == 0:
        availability_severity = "Warning"

    if current_values.size == 0:
        availability_severity = "Critical"

    metric_severities = {
        "mean_shift": mean_severity,
        "psi": psi_severity,
        "ks_statistic": ks_severity,
        "availability": availability_severity,
    }

    overall_severity = maximum_severity(
        metric_severities.values()
    )

    health_score = weighted_health_from_severities(
        metric_severities,
        {
            "mean_shift": 0.35,
            "psi": 0.30,
            "ks_statistic": 0.25,
            "availability": 0.10,
        },
    )

    notes = build_distribution_notes(
        metric_name=metric_name,
        reference_count=int(
            reference_values.size
        ),
        current_count=int(
            current_values.size
        ),
        reference_mean=reference_mean,
        current_mean=current_mean,
    )

    return DistributionDriftMetric(
        metric_name=metric_name,
        reference_count=int(
            reference_values.size
        ),
        current_count=int(
            current_values.size
        ),
        reference_mean=reference_mean,
        current_mean=current_mean,
        absolute_mean_shift=mean_shift,
        reference_std=reference_std,
        current_std=current_std,
        psi=psi_value,
        ks_statistic=ks_statistic,
        ks_p_value=ks_p_value,
        wasserstein_distance=wasserstein,
        severity=overall_severity,
        health_score=health_score,
        notes=notes,
    )


def build_prediction_drift_table(
    context: MonitoringContext,
) -> pd.DataFrame:
    reference_mapping = (
        context.reference.descriptor.column_mapping
    )

    current_mapping = (
        context.current.descriptor.column_mapping
    )

    aligned = align_operational_columns(
        reference_mapping,
        current_mapping,
    )

    records: list[dict[str, Any]] = []

    for metric_name, (
        reference_column,
        current_column,
    ) in aligned.items():
        if (
            reference_column is None
            or current_column is None
        ):
            records.append(
                {
                    "metric_name": metric_name,
                    "reference_column": reference_column,
                    "current_column": current_column,
                    "reference_count": 0,
                    "current_count": 0,
                    "reference_mean": None,
                    "current_mean": None,
                    "absolute_mean_shift": None,
                    "reference_std": None,
                    "current_std": None,
                    "psi": None,
                    "ks_statistic": None,
                    "ks_p_value": None,
                    "wasserstein_distance": None,
                    "severity": "Warning",
                    "health_score": 70.0,
                    "notes": (
                        "The operational metric is unavailable in one or both datasets."
                    ),
                }
            )
            continue

        metric = evaluate_distribution_drift(
            metric_name=metric_name,
            reference_series=context.reference.dataframe[
                reference_column
            ],
            current_series=context.current.dataframe[
                current_column
            ],
            config=context.config,
        )

        record = asdict(
            metric
        )

        record[
            "reference_column"
        ] = reference_column

        record[
            "current_column"
        ] = current_column

        records.append(
            record
        )

    dataframe = pd.DataFrame(
        records
    )

    if dataframe.empty:
        return pd.DataFrame(
            columns=[
                "metric_name",
                "reference_column",
                "current_column",
                "reference_count",
                "current_count",
                "reference_mean",
                "current_mean",
                "absolute_mean_shift",
                "reference_std",
                "current_std",
                "psi",
                "ks_statistic",
                "ks_p_value",
                "wasserstein_distance",
                "severity",
                "health_score",
                "notes",
            ]
        )

    dataframe[
        "severity_rank"
    ] = dataframe[
        "severity"
    ].map(
        SEVERITY_ORDER
    ).fillna(
        0
    ).astype(
        int
    )

    dataframe = dataframe.sort_values(
        by=[
            "severity_rank",
            "health_score",
        ],
        ascending=[
            False,
            True,
        ],
    ).reset_index(
        drop=True
    )

    return dataframe


def normalise_category_series(
    values: pd.Series,
) -> pd.Series:
    output = values.astype(
        "string"
    )

    output = output.fillna(
        "Missing"
    )

    output = output.str.strip()

    output = output.replace(
        {
            "": "Missing",
            "<NA>": "Missing",
            "nan": "Missing",
            "None": "Missing",
        }
    )

    return output


def categorical_distribution(
    values: pd.Series,
    categories: Sequence[str],
    epsilon: float,
) -> dict[str, float]:
    normalised = normalise_category_series(
        values
    )

    counts = normalised.value_counts(
        dropna=False
    )

    total = float(
        counts.sum()
    )

    if total <= 0:
        return {
            category: 0.0
            for category in categories
        }

    proportions = {
        category: float(
            counts.get(
                category,
                0
            )
            / total
        )
        for category in categories
    }

    clipped = {
        category: max(
            proportion,
            epsilon,
        )
        for category, proportion in proportions.items()
    }

    normalisation = sum(
        clipped.values()
    )

    return {
        category: proportion
        / normalisation
        for category, proportion in clipped.items()
    }


def total_variation_distance(
    first: Mapping[str, float],
    second: Mapping[str, float],
) -> float:
    categories = sorted(
        set(
            first
        )
        | set(
            second
        )
    )

    return 0.5 * sum(
        abs(
            float(
                first.get(
                    category,
                    0.0,
                )
            )
            - float(
                second.get(
                    category,
                    0.0,
                )
            )
        )
        for category in categories
    )


def evaluate_decision_drift(
    context: MonitoringContext,
) -> DecisionDriftMetric | None:
    reference_column = (
        context.reference.descriptor.column_mapping.decision_column
    )

    current_column = (
        context.current.descriptor.column_mapping.decision_column
    )

    if (
        reference_column is None
        or current_column is None
    ):
        return None

    reference_series = normalise_category_series(
        context.reference.dataframe[
            reference_column
        ]
    )

    current_series = normalise_category_series(
        context.current.dataframe[
            current_column
        ]
    )

    reference_categories = set(
        reference_series.unique().tolist()
    )

    current_categories = set(
        current_series.unique().tolist()
    )

    categories = sorted(
        reference_categories
        | current_categories
    )

    epsilon = safe_float(
        context.config.get(
            "epsilon"
        ),
        1e-8,
    ) or 1e-8

    reference_distribution = categorical_distribution(
        reference_series,
        categories,
        epsilon,
    )

    current_distribution = categorical_distribution(
        current_series,
        categories,
        epsilon,
    )

    total_variation = total_variation_distance(
        reference_distribution,
        current_distribution,
    )

    category_shifts = {
        category: abs(
            current_distribution.get(
                category,
                0.0,
            )
            - reference_distribution.get(
                category,
                0.0,
            )
        )
        for category in categories
    }

    largest_category = (
        max(
            category_shifts,
            key=category_shifts.get,
        )
        if category_shifts
        else None
    )

    largest_shift = (
        category_shifts[
            largest_category
        ]
        if largest_category is not None
        else 0.0
    )

    unseen = tuple(
        sorted(
            current_categories
            - reference_categories
        )
    )

    missing = tuple(
        sorted(
            reference_categories
            - current_categories
        )
    )

    thresholds = context.config.get(
        "prediction_drift_thresholds",
        {},
    ).get(
        "decision_total_variation",
        {},
    )

    distribution_severity = severity_from_thresholds(
        total_variation,
        thresholds,
    )

    category_severity = "Normal"

    if unseen:
        category_severity = "Warning"

    if missing:
        category_severity = maximum_severity(
            [
                category_severity,
                "Watch",
            ]
        )

    severity = maximum_severity(
        [
            distribution_severity,
            category_severity,
        ]
    )

    health_score = weighted_health_from_severities(
        {
            "distribution_shift": distribution_severity,
            "category_structure": category_severity,
        },
        {
            "distribution_shift": 0.80,
            "category_structure": 0.20,
        },
    )

    notes: list[str] = []

    if largest_category is not None:
        notes.append(
            f"Largest decision share shift: {largest_category} ({largest_shift:.4f})."
        )

    if unseen:
        notes.append(
            "New current categories: "
            + ", ".join(
                unseen
            )
            + "."
        )

    if missing:
        notes.append(
            "Reference categories absent from current data: "
            + ", ".join(
                missing
            )
            + "."
        )

    return DecisionDriftMetric(
        reference_count=int(
            len(
                reference_series
            )
        ),
        current_count=int(
            len(
                current_series
            )
        ),
        reference_distribution=reference_distribution,
        current_distribution=current_distribution,
        total_variation_distance=total_variation,
        largest_category_shift=largest_shift,
        largest_shift_category=largest_category,
        unseen_current_categories=unseen,
        missing_reference_categories=missing,
        severity=severity,
        health_score=health_score,
        notes=" ".join(
            notes
        ) or "No material decision-category issue detected.",
    )


def decision_drift_to_dataframe(
    metric: DecisionDriftMetric | None,
) -> pd.DataFrame:
    if metric is None:
        return pd.DataFrame(
            [
                {
                    "category": "Unavailable",
                    "reference_share": None,
                    "current_share": None,
                    "absolute_shift": None,
                    "total_variation_distance": None,
                    "severity": "Warning",
                    "health_score": 70.0,
                    "notes": (
                        "Decision column is unavailable in one or both datasets."
                    ),
                }
            ]
        )

    categories = sorted(
        set(
            metric.reference_distribution
        )
        | set(
            metric.current_distribution
        )
    )

    records = []

    for category in categories:
        reference_share = metric.reference_distribution.get(
            category,
            0.0,
        )

        current_share = metric.current_distribution.get(
            category,
            0.0,
        )

        records.append(
            {
                "category": category,
                "reference_share": reference_share,
                "current_share": current_share,
                "absolute_shift": abs(
                    current_share
                    - reference_share
                ),
                "total_variation_distance": (
                    metric.total_variation_distance
                ),
                "severity": metric.severity,
                "health_score": metric.health_score,
                "notes": metric.notes,
            }
        )

    return pd.DataFrame(
        records
    ).sort_values(
        "absolute_shift",
        ascending=False,
    ).reset_index(
        drop=True
    )


def total_missing_rate(
    dataframe: pd.DataFrame,
    columns: Sequence[str],
) -> float:
    if not columns:
        return 0.0

    selected = dataframe.loc[
        :,
        list(
            columns
        ),
    ]

    denominator = (
        selected.shape[
            0
        ]
        * selected.shape[
            1
        ]
    )

    if denominator <= 0:
        return 0.0

    return float(
        selected.isna().sum().sum()
        / denominator
    )


def constant_feature_count(
    dataframe: pd.DataFrame,
    features: Sequence[str],
    epsilon: float,
) -> int:
    count = 0

    for feature in features:
        values = finite_numeric_array(
            dataframe[
                feature
            ]
        )

        if is_effectively_constant(
            values,
            epsilon,
        ):
            count += 1

    return count


def data_quality_severity(
    metrics: Mapping[str, float | None],
    config: Mapping[str, Any],
) -> tuple[
    str,
    dict[str, str],
]:
    thresholds = config.get(
        "data_quality_thresholds",
        {},
    )

    severities = {
        "missing_rate": severity_from_thresholds(
            metrics.get(
                "current_total_missing_rate"
            ),
            thresholds.get(
                "missing_rate",
                {},
            ),
        ),
        "duplicate_rate": severity_from_thresholds(
            metrics.get(
                "current_duplicate_rate"
            ),
            thresholds.get(
                "duplicate_rate",
                {},
            ),
        ),
        "constant_feature_rate": severity_from_thresholds(
            metrics.get(
                "current_constant_feature_rate"
            ),
            thresholds.get(
                "constant_feature_rate",
                {},
            ),
        ),
        "unmatched_feature_rate": severity_from_thresholds(
            metrics.get(
                "unmatched_feature_rate"
            ),
            thresholds.get(
                "unmatched_feature_rate",
                {},
            ),
        ),
    }

    for validation_name in [
        "invalid_probability_rate",
        "invalid_confidence_rate",
        "invalid_uncertainty_rate",
        "id_duplicate_rate",
    ]:
        value = metrics.get(
            validation_name
        )

        if value is None:
            severities[
                validation_name
            ] = "Normal"

        elif value >= 0.10:
            severities[
                validation_name
            ] = "Critical"

        elif value >= 0.03:
            severities[
                validation_name
            ] = "Warning"

        elif value > 0:
            severities[
                validation_name
            ] = "Watch"

        else:
            severities[
                validation_name
            ] = "Normal"

    return (
        maximum_severity(
            severities.values()
        ),
        severities,
    )


def evaluate_data_quality(
    context: MonitoringContext,
) -> tuple[
    DataQualitySummary,
    pd.DataFrame,
]:
    epsilon = safe_float(
        context.config.get(
            "epsilon"
        ),
        1e-8,
    ) or 1e-8

    reference_dataframe = (
        context.reference.dataframe
    )

    current_dataframe = (
        context.current.dataframe
    )

    shared_features = context.shared_features

    reference_missing = total_missing_rate(
        reference_dataframe,
        shared_features,
    )

    current_missing = total_missing_rate(
        current_dataframe,
        shared_features,
    )

    current_duplicate_rows = int(
        current_dataframe.duplicated().sum()
    )

    current_duplicate_rate = float(
        current_duplicate_rows
        / max(
            len(
                current_dataframe
            ),
            1,
        )
    )

    constant_count = constant_feature_count(
        current_dataframe,
        shared_features,
        epsilon,
    )

    constant_rate = float(
        constant_count
        / max(
            len(
                shared_features
            ),
            1,
        )
    )

    unmatched_count = (
        len(
            context.reference_only_features
        )
        + len(
            context.current_only_features
        )
    )

    total_distinct_features = len(
        set(
            context.reference.descriptor.feature_columns
        )
        | set(
            context.current.descriptor.feature_columns
        )
    )

    unmatched_rate = float(
        unmatched_count
        / max(
            total_distinct_features,
            1,
        )
    )

    current_mapping = (
        context.current.descriptor.column_mapping
    )

    id_duplicate_count: int | None = None
    id_duplicate_rate: float | None = None

    if (
        current_mapping.id_column is not None
        and current_mapping.id_column
        in current_dataframe.columns
    ):
        id_series = current_dataframe[
            current_mapping.id_column
        ]

        non_missing_ids = id_series.dropna()

        id_duplicate_count = int(
            non_missing_ids.duplicated().sum()
        )

        id_duplicate_rate = float(
            id_duplicate_count
            / max(
                len(
                    non_missing_ids
                ),
                1,
            )
        )

    invalid_probability_count, invalid_probability_rate = (
        validate_probability_like_values(
            current_dataframe,
            current_mapping.probability_column,
        )
    )

    invalid_confidence_count, invalid_confidence_rate = (
        validate_probability_like_values(
            current_dataframe,
            current_mapping.confidence_column,
        )
    )

    invalid_uncertainty_count, invalid_uncertainty_rate = (
        validate_probability_like_values(
            current_dataframe,
            current_mapping.uncertainty_column,
        )
    )

    metric_values: dict[str, float | None] = {
        "current_total_missing_rate": current_missing,
        "current_duplicate_rate": current_duplicate_rate,
        "current_constant_feature_rate": constant_rate,
        "unmatched_feature_rate": unmatched_rate,
        "id_duplicate_rate": id_duplicate_rate,
        "invalid_probability_rate": invalid_probability_rate,
        "invalid_confidence_rate": invalid_confidence_rate,
        "invalid_uncertainty_rate": invalid_uncertainty_rate,
    }

    overall_severity, severities = data_quality_severity(
        metric_values,
        context.config,
    )

    health_score = weighted_health_from_severities(
        severities,
        {
            "missing_rate": 0.30,
            "duplicate_rate": 0.15,
            "constant_feature_rate": 0.20,
            "unmatched_feature_rate": 0.15,
            "id_duplicate_rate": 0.05,
            "invalid_probability_rate": 0.05,
            "invalid_confidence_rate": 0.05,
            "invalid_uncertainty_rate": 0.05,
        },
    )

    summary = DataQualitySummary(
        reference_rows=int(
            len(
                reference_dataframe
            )
        ),
        current_rows=int(
            len(
                current_dataframe
            )
        ),
        current_duplicate_rows=current_duplicate_rows,
        current_duplicate_rate=current_duplicate_rate,
        reference_total_missing_rate=reference_missing,
        current_total_missing_rate=current_missing,
        current_constant_features=constant_count,
        current_constant_feature_rate=constant_rate,
        reference_only_features=len(
            context.reference_only_features
        ),
        current_only_features=len(
            context.current_only_features
        ),
        unmatched_feature_rate=unmatched_rate,
        id_duplicate_count=id_duplicate_count,
        id_duplicate_rate=id_duplicate_rate,
        invalid_probability_count=invalid_probability_count,
        invalid_confidence_count=invalid_confidence_count,
        invalid_uncertainty_count=invalid_uncertainty_count,
        severity=overall_severity,
        health_score=health_score,
    )

    rows = [
        {
            "metric": "Current total missing rate",
            "value": current_missing,
            "reference_value": reference_missing,
            "severity": severities[
                "missing_rate"
            ],
        },
        {
            "metric": "Current duplicate-row rate",
            "value": current_duplicate_rate,
            "reference_value": None,
            "severity": severities[
                "duplicate_rate"
            ],
        },
        {
            "metric": "Current constant-feature rate",
            "value": constant_rate,
            "reference_value": None,
            "severity": severities[
                "constant_feature_rate"
            ],
        },
        {
            "metric": "Unmatched-feature rate",
            "value": unmatched_rate,
            "reference_value": None,
            "severity": severities[
                "unmatched_feature_rate"
            ],
        },
        {
            "metric": "Duplicate ID rate",
            "value": id_duplicate_rate,
            "reference_value": None,
            "severity": severities[
                "id_duplicate_rate"
            ],
        },
        {
            "metric": "Invalid probability rate",
            "value": invalid_probability_rate,
            "reference_value": None,
            "severity": severities[
                "invalid_probability_rate"
            ],
        },
        {
            "metric": "Invalid confidence rate",
            "value": invalid_confidence_rate,
            "reference_value": None,
            "severity": severities[
                "invalid_confidence_rate"
            ],
        },
        {
            "metric": "Invalid uncertainty rate",
            "value": invalid_uncertainty_rate,
            "reference_value": None,
            "severity": severities[
                "invalid_uncertainty_rate"
            ],
        },
    ]

    table = pd.DataFrame(
        rows
    )

    table[
        "severity_rank"
    ] = table[
        "severity"
    ].map(
        SEVERITY_ORDER
    ).fillna(
        0
    ).astype(
        int
    )

    table = table.sort_values(
        by=[
            "severity_rank",
            "metric",
        ],
        ascending=[
            False,
            True,
        ],
    ).reset_index(
        drop=True
    )

    return summary, table


def validate_prediction_drift_table(
    dataframe: pd.DataFrame,
) -> None:
    if dataframe.empty:
        raise RuntimeError(
            "Prediction drift table is empty."
        )

    invalid_severities = set(
        dataframe[
            "severity"
        ].dropna().unique()
    ) - set(
        SEVERITY_ORDER
    )

    if invalid_severities:
        raise RuntimeError(
            "Unexpected prediction-drift severities: "
            + ", ".join(
                sorted(
                    invalid_severities
                )
            )
        )

    bounded_columns = {
        "ks_statistic": (
            0.0,
            1.0,
        ),
        "ks_p_value": (
            0.0,
            1.0,
        ),
        "health_score": (
            0.0,
            100.0,
        ),
    }

    for column, (
        lower,
        upper,
    ) in bounded_columns.items():
        numeric = pd.to_numeric(
            dataframe[
                column
            ],
            errors="coerce",
        ).dropna()

        if (
            (
                numeric
                < lower
            )
            | (
                numeric
                > upper
            )
        ).any():
            raise RuntimeError(
                f"{column} contains values outside [{lower}, {upper}]."
            )

    psi_values = pd.to_numeric(
        dataframe[
            "psi"
        ],
        errors="coerce",
    ).dropna()

    if (
        psi_values
        < 0
    ).any():
        raise RuntimeError(
            "Prediction PSI contains negative values."
        )


def summarise_prediction_monitoring(
    prediction_drift: pd.DataFrame,
    decision_metric: DecisionDriftMetric | None,
    data_quality_summary: DataQualitySummary,
) -> PredictionMonitoringSummary:
    numeric_health = pd.to_numeric(
        prediction_drift[
            "health_score"
        ],
        errors="coerce",
    )

    mean_numeric_health = (
        float(
            numeric_health.mean()
        )
        if numeric_health.notna().any()
        else 100.0
    )

    numeric_severity = maximum_severity(
        prediction_drift[
            "severity"
        ].tolist()
    )

    decision_severity = (
        decision_metric.severity
        if decision_metric is not None
        else "Warning"
    )

    decision_health = (
        decision_metric.health_score
        if decision_metric is not None
        else None
    )

    dominant = maximum_severity(
        [
            numeric_severity,
            decision_severity,
            data_quality_summary.severity,
        ]
    )

    return PredictionMonitoringSummary(
        monitored_numeric_outputs=int(
            len(
                prediction_drift
            )
        ),
        numeric_output_severity=numeric_severity,
        decision_severity=decision_severity,
        data_quality_severity=(
            data_quality_summary.severity
        ),
        mean_numeric_health_score=round(
            mean_numeric_health,
            2,
        ),
        decision_health_score=decision_health,
        data_quality_health_score=(
            data_quality_summary.health_score
        ),
        dominant_severity=dominant,
    )


def run_prediction_and_quality_engine(
    context: MonitoringContext,
    *,
    save_output: bool = True,
) -> tuple[
    pd.DataFrame,
    DecisionDriftMetric | None,
    pd.DataFrame,
    DataQualitySummary,
    pd.DataFrame,
    PredictionMonitoringSummary,
]:
    prediction_drift = build_prediction_drift_table(
        context
    )

    validate_prediction_drift_table(
        prediction_drift
    )

    decision_metric = evaluate_decision_drift(
        context
    )

    decision_table = decision_drift_to_dataframe(
        decision_metric
    )

    data_quality_summary, data_quality_table = evaluate_data_quality(
        context
    )

    monitoring_summary = summarise_prediction_monitoring(
        prediction_drift,
        decision_metric,
        data_quality_summary,
    )

    if save_output:
        ensure_directories()

        prediction_drift.to_csv(
            PREDICTION_DRIFT_PATH,
            index=False,
        )

        decision_table.to_csv(
            TABLES_DIR
            / "decision_drift.csv",
            index=False,
        )

        data_quality_table.to_csv(
            DATA_QUALITY_PATH,
            index=False,
        )

        atomic_write_json(
            TABLES_DIR
            / "prediction_monitoring_summary.json",
            {
                "run_id": context.run_identity.run_id,
                "generated_utc": context.run_identity.generated_utc,
                "summary": asdict(
                    monitoring_summary
                ),
                "decision_metric": (
                    asdict(
                        decision_metric
                    )
                    if decision_metric is not None
                    else None
                ),
                "data_quality_summary": asdict(
                    data_quality_summary
                ),
                "prediction_drift_path": str(
                    PREDICTION_DRIFT_PATH
                ),
                "decision_drift_path": str(
                    TABLES_DIR
                    / "decision_drift.csv"
                ),
                "data_quality_path": str(
                    DATA_QUALITY_PATH
                ),
            },
        )

    return (
        prediction_drift,
        decision_metric,
        decision_table,
        data_quality_summary,
        data_quality_table,
        monitoring_summary,
    )


def build_prediction_monitoring_console_summary(
    summary: PredictionMonitoringSummary,
    decision_metric: DecisionDriftMetric | None,
    data_quality_summary: DataQualitySummary,
) -> str:
    lines = [
        "="
        * 118,
        "HEVEMIND PREDICTION, DECISION AND DATA-QUALITY MONITORING",
        "="
        * 118,
        "",
        (
            "Numeric operational outputs:      "
            f"{summary.monitored_numeric_outputs}"
        ),
        (
            "Numeric-output severity:          "
            f"{summary.numeric_output_severity}"
        ),
        (
            "Mean numeric health score:        "
            f"{summary.mean_numeric_health_score:.2f}"
        ),
        (
            "Decision severity:                "
            f"{summary.decision_severity}"
        ),
        (
            "Decision total variation:         "
            f"{decision_metric.total_variation_distance if decision_metric else None}"
        ),
        (
            "Data-quality severity:            "
            f"{summary.data_quality_severity}"
        ),
        (
            "Data-quality health score:        "
            f"{summary.data_quality_health_score:.2f}"
        ),
        (
            "Current missing rate:             "
            f"{data_quality_summary.current_total_missing_rate:.4f}"
        ),
        (
            "Current duplicate-row rate:       "
            f"{data_quality_summary.current_duplicate_rate:.4f}"
        ),
        (
            "Current constant-feature rate:    "
            f"{data_quality_summary.current_constant_feature_rate:.4f}"
        ),
        (
            "Unmatched-feature rate:           "
            f"{data_quality_summary.unmatched_feature_rate:.4f}"
        ),
        (
            "Dominant monitoring severity:     "
            f"{summary.dominant_severity}"
        ),
        "",
        f"Prediction drift table:             {PREDICTION_DRIFT_PATH}",
        f"Decision drift table:               {TABLES_DIR / 'decision_drift.csv'}",
        f"Data quality table:                 {DATA_QUALITY_PATH}",
    ]

    return "\n".join(
        lines
    )


# End of Part 3.
# Append Part 4 immediately below this line.


# ============================================================
# PART 4 — DEPLOYMENT HEALTH, ALERTING AND RETRAINING POLICY
# ============================================================

@dataclass(frozen=True)
class MonitoringAlert:
    alert_id: str
    run_id: str
    generated_utc: str
    category: str
    alert_type: str
    severity: str
    resource_type: str
    resource_id: str | None
    metric_name: str
    observed_value: float | None
    threshold_value: float | None
    message: str
    recommended_action: str
    evidence: dict[str, Any]


@dataclass(frozen=True)
class HealthComponent:
    component_name: str
    weight: float
    raw_score: float
    weighted_score: float
    severity: str
    status_label: str
    monitored_items: int
    warning_items: int
    critical_items: int
    explanation: str


@dataclass(frozen=True)
class DeploymentHealthAssessment:
    run_id: str
    generated_utc: str
    overall_health_score: float
    overall_health_band: str
    overall_severity: str
    recommendation: str
    recommendation_reason: str
    production_monitoring_status: str
    component_scores: dict[str, float]
    component_weights: dict[str, float]
    alert_count: int
    watch_alerts: int
    warning_alerts: int
    critical_alerts: int
    retraining_triggered: bool
    immediate_action_required: bool
    model_modification_performed: bool
    primary_risk_driver: str
    summary_statement: str


@dataclass(frozen=True)
class MonitoringBundle:
    context: MonitoringContext
    feature_drift: pd.DataFrame
    feature_summary: FeatureDriftSummary
    unmatched_features: pd.DataFrame
    prediction_drift: pd.DataFrame
    decision_metric: DecisionDriftMetric | None
    decision_drift: pd.DataFrame
    data_quality_summary: DataQualitySummary
    data_quality_table: pd.DataFrame
    prediction_summary: PredictionMonitoringSummary
    alerts: pd.DataFrame
    health_components: pd.DataFrame
    health_assessment: DeploymentHealthAssessment


def deterministic_alert_id(
    *,
    run_id: str,
    category: str,
    alert_type: str,
    resource_id: str | None,
    metric_name: str,
) -> str:
    payload = "|".join(
        [
            run_id,
            category,
            alert_type,
            resource_id or "",
            metric_name,
        ]
    )

    return hashlib.sha256(
        payload.encode(
            "utf-8"
        )
    ).hexdigest()[
        :20
    ]


def alert_threshold_for_severity(
    thresholds: Mapping[str, Any],
    severity: str,
) -> float | None:
    key_mapping = {
        "Watch": "watch",
        "Warning": "warning",
        "Critical": "critical",
    }

    threshold_key = key_mapping.get(
        severity
    )

    if threshold_key is None:
        return None

    return safe_float(
        thresholds.get(
            threshold_key
        )
    )


def severity_action(
    severity: str,
    category: str,
) -> str:
    if severity == "Critical":
        if category == "Data Quality":
            return (
                "Quarantine affected records, verify data ingestion and "
                "require engineering review before relying on predictions."
            )

        if category == "Feature Drift":
            return (
                "Escalate process and sensor review, compare with maintenance "
                "and recipe changes, and assess retraining immediately."
            )

        if category == "Prediction Drift":
            return (
                "Review recent prediction distributions and actual outcomes; "
                "consider immediate model recalibration or retraining."
            )

        return (
            "Escalate immediately to the model owner and manufacturing "
            "engineering lead."
        )

    if severity == "Warning":
        return (
            "Increase monitoring frequency, investigate the affected evidence "
            "and prepare a controlled retraining assessment."
        )

    if severity == "Watch":
        return (
            "Continue deployment with enhanced monitoring and document the "
            "observed change for trend analysis."
        )

    return (
        "Continue standard monitoring."
    )


def feature_alert_message(
    row: Mapping[str, Any],
) -> str:
    feature = safe_text(
        row.get(
            "feature"
        )
    )

    severity = safe_text(
        row.get(
            "overall_severity"
        ),
        "Normal",
    )

    metric_fragments: list[str] = []

    for label, column in [
        (
            "PSI",
            "psi",
        ),
        (
            "KS",
            "ks_statistic",
        ),
        (
            "standardised mean shift",
            "standardised_mean_shift",
        ),
        (
            "missing-rate difference",
            "missing_rate_difference",
        ),
    ]:
        value = safe_float(
            row.get(
                column
            )
        )

        if value is not None:
            metric_fragments.append(
                f"{label}={value:.4f}"
            )

    metric_text = (
        ", ".join(
            metric_fragments
        )
        if metric_fragments
        else "limited metric evidence"
    )

    return (
        f"{feature} is classified as {severity} for distribution drift "
        f"({metric_text})."
    )


def feature_primary_alert_metric(
    row: Mapping[str, Any],
) -> tuple[
    str,
    float | None,
    str,
]:
    candidate_metrics = [
        (
            "psi",
            safe_float(
                row.get(
                    "psi"
                )
            ),
            safe_text(
                row.get(
                    "psi_severity"
                ),
                "Normal",
            ),
        ),
        (
            "ks_statistic",
            safe_float(
                row.get(
                    "ks_statistic"
                )
            ),
            safe_text(
                row.get(
                    "ks_statistic_severity"
                ),
                "Normal",
            ),
        ),
        (
            "standardised_mean_shift",
            safe_float(
                row.get(
                    "standardised_mean_shift"
                )
            ),
            safe_text(
                row.get(
                    "standardised_mean_shift_severity"
                ),
                "Normal",
            ),
        ),
        (
            "missing_rate_difference",
            safe_float(
                row.get(
                    "missing_rate_difference"
                )
            ),
            safe_text(
                row.get(
                    "missing_rate_difference_severity"
                ),
                "Normal",
            ),
        ),
        (
            "constant_sensor_change",
            None,
            safe_text(
                row.get(
                    "constant_sensor_change_severity"
                ),
                "Normal",
            ),
        ),
        (
            "measurement_availability",
            None,
            safe_text(
                row.get(
                    "measurement_availability_severity"
                ),
                "Normal",
            ),
        ),
    ]

    return max(
        candidate_metrics,
        key=lambda item: SEVERITY_ORDER.get(
            item[
                2
            ],
            0,
        ),
    )


def build_feature_alerts(
    context: MonitoringContext,
    feature_drift: pd.DataFrame,
) -> list[MonitoringAlert]:
    alerts: list[MonitoringAlert] = []

    thresholds = context.config.get(
        "feature_drift_thresholds",
        {},
    )

    for row in feature_drift.to_dict(
        orient="records"
    ):
        severity = safe_text(
            row.get(
                "overall_severity"
            ),
            "Normal",
        )

        if severity == "Normal":
            continue

        metric_name, observed_value, metric_severity = (
            feature_primary_alert_metric(
                row
            )
        )

        metric_thresholds = thresholds.get(
            metric_name,
            {},
        )

        threshold_value = alert_threshold_for_severity(
            metric_thresholds,
            metric_severity,
        )

        feature = safe_text(
            row.get(
                "feature"
            )
        )

        alert_id = deterministic_alert_id(
            run_id=context.run_identity.run_id,
            category="Feature Drift",
            alert_type="sensor_distribution_shift",
            resource_id=feature,
            metric_name=metric_name,
        )

        alerts.append(
            MonitoringAlert(
                alert_id=alert_id,
                run_id=context.run_identity.run_id,
                generated_utc=context.run_identity.generated_utc,
                category="Feature Drift",
                alert_type="sensor_distribution_shift",
                severity=severity,
                resource_type="sensor_feature",
                resource_id=feature,
                metric_name=metric_name,
                observed_value=observed_value,
                threshold_value=threshold_value,
                message=feature_alert_message(
                    row
                ),
                recommended_action=severity_action(
                    severity,
                    "Feature Drift",
                ),
                evidence={
                    "feature_health_score": row.get(
                        "feature_health_score"
                    ),
                    "psi": row.get(
                        "psi"
                    ),
                    "ks_statistic": row.get(
                        "ks_statistic"
                    ),
                    "ks_p_value": row.get(
                        "ks_p_value"
                    ),
                    "standardised_mean_shift": row.get(
                        "standardised_mean_shift"
                    ),
                    "missing_rate_difference": row.get(
                        "missing_rate_difference"
                    ),
                    "reference_constant": row.get(
                        "reference_constant"
                    ),
                    "current_constant": row.get(
                        "current_constant"
                    ),
                    "notes": row.get(
                        "notes"
                    ),
                },
            )
        )

    return alerts


def prediction_alert_thresholds(
    context: MonitoringContext,
    metric_name: str,
) -> Mapping[str, Any]:
    key = operational_threshold_key(
        metric_name
    )

    return (
        context.config.get(
            "prediction_drift_thresholds",
            {}
        ).get(
            key,
            {},
        )
    )


def build_prediction_alerts(
    context: MonitoringContext,
    prediction_drift: pd.DataFrame,
) -> list[MonitoringAlert]:
    alerts: list[MonitoringAlert] = []

    for row in prediction_drift.to_dict(
        orient="records"
    ):
        severity = safe_text(
            row.get(
                "severity"
            ),
            "Normal",
        )

        if severity == "Normal":
            continue

        metric_name = safe_text(
            row.get(
                "metric_name"
            )
        )

        observed_value = safe_float(
            row.get(
                "absolute_mean_shift"
            )
        )

        threshold_value = alert_threshold_for_severity(
            prediction_alert_thresholds(
                context,
                metric_name,
            ),
            severity,
        )

        message = (
            f"{metric_name.replace('_', ' ').title()} is classified as "
            f"{severity}; reference mean={row.get('reference_mean')}, "
            f"current mean={row.get('current_mean')}."
        )

        alerts.append(
            MonitoringAlert(
                alert_id=deterministic_alert_id(
                    run_id=context.run_identity.run_id,
                    category="Prediction Drift",
                    alert_type="operational_output_shift",
                    resource_id=metric_name,
                    metric_name=metric_name,
                ),
                run_id=context.run_identity.run_id,
                generated_utc=context.run_identity.generated_utc,
                category="Prediction Drift",
                alert_type="operational_output_shift",
                severity=severity,
                resource_type="model_output",
                resource_id=metric_name,
                metric_name=metric_name,
                observed_value=observed_value,
                threshold_value=threshold_value,
                message=message,
                recommended_action=severity_action(
                    severity,
                    "Prediction Drift",
                ),
                evidence={
                    "reference_count": row.get(
                        "reference_count"
                    ),
                    "current_count": row.get(
                        "current_count"
                    ),
                    "reference_mean": row.get(
                        "reference_mean"
                    ),
                    "current_mean": row.get(
                        "current_mean"
                    ),
                    "psi": row.get(
                        "psi"
                    ),
                    "ks_statistic": row.get(
                        "ks_statistic"
                    ),
                    "ks_p_value": row.get(
                        "ks_p_value"
                    ),
                    "wasserstein_distance": row.get(
                        "wasserstein_distance"
                    ),
                    "health_score": row.get(
                        "health_score"
                    ),
                    "notes": row.get(
                        "notes"
                    ),
                },
            )
        )

    return alerts


def build_decision_alert(
    context: MonitoringContext,
    decision_metric: DecisionDriftMetric | None,
) -> list[MonitoringAlert]:
    if (
        decision_metric is None
        or decision_metric.severity
        == "Normal"
    ):
        return []

    thresholds = (
        context.config.get(
            "prediction_drift_thresholds",
            {}
        ).get(
            "decision_total_variation",
            {},
        )
    )

    return [
        MonitoringAlert(
            alert_id=deterministic_alert_id(
                run_id=context.run_identity.run_id,
                category="Prediction Drift",
                alert_type="decision_distribution_shift",
                resource_id="operational_decision",
                metric_name="decision_total_variation",
            ),
            run_id=context.run_identity.run_id,
            generated_utc=context.run_identity.generated_utc,
            category="Prediction Drift",
            alert_type="decision_distribution_shift",
            severity=decision_metric.severity,
            resource_type="decision_distribution",
            resource_id="operational_decision",
            metric_name="decision_total_variation",
            observed_value=(
                decision_metric.total_variation_distance
            ),
            threshold_value=alert_threshold_for_severity(
                thresholds,
                decision_metric.severity,
            ),
            message=(
                "Operational decision distribution shifted with total "
                f"variation distance "
                f"{decision_metric.total_variation_distance:.4f}."
            ),
            recommended_action=severity_action(
                decision_metric.severity,
                "Prediction Drift",
            ),
            evidence={
                "reference_distribution": (
                    decision_metric.reference_distribution
                ),
                "current_distribution": (
                    decision_metric.current_distribution
                ),
                "largest_category_shift": (
                    decision_metric.largest_category_shift
                ),
                "largest_shift_category": (
                    decision_metric.largest_shift_category
                ),
                "unseen_current_categories": (
                    decision_metric.unseen_current_categories
                ),
                "missing_reference_categories": (
                    decision_metric.missing_reference_categories
                ),
                "health_score": (
                    decision_metric.health_score
                ),
                "notes": decision_metric.notes,
            },
        )
    ]


def quality_metric_thresholds(
    context: MonitoringContext,
    metric: str,
) -> Mapping[str, Any]:
    mapping = {
        "Current total missing rate": "missing_rate",
        "Current duplicate-row rate": "duplicate_rate",
        "Current constant-feature rate": "constant_feature_rate",
        "Unmatched-feature rate": "unmatched_feature_rate",
    }

    threshold_key = mapping.get(
        metric
    )

    if threshold_key is None:
        return {}

    return (
        context.config.get(
            "data_quality_thresholds",
            {}
        ).get(
            threshold_key,
            {},
        )
    )


def build_data_quality_alerts(
    context: MonitoringContext,
    data_quality_table: pd.DataFrame,
) -> list[MonitoringAlert]:
    alerts: list[MonitoringAlert] = []

    for row in data_quality_table.to_dict(
        orient="records"
    ):
        severity = safe_text(
            row.get(
                "severity"
            ),
            "Normal",
        )

        if severity == "Normal":
            continue

        metric_name = safe_text(
            row.get(
                "metric"
            )
        )

        observed_value = safe_float(
            row.get(
                "value"
            )
        )

        thresholds = quality_metric_thresholds(
            context,
            metric_name,
        )

        threshold_value = alert_threshold_for_severity(
            thresholds,
            severity,
        )

        alerts.append(
            MonitoringAlert(
                alert_id=deterministic_alert_id(
                    run_id=context.run_identity.run_id,
                    category="Data Quality",
                    alert_type="data_quality_degradation",
                    resource_id=metric_name,
                    metric_name=normalise_column_name(
                        metric_name
                    ),
                ),
                run_id=context.run_identity.run_id,
                generated_utc=context.run_identity.generated_utc,
                category="Data Quality",
                alert_type="data_quality_degradation",
                severity=severity,
                resource_type="deployment_batch",
                resource_id=metric_name,
                metric_name=normalise_column_name(
                    metric_name
                ),
                observed_value=observed_value,
                threshold_value=threshold_value,
                message=(
                    f"{metric_name} is classified as {severity}; "
                    f"observed value={observed_value}."
                ),
                recommended_action=severity_action(
                    severity,
                    "Data Quality",
                ),
                evidence={
                    "reference_value": row.get(
                        "reference_value"
                    ),
                    "severity_rank": row.get(
                        "severity_rank"
                    ),
                },
            )
        )

    return alerts


def build_unmatched_feature_alerts(
    context: MonitoringContext,
    unmatched_features: pd.DataFrame,
) -> list[MonitoringAlert]:
    alerts: list[MonitoringAlert] = []

    for row in unmatched_features.to_dict(
        orient="records"
    ):
        severity = safe_text(
            row.get(
                "severity"
            ),
            "Warning",
        )

        feature = safe_text(
            row.get(
                "feature"
            )
        )

        status = safe_text(
            row.get(
                "feature_status"
            )
        )

        alerts.append(
            MonitoringAlert(
                alert_id=deterministic_alert_id(
                    run_id=context.run_identity.run_id,
                    category="Data Quality",
                    alert_type="feature_schema_mismatch",
                    resource_id=feature,
                    metric_name="feature_schema_alignment",
                ),
                run_id=context.run_identity.run_id,
                generated_utc=context.run_identity.generated_utc,
                category="Data Quality",
                alert_type="feature_schema_mismatch",
                severity=severity,
                resource_type="sensor_feature",
                resource_id=feature,
                metric_name="feature_schema_alignment",
                observed_value=None,
                threshold_value=None,
                message=(
                    f"{feature}: {status}."
                ),
                recommended_action=severity_action(
                    severity,
                    "Data Quality",
                ),
                evidence={
                    "reference_available": row.get(
                        "reference_available"
                    ),
                    "current_available": row.get(
                        "current_available"
                    ),
                },
            )
        )

    return alerts


def alerts_to_dataframe(
    alerts: Sequence[MonitoringAlert],
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []

    for alert in alerts:
        record = asdict(
            alert
        )

        record[
            "evidence_json"
        ] = json.dumps(
            json_safe(
                record.pop(
                    "evidence"
                )
            ),
            ensure_ascii=False,
            sort_keys=True,
        )

        records.append(
            record
        )

    columns = [
        "alert_id",
        "run_id",
        "generated_utc",
        "category",
        "alert_type",
        "severity",
        "resource_type",
        "resource_id",
        "metric_name",
        "observed_value",
        "threshold_value",
        "message",
        "recommended_action",
        "evidence_json",
    ]

    dataframe = pd.DataFrame(
        records,
        columns=columns,
    )

    if dataframe.empty:
        dataframe[
            "severity_rank"
        ] = pd.Series(
            dtype=int
        )

        return dataframe

    dataframe[
        "severity_rank"
    ] = dataframe[
        "severity"
    ].map(
        SEVERITY_ORDER
    ).fillna(
        0
    ).astype(
        int
    )

    dataframe = dataframe.sort_values(
        by=[
            "severity_rank",
            "category",
            "resource_id",
        ],
        ascending=[
            False,
            True,
            True,
        ],
        na_position="last",
    ).reset_index(
        drop=True
    )

    return dataframe


def build_alert_engine(
    context: MonitoringContext,
    feature_drift: pd.DataFrame,
    unmatched_features: pd.DataFrame,
    prediction_drift: pd.DataFrame,
    decision_metric: DecisionDriftMetric | None,
    data_quality_table: pd.DataFrame,
) -> pd.DataFrame:
    alerts: list[MonitoringAlert] = []

    alerts.extend(
        build_feature_alerts(
            context,
            feature_drift,
        )
    )

    alerts.extend(
        build_prediction_alerts(
            context,
            prediction_drift,
        )
    )

    alerts.extend(
        build_decision_alert(
            context,
            decision_metric,
        )
    )

    alerts.extend(
        build_data_quality_alerts(
            context,
            data_quality_table,
        )
    )

    alerts.extend(
        build_unmatched_feature_alerts(
            context,
            unmatched_features,
        )
    )

    return alerts_to_dataframe(
        alerts
    )


def health_band_from_score(
    score: float,
) -> str:
    if score >= 97.0:
        return "Excellent"

    if score >= 90.0:
        return "Healthy"

    if score >= 80.0:
        return "Watch"

    if score >= 65.0:
        return "Warning"

    return "Critical"


def severity_from_health_band(
    band: str,
) -> str:
    mapping = {
        "Excellent": "Normal",
        "Healthy": "Normal",
        "Watch": "Watch",
        "Warning": "Warning",
        "Critical": "Critical",
    }

    return mapping.get(
        band,
        "Warning",
    )


def count_severity(
    dataframe: pd.DataFrame,
    column: str,
    severity: str,
) -> int:
    if (
        dataframe.empty
        or column not in dataframe.columns
    ):
        return 0

    return int(
        (
            dataframe[
                column
            ]
            == severity
        ).sum()
    )


def component_status_label(
    score: float,
) -> str:
    return health_band_from_score(
        score
    )


def feature_component(
    feature_drift: pd.DataFrame,
    feature_summary: FeatureDriftSummary,
    weight: float,
) -> HealthComponent:
    score = float(
        feature_summary.mean_feature_health_score
    )

    return HealthComponent(
        component_name="feature_drift",
        weight=weight,
        raw_score=score,
        weighted_score=score
        * weight,
        severity=feature_summary.dominant_severity,
        status_label=component_status_label(
            score
        ),
        monitored_items=feature_summary.monitored_features,
        warning_items=(
            feature_summary.warning_features
        ),
        critical_items=(
            feature_summary.critical_features
        ),
        explanation=(
            f"{feature_summary.drifted_features} of "
            f"{feature_summary.monitored_features} monitored features "
            f"show Watch-or-higher drift."
        ),
    )


def prediction_component(
    prediction_drift: pd.DataFrame,
    weight: float,
) -> HealthComponent:
    numeric = pd.to_numeric(
        prediction_drift[
            "health_score"
        ],
        errors="coerce",
    )

    score = (
        float(
            numeric.mean()
        )
        if numeric.notna().any()
        else 70.0
    )

    severity = maximum_severity(
        prediction_drift[
            "severity"
        ].tolist()
    )

    return HealthComponent(
        component_name="prediction_drift",
        weight=weight,
        raw_score=round(
            score,
            2,
        ),
        weighted_score=score
        * weight,
        severity=severity,
        status_label=component_status_label(
            score
        ),
        monitored_items=int(
            len(
                prediction_drift
            )
        ),
        warning_items=count_severity(
            prediction_drift,
            "severity",
            "Warning",
        ),
        critical_items=count_severity(
            prediction_drift,
            "severity",
            "Critical",
        ),
        explanation=(
            "Monitors failure probability and other available operational "
            "output distributions."
        ),
    )


def confidence_uncertainty_component(
    prediction_drift: pd.DataFrame,
    weight: float,
) -> HealthComponent:
    subset = prediction_drift.loc[
        prediction_drift[
            "metric_name"
        ].isin(
            [
                "prediction_confidence",
                "combined_uncertainty",
                "data_confidence",
            ]
        )
    ].copy()

    if subset.empty:
        return HealthComponent(
            component_name="confidence_and_uncertainty",
            weight=weight,
            raw_score=70.0,
            weighted_score=70.0
            * weight,
            severity="Warning",
            status_label="Warning",
            monitored_items=0,
            warning_items=1,
            critical_items=0,
            explanation=(
                "Confidence and uncertainty outputs are unavailable for "
                "complete monitoring."
            ),
        )

    numeric = pd.to_numeric(
        subset[
            "health_score"
        ],
        errors="coerce",
    )

    score = (
        float(
            numeric.mean()
        )
        if numeric.notna().any()
        else 70.0
    )

    severity = maximum_severity(
        subset[
            "severity"
        ].tolist()
    )

    return HealthComponent(
        component_name="confidence_and_uncertainty",
        weight=weight,
        raw_score=round(
            score,
            2,
        ),
        weighted_score=score
        * weight,
        severity=severity,
        status_label=component_status_label(
            score
        ),
        monitored_items=int(
            len(
                subset
            )
        ),
        warning_items=count_severity(
            subset,
            "severity",
            "Warning",
        ),
        critical_items=count_severity(
            subset,
            "severity",
            "Critical",
        ),
        explanation=(
            "Measures drift in prediction confidence, uncertainty and "
            "available input-data confidence."
        ),
    )


def data_quality_component(
    data_quality_summary: DataQualitySummary,
    data_quality_table: pd.DataFrame,
    weight: float,
) -> HealthComponent:
    score = float(
        data_quality_summary.health_score
    )

    return HealthComponent(
        component_name="data_quality",
        weight=weight,
        raw_score=score,
        weighted_score=score
        * weight,
        severity=data_quality_summary.severity,
        status_label=component_status_label(
            score
        ),
        monitored_items=int(
            len(
                data_quality_table
            )
        ),
        warning_items=count_severity(
            data_quality_table,
            "severity",
            "Warning",
        ),
        critical_items=count_severity(
            data_quality_table,
            "severity",
            "Critical",
        ),
        explanation=(
            "Covers missingness, duplication, constant sensors, invalid "
            "probability-like values and schema mismatch."
        ),
    )


def normalised_health_weights(
    config: Mapping[str, Any],
) -> dict[str, float]:
    expected_components = [
        "data_quality",
        "feature_drift",
        "prediction_drift",
        "confidence_and_uncertainty",
    ]

    configured = config.get(
        "health_weights",
        {}
    )

    weights = {
        component: (
            safe_float(
                configured.get(
                    component
                ),
                0.0,
            )
            or 0.0
        )
        for component in expected_components
    }

    total = sum(
        weights.values()
    )

    if total <= 0:
        raise ValueError(
            "Health weights must contain a positive total."
        )

    return {
        component: weight
        / total
        for component, weight in weights.items()
    }


def health_components_to_dataframe(
    components: Sequence[HealthComponent],
) -> pd.DataFrame:
    dataframe = pd.DataFrame(
        [
            asdict(
                component
            )
            for component in components
        ]
    )

    if dataframe.empty:
        raise RuntimeError(
            "Health component table is empty."
        )

    return dataframe


def alert_counts(
    alerts: pd.DataFrame,
) -> dict[str, int]:
    if alerts.empty:
        return {
            "total": 0,
            "watch": 0,
            "warning": 0,
            "critical": 0,
        }

    return {
        "total": int(
            len(
                alerts
            )
        ),
        "watch": count_severity(
            alerts,
            "severity",
            "Watch",
        ),
        "warning": count_severity(
            alerts,
            "severity",
            "Warning",
        ),
        "critical": count_severity(
            alerts,
            "severity",
            "Critical",
        ),
    }


def determine_retraining_recommendation(
    *,
    overall_score: float,
    critical_alerts: int,
    warning_alerts: int,
    critical_feature_rate: float,
    data_quality_severity: str,
    current_rows: int,
    minimum_current_rows: int,
) -> tuple[
    str,
    str,
    bool,
    bool,
]:
    if current_rows < minimum_current_rows:
        return (
            "Increase Monitoring",
            (
                "The current monitoring window is below the configured "
                "minimum; collect more production records before retraining."
            ),
            False,
            False,
        )

    if (
        critical_alerts
        >= 3
        or critical_feature_rate
        >= 0.15
        or (
            overall_score
            < 60.0
            and data_quality_severity
            != "Critical"
        )
    ):
        return (
            "Immediate Retraining",
            (
                "Multiple critical monitoring signals indicate that the "
                "deployed model may no longer represent the current process."
            ),
            True,
            True,
        )

    if (
        overall_score
        < 65.0
        or critical_alerts
        > 0
    ):
        return (
            "Retrain Soon",
            (
                "At least one critical signal or a low deployment-health "
                "score requires a controlled retraining assessment."
            ),
            True,
            True,
        )

    if (
        overall_score
        < 80.0
        or warning_alerts
        >= 5
    ):
        return (
            "Retrain Soon",
            (
                "Sustained warning-level drift warrants preparation of a "
                "new labelled validation and retraining cycle."
            ),
            True,
            False,
        )

    if (
        overall_score
        < 90.0
        or warning_alerts
        > 0
    ):
        return (
            "Increase Monitoring",
            (
                "Moderate drift is present but does not yet justify an "
                "immediate model replacement."
            ),
            False,
            False,
        )

    return (
        "Continue Monitoring",
        (
            "Current deployment evidence remains within the accepted "
            "monitoring envelope."
        ),
        False,
        False,
    )


def primary_risk_driver(
    components: Sequence[HealthComponent],
) -> str:
    if not components:
        return "Unavailable"

    worst = min(
        components,
        key=lambda component: (
            component.raw_score,
            -SEVERITY_ORDER.get(
                component.severity,
                0,
            ),
        ),
    )

    readable = worst.component_name.replace(
        "_",
        " ",
    ).title()

    return (
        f"{readable} ({worst.raw_score:.2f}/100, "
        f"{worst.severity})"
    )


def build_deployment_health_assessment(
    context: MonitoringContext,
    feature_drift: pd.DataFrame,
    feature_summary: FeatureDriftSummary,
    prediction_drift: pd.DataFrame,
    data_quality_summary: DataQualitySummary,
    data_quality_table: pd.DataFrame,
    alerts: pd.DataFrame,
) -> tuple[
    DeploymentHealthAssessment,
    pd.DataFrame,
]:
    weights = normalised_health_weights(
        context.config
    )

    components = [
        data_quality_component(
            data_quality_summary,
            data_quality_table,
            weights[
                "data_quality"
            ],
        ),
        feature_component(
            feature_drift,
            feature_summary,
            weights[
                "feature_drift"
            ],
        ),
        prediction_component(
            prediction_drift,
            weights[
                "prediction_drift"
            ],
        ),
        confidence_uncertainty_component(
            prediction_drift,
            weights[
                "confidence_and_uncertainty"
            ],
        ),
    ]

    component_table = health_components_to_dataframe(
        components
    )

    overall_score = float(
        sum(
            component.weight
            * component.raw_score
            for component in components
        )
    )

    overall_score = round(
        max(
            min(
                overall_score,
                100.0,
            ),
            0.0,
        ),
        2,
    )

    band = health_band_from_score(
        overall_score
    )

    overall_severity = maximum_severity(
        [
            severity_from_health_band(
                band
            ),
            *[
                component.severity
                for component in components
            ],
        ]
    )

    counts = alert_counts(
        alerts
    )

    critical_feature_rate = (
        feature_summary.critical_features
        / max(
            feature_summary.monitored_features,
            1,
        )
    )

    minimum_current_rows = int(
        context.config.get(
            "minimum_current_rows",
            30,
        )
    )

    (
        recommendation,
        recommendation_reason,
        retraining_triggered,
        immediate_action_required,
    ) = determine_retraining_recommendation(
        overall_score=overall_score,
        critical_alerts=counts[
            "critical"
        ],
        warning_alerts=counts[
            "warning"
        ],
        critical_feature_rate=critical_feature_rate,
        data_quality_severity=(
            data_quality_summary.severity
        ),
        current_rows=(
            context.current.descriptor.row_count
        ),
        minimum_current_rows=minimum_current_rows,
    )

    if (
        data_quality_summary.severity
        == "Critical"
    ):
        production_status = (
            "Predictions require data verification before operational use"
        )
    elif immediate_action_required:
        production_status = (
            "Escalated engineering and model-owner review required"
        )
    elif overall_severity in {
        "Warning",
        "Critical",
    }:
        production_status = (
            "Deployment may continue only under enhanced review"
        )
    elif overall_severity == "Watch":
        production_status = (
            "Deployment active with increased monitoring"
        )
    else:
        production_status = (
            "Deployment monitoring within accepted limits"
        )

    risk_driver = primary_risk_driver(
        components
    )

    summary_statement = (
        f"Deployment health is {band} at {overall_score:.2f}/100. "
        f"The primary risk driver is {risk_driver}. "
        f"Recommendation: {recommendation}."
    )

    assessment = DeploymentHealthAssessment(
        run_id=context.run_identity.run_id,
        generated_utc=context.run_identity.generated_utc,
        overall_health_score=overall_score,
        overall_health_band=band,
        overall_severity=overall_severity,
        recommendation=recommendation,
        recommendation_reason=recommendation_reason,
        production_monitoring_status=production_status,
        component_scores={
            component.component_name: (
                component.raw_score
            )
            for component in components
        },
        component_weights={
            component.component_name: (
                component.weight
            )
            for component in components
        },
        alert_count=counts[
            "total"
        ],
        watch_alerts=counts[
            "watch"
        ],
        warning_alerts=counts[
            "warning"
        ],
        critical_alerts=counts[
            "critical"
        ],
        retraining_triggered=retraining_triggered,
        immediate_action_required=(
            immediate_action_required
        ),
        model_modification_performed=False,
        primary_risk_driver=risk_driver,
        summary_statement=summary_statement,
    )

    return (
        assessment,
        component_table,
    )


def validate_health_assessment(
    assessment: DeploymentHealthAssessment,
    component_table: pd.DataFrame,
    alerts: pd.DataFrame,
) -> None:
    if not (
        0.0
        <= assessment.overall_health_score
        <= 100.0
    ):
        raise RuntimeError(
            "Overall deployment-health score is outside [0, 100]."
        )

    expected_bands = {
        "Excellent",
        "Healthy",
        "Watch",
        "Warning",
        "Critical",
    }

    if assessment.overall_health_band not in expected_bands:
        raise RuntimeError(
            "Unexpected deployment-health band."
        )

    if (
        assessment.overall_severity
        not in SEVERITY_ORDER
    ):
        raise RuntimeError(
            "Unexpected overall monitoring severity."
        )

    required_components = {
        "data_quality",
        "feature_drift",
        "prediction_drift",
        "confidence_and_uncertainty",
    }

    actual_components = set(
        component_table[
            "component_name"
        ].tolist()
    )

    if actual_components != required_components:
        raise RuntimeError(
            "Health component set does not match the monitoring policy."
        )

    raw_scores = pd.to_numeric(
        component_table[
            "raw_score"
        ],
        errors="coerce",
    )

    if (
        (
            raw_scores
            < 0
        )
        | (
            raw_scores
            > 100
        )
    ).any():
        raise RuntimeError(
            "A component health score is outside [0, 100]."
        )

    weights = pd.to_numeric(
        component_table[
            "weight"
        ],
        errors="coerce",
    )

    if not math.isclose(
        float(
            weights.sum()
        ),
        1.0,
        rel_tol=1e-6,
        abs_tol=1e-6,
    ):
        raise RuntimeError(
            "Health component weights do not sum to 1.0."
        )

    if not alerts.empty:
        invalid_alert_severities = set(
            alerts[
                "severity"
            ].dropna().unique()
        ) - set(
            SEVERITY_ORDER
        )

        if invalid_alert_severities:
            raise RuntimeError(
                "Alert table contains invalid severity labels."
            )

        if alerts[
            "alert_id"
        ].duplicated().any():
            raise RuntimeError(
                "Alert table contains duplicate alert identifiers."
            )


def save_health_outputs(
    context: MonitoringContext,
    alerts: pd.DataFrame,
    component_table: pd.DataFrame,
    assessment: DeploymentHealthAssessment,
) -> None:
    ensure_directories()

    alerts.to_csv(
        ALERTS_PATH,
        index=False,
    )

    component_table.to_csv(
        HEALTH_COMPONENTS_PATH,
        index=False,
    )

    atomic_write_json(
        DEPLOYMENT_HEALTH_PATH,
        {
            "run_identity": asdict(
                context.run_identity
            ),
            "deployment_health": asdict(
                assessment
            ),
            "alerts_path": str(
                ALERTS_PATH
            ),
            "health_components_path": str(
                HEALTH_COMPONENTS_PATH
            ),
            "monitoring_policy": {
                "health_weights": (
                    context.config.get(
                        "health_weights"
                    )
                ),
                "feature_drift_thresholds": (
                    context.config.get(
                        "feature_drift_thresholds"
                    )
                ),
                "prediction_drift_thresholds": (
                    context.config.get(
                        "prediction_drift_thresholds"
                    )
                ),
                "data_quality_thresholds": (
                    context.config.get(
                        "data_quality_thresholds"
                    )
                ),
            },
        },
    )


def initialise_monitoring_history_database() -> None:
    ensure_directories()

    with sqlite3.connect(
        HISTORY_DATABASE_PATH
    ) as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS monitoring_runs (
                run_id TEXT PRIMARY KEY,
                generated_utc TEXT NOT NULL,
                reference_sha256 TEXT NOT NULL,
                current_sha256 TEXT NOT NULL,
                reference_path TEXT NOT NULL,
                current_path TEXT NOT NULL,
                overall_health_score REAL NOT NULL,
                overall_health_band TEXT NOT NULL,
                overall_severity TEXT NOT NULL,
                recommendation TEXT NOT NULL,
                alert_count INTEGER NOT NULL,
                watch_alerts INTEGER NOT NULL,
                warning_alerts INTEGER NOT NULL,
                critical_alerts INTEGER NOT NULL,
                retraining_triggered INTEGER NOT NULL,
                immediate_action_required INTEGER NOT NULL,
                primary_risk_driver TEXT NOT NULL,
                summary_statement TEXT NOT NULL,
                created_utc TEXT NOT NULL
            )
            """
        )

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS monitoring_alert_history (
                alert_id TEXT PRIMARY KEY,
                run_id TEXT NOT NULL,
                generated_utc TEXT NOT NULL,
                category TEXT NOT NULL,
                alert_type TEXT NOT NULL,
                severity TEXT NOT NULL,
                resource_type TEXT NOT NULL,
                resource_id TEXT,
                metric_name TEXT NOT NULL,
                observed_value REAL,
                threshold_value REAL,
                message TEXT NOT NULL,
                recommended_action TEXT NOT NULL,
                evidence_json TEXT NOT NULL,
                FOREIGN KEY(run_id) REFERENCES monitoring_runs(run_id)
            )
            """
        )

        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_monitoring_runs_generated
            ON monitoring_runs(generated_utc)
            """
        )

        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_monitoring_alerts_run
            ON monitoring_alert_history(run_id)
            """
        )

        connection.commit()


def append_monitoring_history_csv(
    assessment: DeploymentHealthAssessment,
) -> None:
    row = pd.DataFrame(
        [
            {
                "run_id": assessment.run_id,
                "generated_utc": (
                    assessment.generated_utc
                ),
                "overall_health_score": (
                    assessment.overall_health_score
                ),
                "overall_health_band": (
                    assessment.overall_health_band
                ),
                "overall_severity": (
                    assessment.overall_severity
                ),
                "recommendation": (
                    assessment.recommendation
                ),
                "alert_count": (
                    assessment.alert_count
                ),
                "watch_alerts": (
                    assessment.watch_alerts
                ),
                "warning_alerts": (
                    assessment.warning_alerts
                ),
                "critical_alerts": (
                    assessment.critical_alerts
                ),
                "retraining_triggered": (
                    assessment.retraining_triggered
                ),
                "immediate_action_required": (
                    assessment.immediate_action_required
                ),
                "primary_risk_driver": (
                    assessment.primary_risk_driver
                ),
            }
        ]
    )

    if HISTORY_CSV_PATH.exists():
        existing = pd.read_csv(
            HISTORY_CSV_PATH
        )

        existing = existing.loc[
            existing[
                "run_id"
            ]
            != assessment.run_id
        ].copy()

        combined = pd.concat(
            [
                existing,
                row,
            ],
            ignore_index=True,
        )
    else:
        combined = row

    combined = combined.sort_values(
        "generated_utc"
    ).reset_index(
        drop=True
    )

    combined.to_csv(
        HISTORY_CSV_PATH,
        index=False,
    )


def persist_monitoring_history(
    context: MonitoringContext,
    assessment: DeploymentHealthAssessment,
    alerts: pd.DataFrame,
) -> None:
    initialise_monitoring_history_database()

    with sqlite3.connect(
        HISTORY_DATABASE_PATH
    ) as connection:
        connection.execute(
            """
            INSERT OR REPLACE INTO monitoring_runs (
                run_id,
                generated_utc,
                reference_sha256,
                current_sha256,
                reference_path,
                current_path,
                overall_health_score,
                overall_health_band,
                overall_severity,
                recommendation,
                alert_count,
                watch_alerts,
                warning_alerts,
                critical_alerts,
                retraining_triggered,
                immediate_action_required,
                primary_risk_driver,
                summary_statement,
                created_utc
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
            """,
            (
                assessment.run_id,
                assessment.generated_utc,
                context.run_identity.reference_sha256,
                context.run_identity.current_sha256,
                context.run_identity.reference_path,
                context.run_identity.current_path,
                assessment.overall_health_score,
                assessment.overall_health_band,
                assessment.overall_severity,
                assessment.recommendation,
                assessment.alert_count,
                assessment.watch_alerts,
                assessment.warning_alerts,
                assessment.critical_alerts,
                int(
                    assessment.retraining_triggered
                ),
                int(
                    assessment.immediate_action_required
                ),
                assessment.primary_risk_driver,
                assessment.summary_statement,
                utc_now_iso(),
            ),
        )

        if not alerts.empty:
            for row in alerts.to_dict(
                orient="records"
            ):
                connection.execute(
                    """
                    INSERT OR REPLACE INTO monitoring_alert_history (
                        alert_id,
                        run_id,
                        generated_utc,
                        category,
                        alert_type,
                        severity,
                        resource_type,
                        resource_id,
                        metric_name,
                        observed_value,
                        threshold_value,
                        message,
                        recommended_action,
                        evidence_json
                    ) VALUES (
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    )
                    """,
                    (
                        row.get(
                            "alert_id"
                        ),
                        row.get(
                            "run_id"
                        ),
                        row.get(
                            "generated_utc"
                        ),
                        row.get(
                            "category"
                        ),
                        row.get(
                            "alert_type"
                        ),
                        row.get(
                            "severity"
                        ),
                        row.get(
                            "resource_type"
                        ),
                        row.get(
                            "resource_id"
                        ),
                        row.get(
                            "metric_name"
                        ),
                        safe_float(
                            row.get(
                                "observed_value"
                            )
                        ),
                        safe_float(
                            row.get(
                                "threshold_value"
                            )
                        ),
                        row.get(
                            "message"
                        ),
                        row.get(
                            "recommended_action"
                        ),
                        row.get(
                            "evidence_json",
                            "{}",
                        ),
                    ),
                )

        connection.commit()

    append_monitoring_history_csv(
        assessment
    )


def run_health_and_alert_engine(
    *,
    context: MonitoringContext,
    feature_drift: pd.DataFrame,
    feature_summary: FeatureDriftSummary,
    unmatched_features: pd.DataFrame,
    prediction_drift: pd.DataFrame,
    decision_metric: DecisionDriftMetric | None,
    data_quality_summary: DataQualitySummary,
    data_quality_table: pd.DataFrame,
    save_output: bool = True,
    persist_history: bool = True,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    DeploymentHealthAssessment,
]:
    alerts = build_alert_engine(
        context,
        feature_drift,
        unmatched_features,
        prediction_drift,
        decision_metric,
        data_quality_table,
    )

    (
        assessment,
        component_table,
    ) = build_deployment_health_assessment(
        context,
        feature_drift,
        feature_summary,
        prediction_drift,
        data_quality_summary,
        data_quality_table,
        alerts,
    )

    validate_health_assessment(
        assessment,
        component_table,
        alerts,
    )

    if save_output:
        save_health_outputs(
            context,
            alerts,
            component_table,
            assessment,
        )

    if persist_history:
        persist_monitoring_history(
            context,
            assessment,
            alerts,
        )

    record_monitoring_event(
        event_type="model_monitoring_health_assessed",
        outcome=(
            "warning"
            if assessment.overall_severity
            in {
                "Warning",
                "Critical",
            }
            else "success"
        ),
        action="assess_deployment_health",
        resource_id=assessment.run_id,
        details={
            "overall_health_score": (
                assessment.overall_health_score
            ),
            "overall_health_band": (
                assessment.overall_health_band
            ),
            "overall_severity": (
                assessment.overall_severity
            ),
            "recommendation": (
                assessment.recommendation
            ),
            "critical_alerts": (
                assessment.critical_alerts
            ),
            "warning_alerts": (
                assessment.warning_alerts
            ),
            "retraining_triggered": (
                assessment.retraining_triggered
            ),
            "model_modification_performed": False,
        },
    )

    return (
        alerts,
        component_table,
        assessment,
    )


def build_health_console_summary(
    assessment: DeploymentHealthAssessment,
) -> str:
    lines = [
        "="
        * 118,
        "HEVEMIND DEPLOYMENT HEALTH AND RETRAINING POLICY",
        "="
        * 118,
        "",
        (
            "Overall health score:            "
            f"{assessment.overall_health_score:.2f}"
        ),
        (
            "Health band:                     "
            f"{assessment.overall_health_band}"
        ),
        (
            "Overall severity:                "
            f"{assessment.overall_severity}"
        ),
        (
            "Production monitoring status:    "
            f"{assessment.production_monitoring_status}"
        ),
        (
            "Recommendation:                  "
            f"{assessment.recommendation}"
        ),
        (
            "Retraining trigger:              "
            f"{assessment.retraining_triggered}"
        ),
        (
            "Immediate action required:       "
            f"{assessment.immediate_action_required}"
        ),
        (
            "Model modified by monitoring:    "
            f"{assessment.model_modification_performed}"
        ),
        (
            "Primary risk driver:             "
            f"{assessment.primary_risk_driver}"
        ),
        "",
        (
            "Total alerts:                    "
            f"{assessment.alert_count}"
        ),
        (
            "Watch alerts:                    "
            f"{assessment.watch_alerts}"
        ),
        (
            "Warning alerts:                  "
            f"{assessment.warning_alerts}"
        ),
        (
            "Critical alerts:                 "
            f"{assessment.critical_alerts}"
        ),
        "",
        (
            "Recommendation reason:           "
            f"{assessment.recommendation_reason}"
        ),
        "",
        f"Alerts table:                     {ALERTS_PATH}",
        f"Health components:                {HEALTH_COMPONENTS_PATH}",
        f"Deployment health:                {DEPLOYMENT_HEALTH_PATH}",
        f"Monitoring history database:      {HISTORY_DATABASE_PATH}",
    ]

    return "\n".join(
        lines
    )


# End of Part 4.
# Append Part 5 immediately below this line.


# ============================================================
# PART 5 — FIGURES, JSON, EXCEL AND PDF REPORTING
# ============================================================

@dataclass(frozen=True)
class MonitoringReportArtifacts:
    run_id: str
    generated_utc: str
    drift_summary_json: Path
    deployment_health_json: Path
    feature_drift_csv: Path
    prediction_drift_csv: Path
    decision_drift_csv: Path
    data_quality_csv: Path
    alerts_csv: Path
    health_components_csv: Path
    monitoring_excel: Path
    monitoring_pdf: Path | None
    figure_paths: tuple[Path, ...]
    report_sha256: str | None


def report_timestamp(
    generated_utc: str,
) -> str:
    try:
        parsed = datetime.fromisoformat(
            generated_utc.replace(
                "Z",
                "+00:00",
            )
        )
    except ValueError:
        parsed = datetime.now(
            timezone.utc
        )

    return parsed.astimezone(
        timezone.utc
    ).strftime(
        "%Y%m%d_%H%M%S"
    )


def path_with_run_suffix(
    base_path: Path,
    run_id: str,
    generated_utc: str,
) -> Path:
    timestamp = report_timestamp(
        generated_utc
    )

    return base_path.with_name(
        f"{base_path.stem}_{timestamp}_{run_id}{base_path.suffix}"
    )


def monitoring_bundle_to_summary_dict(
    bundle: MonitoringBundle,
) -> dict[str, Any]:
    return {
        "application": {
            "name": APP_NAME,
            "version": APP_VERSION,
        },
        "run_identity": asdict(
            bundle.context.run_identity
        ),
        "reference_dataset": json_safe(
            asdict(
                bundle.context.reference.descriptor
            )
        ),
        "current_dataset": json_safe(
            asdict(
                bundle.context.current.descriptor
            )
        ),
        "feature_alignment": {
            "shared_features": len(
                bundle.context.shared_features
            ),
            "reference_only_features": (
                bundle.context.reference_only_features
            ),
            "current_only_features": (
                bundle.context.current_only_features
            ),
        },
        "feature_drift_summary": asdict(
            bundle.feature_summary
        ),
        "prediction_monitoring_summary": asdict(
            bundle.prediction_summary
        ),
        "decision_drift": (
            asdict(
                bundle.decision_metric
            )
            if bundle.decision_metric is not None
            else None
        ),
        "data_quality_summary": asdict(
            bundle.data_quality_summary
        ),
        "deployment_health": asdict(
            bundle.health_assessment
        ),
        "warnings": (
            bundle.context.warnings
        ),
        "model_modification_performed": False,
        "monitoring_scope": (
            "Monitoring, alerting and retraining recommendation only"
        ),
    }


def save_consolidated_json_reports(
    bundle: MonitoringBundle,
) -> tuple[
    Path,
    Path,
]:
    ensure_directories()

    summary_payload = monitoring_bundle_to_summary_dict(
        bundle
    )

    atomic_write_json(
        DRIFT_SUMMARY_PATH,
        summary_payload,
    )

    atomic_write_json(
        DEPLOYMENT_HEALTH_PATH,
        {
            "run_identity": asdict(
                bundle.context.run_identity
            ),
            "deployment_health": asdict(
                bundle.health_assessment
            ),
            "health_components": (
                bundle.health_components.to_dict(
                    orient="records"
                )
            ),
            "alerts": (
                bundle.alerts.to_dict(
                    orient="records"
                )
            ),
            "model_modification_performed": False,
        },
    )

    archival_summary_path = path_with_run_suffix(
        DRIFT_SUMMARY_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    archival_health_path = path_with_run_suffix(
        DEPLOYMENT_HEALTH_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    atomic_write_json(
        archival_summary_path,
        summary_payload,
    )

    atomic_write_json(
        archival_health_path,
        {
            "run_identity": asdict(
                bundle.context.run_identity
            ),
            "deployment_health": asdict(
                bundle.health_assessment
            ),
            "health_components": (
                bundle.health_components.to_dict(
                    orient="records"
                )
            ),
            "alerts": (
                bundle.alerts.to_dict(
                    orient="records"
                )
            ),
            "model_modification_performed": False,
        },
    )

    return (
        archival_summary_path,
        archival_health_path,
    )


def dataframe_for_excel(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    output = dataframe.copy()

    for column in output.columns:
        if output[
            column
        ].map(
            lambda value: isinstance(
                value,
                (
                    dict,
                    list,
                    tuple,
                    set,
                ),
            )
        ).any():
            output[
                column
            ] = output[
                column
            ].map(
                lambda value: json.dumps(
                    json_safe(
                        value
                    ),
                    ensure_ascii=False,
                    sort_keys=True,
                )
                if isinstance(
                    value,
                    (
                        dict,
                        list,
                        tuple,
                        set,
                    ),
                )
                else value
            )

    return output


def autosize_excel_worksheet(
    worksheet: Any,
    dataframe: pd.DataFrame,
    maximum_width: int = 50,
) -> None:
    from openpyxl.utils import get_column_letter

    for index, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        values = [
            safe_text(
                column,
                "",
            )
        ]

        values.extend(
            safe_text(
                value,
                "",
            )
            for value in dataframe[
                column
            ].head(
                500
            )
        )

        width = min(
            max(
                len(
                    value
                )
                for value in values
            )
            + 2,
            maximum_width,
        )

        worksheet.column_dimensions[
            get_column_letter(
                index
            )
        ].width = max(
            width,
            10,
        )

    worksheet.freeze_panes = "A2"

    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


def format_excel_workbook(
    writer: Any,
    sheets: Mapping[str, pd.DataFrame],
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="123252",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for sheet_name, dataframe in sheets.items():
        worksheet = writer.book[
            sheet_name
        ]

        for cell in worksheet[
            1
        ]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        autosize_excel_worksheet(
            worksheet,
            dataframe,
        )

        for row in worksheet.iter_rows(
            min_row=2,
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )


def export_monitoring_excel(
    bundle: MonitoringBundle,
) -> Path:
    ensure_directories()

    output_path = path_with_run_suffix(
        EXCEL_REPORT_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    deployment_health_frame = pd.DataFrame(
        [
            asdict(
                bundle.health_assessment
            )
        ]
    )

    feature_summary_frame = pd.DataFrame(
        [
            asdict(
                bundle.feature_summary
            )
        ]
    )

    prediction_summary_frame = pd.DataFrame(
        [
            asdict(
                bundle.prediction_summary
            )
        ]
    )

    data_quality_summary_frame = pd.DataFrame(
        [
            asdict(
                bundle.data_quality_summary
            )
        ]
    )

    context_frame = pd.DataFrame(
        [
            {
                "run_id": (
                    bundle.context.run_identity.run_id
                ),
                "generated_utc": (
                    bundle.context.run_identity.generated_utc
                ),
                "reference_path": (
                    bundle.context.run_identity.reference_path
                ),
                "current_path": (
                    bundle.context.run_identity.current_path
                ),
                "reference_sha256": (
                    bundle.context.run_identity.reference_sha256
                ),
                "current_sha256": (
                    bundle.context.run_identity.current_sha256
                ),
                "reference_rows": (
                    bundle.context.reference.descriptor.row_count
                ),
                "current_rows": (
                    bundle.context.current.descriptor.row_count
                ),
                "shared_features": len(
                    bundle.context.shared_features
                ),
                "reference_only_features": len(
                    bundle.context.reference_only_features
                ),
                "current_only_features": len(
                    bundle.context.current_only_features
                ),
                "model_modified": False,
            }
        ]
    )

    sheets: dict[str, pd.DataFrame] = {
        "Executive Summary": dataframe_for_excel(
            deployment_health_frame
        ),
        "Monitoring Context": dataframe_for_excel(
            context_frame
        ),
        "Health Components": dataframe_for_excel(
            bundle.health_components
        ),
        "Alerts": dataframe_for_excel(
            bundle.alerts
        ),
        "Feature Summary": dataframe_for_excel(
            feature_summary_frame
        ),
        "Feature Drift": dataframe_for_excel(
            bundle.feature_drift
        ),
        "Unmatched Features": dataframe_for_excel(
            bundle.unmatched_features
        ),
        "Prediction Summary": dataframe_for_excel(
            prediction_summary_frame
        ),
        "Prediction Drift": dataframe_for_excel(
            bundle.prediction_drift
        ),
        "Decision Drift": dataframe_for_excel(
            bundle.decision_drift
        ),
        "Data Quality Summary": dataframe_for_excel(
            data_quality_summary_frame
        ),
        "Data Quality": dataframe_for_excel(
            bundle.data_quality_table
        ),
    }

    if HISTORY_CSV_PATH.exists():
        try:
            history = pd.read_csv(
                HISTORY_CSV_PATH
            )

            sheets[
                "Monitoring History"
            ] = dataframe_for_excel(
                history
            )
        except Exception:
            pass

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[
                    :31
                ],
                index=False,
            )

        format_excel_workbook(
            writer,
            {
                sheet_name[
                    :31
                ]: dataframe
                for sheet_name, dataframe in sheets.items()
            },
        )

    shutil.copy2(
        output_path,
        EXCEL_REPORT_PATH,
    )

    return output_path


def require_matplotlib() -> Any:
    try:
        import matplotlib.pyplot as plt
    except ImportError as error:
        raise RuntimeError(
            "Matplotlib is required for monitoring figures."
        ) from error

    return plt


def save_figure(
    figure: Any,
    path: Path,
) -> Path:
    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    figure.savefig(
        path,
        dpi=180,
        bbox_inches="tight",
    )

    return path


def create_health_component_figure(
    bundle: MonitoringBundle,
) -> Path:
    plt = require_matplotlib()

    dataframe = (
        bundle.health_components.sort_values(
            "raw_score",
            ascending=True,
        )
    )

    figure, axis = plt.subplots(
        figsize=(
            9,
            4.8,
        )
    )

    axis.barh(
        dataframe[
            "component_name"
        ].str.replace(
            "_",
            " ",
        ).str.title(),
        dataframe[
            "raw_score"
        ],
    )

    axis.set_xlim(
        0,
        100,
    )

    axis.set_xlabel(
        "Health score"
    )

    axis.set_ylabel(
        "Monitoring component"
    )

    axis.set_title(
        (
            "HeveMind Deployment Health Components\n"
            f"Overall health: "
            f"{bundle.health_assessment.overall_health_score:.2f}/100"
        )
    )

    axis.axvline(
        90,
        linestyle="--",
        linewidth=1,
    )

    axis.axvline(
        80,
        linestyle="--",
        linewidth=1,
    )

    axis.grid(
        axis="x",
        alpha=0.25,
    )

    figure.tight_layout()

    path = path_with_run_suffix(
        FIGURES_DIR
        / "health_components.png",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    save_figure(
        figure,
        path,
    )

    plt.close(
        figure
    )

    shutil.copy2(
        path,
        FIGURES_DIR
        / "health_components.png",
    )

    return path


def create_feature_drift_figure(
    bundle: MonitoringBundle,
    maximum_features: int = 25,
) -> Path:
    plt = require_matplotlib()

    dataframe = bundle.feature_drift.copy()

    dataframe[
        "drift_score"
    ] = (
        100.0
        - pd.to_numeric(
            dataframe[
                "feature_health_score"
            ],
            errors="coerce",
        ).fillna(
            100.0
        )
    )

    dataframe = dataframe.sort_values(
        "drift_score",
        ascending=False,
    ).head(
        maximum_features
    )

    dataframe = dataframe.sort_values(
        "drift_score",
        ascending=True,
    )

    figure, axis = plt.subplots(
        figsize=(
            10,
            max(
                5,
                len(
                    dataframe
                )
                * 0.28,
            ),
        )
    )

    axis.barh(
        dataframe[
            "feature"
        ],
        dataframe[
            "drift_score"
        ],
    )

    axis.set_xlabel(
        "Drift concern score (100 - feature health)"
    )

    axis.set_ylabel(
        "Sensor feature"
    )

    axis.set_title(
        "Highest-Priority Sensor Distribution Shifts"
    )

    axis.grid(
        axis="x",
        alpha=0.25,
    )

    figure.tight_layout()

    path = path_with_run_suffix(
        FIGURES_DIR
        / "feature_drift_priority.png",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    save_figure(
        figure,
        path,
    )

    plt.close(
        figure
    )

    shutil.copy2(
        path,
        FIGURES_DIR
        / "feature_drift_priority.png",
    )

    return path


def create_prediction_shift_figure(
    bundle: MonitoringBundle,
) -> Path:
    plt = require_matplotlib()

    dataframe = bundle.prediction_drift.copy()

    dataframe = dataframe.loc[
        dataframe[
            "reference_mean"
        ].notna()
        & dataframe[
            "current_mean"
        ].notna()
    ].copy()

    if dataframe.empty:
        dataframe = pd.DataFrame(
            {
                "metric_name": [
                    "Unavailable"
                ],
                "reference_mean": [
                    0.0
                ],
                "current_mean": [
                    0.0
                ],
            }
        )

    positions = np.arange(
        len(
            dataframe
        )
    )

    width = 0.36

    figure, axis = plt.subplots(
        figsize=(
            9,
            4.8,
        )
    )

    axis.bar(
        positions
        - width
        / 2,
        dataframe[
            "reference_mean"
        ],
        width=width,
        label="Reference",
    )

    axis.bar(
        positions
        + width
        / 2,
        dataframe[
            "current_mean"
        ],
        width=width,
        label="Current",
    )

    axis.set_xticks(
        positions
    )

    axis.set_xticklabels(
        dataframe[
            "metric_name"
        ].str.replace(
            "_",
            " ",
        ).str.title(),
        rotation=20,
        ha="right",
    )

    axis.set_ylabel(
        "Mean value"
    )

    axis.set_title(
        "Operational Output Distribution Means"
    )

    axis.legend()

    axis.grid(
        axis="y",
        alpha=0.25,
    )

    figure.tight_layout()

    path = path_with_run_suffix(
        FIGURES_DIR
        / "prediction_shift.png",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    save_figure(
        figure,
        path,
    )

    plt.close(
        figure
    )

    shutil.copy2(
        path,
        FIGURES_DIR
        / "prediction_shift.png",
    )

    return path


def create_decision_distribution_figure(
    bundle: MonitoringBundle,
) -> Path:
    plt = require_matplotlib()

    dataframe = bundle.decision_drift.copy()

    dataframe = dataframe.loc[
        dataframe[
            "category"
        ]
        != "Unavailable"
    ].copy()

    if dataframe.empty:
        dataframe = pd.DataFrame(
            {
                "category": [
                    "Unavailable"
                ],
                "reference_share": [
                    0.0
                ],
                "current_share": [
                    0.0
                ],
            }
        )

    positions = np.arange(
        len(
            dataframe
        )
    )

    width = 0.36

    figure, axis = plt.subplots(
        figsize=(
            9,
            4.8,
        )
    )

    axis.bar(
        positions
        - width
        / 2,
        dataframe[
            "reference_share"
        ],
        width=width,
        label="Reference",
    )

    axis.bar(
        positions
        + width
        / 2,
        dataframe[
            "current_share"
        ],
        width=width,
        label="Current",
    )

    axis.set_xticks(
        positions
    )

    axis.set_xticklabels(
        dataframe[
            "category"
        ],
        rotation=20,
        ha="right",
    )

    axis.set_ylim(
        0,
        max(
            1.0,
            float(
                max(
                    dataframe[
                        "reference_share"
                    ].max(),
                    dataframe[
                        "current_share"
                    ].max(),
                )
            )
            * 1.15,
        ),
    )

    axis.set_ylabel(
        "Proportion"
    )

    axis.set_title(
        "Operational Decision Distribution"
    )

    axis.legend()

    axis.grid(
        axis="y",
        alpha=0.25,
    )

    figure.tight_layout()

    path = path_with_run_suffix(
        FIGURES_DIR
        / "decision_distribution.png",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    save_figure(
        figure,
        path,
    )

    plt.close(
        figure
    )

    shutil.copy2(
        path,
        FIGURES_DIR
        / "decision_distribution.png",
    )

    return path


def create_alert_distribution_figure(
    bundle: MonitoringBundle,
) -> Path:
    plt = require_matplotlib()

    severity_order = [
        "Watch",
        "Warning",
        "Critical",
    ]

    counts = [
        count_severity(
            bundle.alerts,
            "severity",
            severity,
        )
        for severity in severity_order
    ]

    figure, axis = plt.subplots(
        figsize=(
            7.5,
            4.5,
        )
    )

    axis.bar(
        severity_order,
        counts,
    )

    axis.set_ylabel(
        "Alert count"
    )

    axis.set_title(
        "Monitoring Alerts by Severity"
    )

    axis.set_ylim(
        0,
        max(
            counts
            + [
                1
            ]
        )
        * 1.2,
    )

    axis.grid(
        axis="y",
        alpha=0.25,
    )

    for index, count in enumerate(
        counts
    ):
        axis.text(
            index,
            count,
            str(
                count
            ),
            ha="center",
            va="bottom",
        )

    figure.tight_layout()

    path = path_with_run_suffix(
        FIGURES_DIR
        / "alert_distribution.png",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    save_figure(
        figure,
        path,
    )

    plt.close(
        figure
    )

    shutil.copy2(
        path,
        FIGURES_DIR
        / "alert_distribution.png",
    )

    return path


def create_monitoring_history_figure(
    bundle: MonitoringBundle,
) -> Path | None:
    if not HISTORY_CSV_PATH.exists():
        return None

    try:
        history = pd.read_csv(
            HISTORY_CSV_PATH
        )
    except Exception:
        return None

    if (
        history.empty
        or "generated_utc"
        not in history.columns
        or "overall_health_score"
        not in history.columns
    ):
        return None

    history[
        "generated_utc"
    ] = pd.to_datetime(
        history[
            "generated_utc"
        ],
        errors="coerce",
        utc=True,
    )

    history[
        "overall_health_score"
    ] = pd.to_numeric(
        history[
            "overall_health_score"
        ],
        errors="coerce",
    )

    history = history.dropna(
        subset=[
            "generated_utc",
            "overall_health_score",
        ]
    ).sort_values(
        "generated_utc"
    )

    if history.empty:
        return None

    plt = require_matplotlib()

    figure, axis = plt.subplots(
        figsize=(
            9,
            4.6,
        )
    )

    axis.plot(
        history[
            "generated_utc"
        ],
        history[
            "overall_health_score"
        ],
        marker="o",
    )

    axis.axhline(
        90,
        linestyle="--",
        linewidth=1,
    )

    axis.axhline(
        80,
        linestyle="--",
        linewidth=1,
    )

    axis.axhline(
        65,
        linestyle="--",
        linewidth=1,
    )

    axis.set_ylim(
        0,
        100,
    )

    axis.set_ylabel(
        "Overall health score"
    )

    axis.set_xlabel(
        "Monitoring run"
    )

    axis.set_title(
        "Deployment Health Trend"
    )

    axis.grid(
        alpha=0.25,
    )

    figure.autofmt_xdate()

    figure.tight_layout()

    path = path_with_run_suffix(
        FIGURES_DIR
        / "monitoring_history.png",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    save_figure(
        figure,
        path,
    )

    plt.close(
        figure
    )

    shutil.copy2(
        path,
        FIGURES_DIR
        / "monitoring_history.png",
    )

    return path


def create_all_monitoring_figures(
    bundle: MonitoringBundle,
) -> tuple[Path, ...]:
    figure_paths: list[Path] = []

    figure_builders = [
        create_health_component_figure,
        create_feature_drift_figure,
        create_prediction_shift_figure,
        create_decision_distribution_figure,
        create_alert_distribution_figure,
        create_monitoring_history_figure,
    ]

    for builder in figure_builders:
        try:
            path = builder(
                bundle
            )
        except Exception as error:
            bundle.context.warnings.append(
                f"Figure generation failed for {builder.__name__}: {error}"
            )

            continue

        if path is not None:
            figure_paths.append(
                path
            )

    return tuple(
        figure_paths
    )


def require_reportlab() -> dict[str, Any]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            BaseDocTemplate,
            Frame,
            Image,
            PageBreak,
            PageTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as error:
        raise RuntimeError(
            "ReportLab is required for PDF monitoring reports."
        ) from error

    return {
        "colors": colors,
        "TA_CENTER": TA_CENTER,
        "TA_LEFT": TA_LEFT,
        "A4": A4,
        "ParagraphStyle": ParagraphStyle,
        "getSampleStyleSheet": getSampleStyleSheet,
        "mm": mm,
        "BaseDocTemplate": BaseDocTemplate,
        "Frame": Frame,
        "Image": Image,
        "PageBreak": PageBreak,
        "PageTemplate": PageTemplate,
        "Paragraph": Paragraph,
        "Spacer": Spacer,
        "Table": Table,
        "TableStyle": TableStyle,
    }


def pdf_paragraph(
    text: Any,
    style: Any,
    reportlab: Mapping[str, Any],
) -> Any:
    Paragraph = reportlab[
        "Paragraph"
    ]

    return Paragraph(
        safe_text(
            text
        ),
        style,
    )


def pdf_table(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    widths: Sequence[float],
    styles: Mapping[str, Any],
    reportlab: Mapping[str, Any],
) -> Any:
    Table = reportlab[
        "Table"
    ]

    TableStyle = reportlab[
        "TableStyle"
    ]

    colors = reportlab[
        "colors"
    ]

    data = [
        [
            pdf_paragraph(
                header,
                styles[
                    "table_header"
                ],
                reportlab,
            )
            for header in headers
        ]
    ]

    for row in rows:
        data.append(
            [
                pdf_paragraph(
                    value,
                    styles[
                        "table_body"
                    ],
                    reportlab,
                )
                for value in row
            ]
        )

    table = Table(
        data,
        colWidths=list(
            widths
        ),
        repeatRows=1,
    )

    commands = [
        (
            "BACKGROUND",
            (
                0,
                0,
            ),
            (
                -1,
                0,
            ),
            colors.HexColor(
                "#123252"
            ),
        ),
        (
            "TEXTCOLOR",
            (
                0,
                0,
            ),
            (
                -1,
                0,
            ),
            colors.white,
        ),
        (
            "GRID",
            (
                0,
                0,
            ),
            (
                -1,
                -1,
            ),
            0.4,
            colors.HexColor(
                "#DCE3E9"
            ),
        ),
        (
            "VALIGN",
            (
                0,
                0,
            ),
            (
                -1,
                -1,
            ),
            "TOP",
        ),
        (
            "LEFTPADDING",
            (
                0,
                0,
            ),
            (
                -1,
                -1,
            ),
            4,
        ),
        (
            "RIGHTPADDING",
            (
                0,
                0,
            ),
            (
                -1,
                -1,
            ),
            4,
        ),
        (
            "TOPPADDING",
            (
                0,
                0,
            ),
            (
                -1,
                -1,
            ),
            4,
        ),
        (
            "BOTTOMPADDING",
            (
                0,
                0,
            ),
            (
                -1,
                -1,
            ),
            4,
        ),
    ]

    for row_index in range(
        1,
        len(
            data
        ),
    ):
        if row_index % 2 == 0:
            commands.append(
                (
                    "BACKGROUND",
                    (
                        0,
                        row_index,
                    ),
                    (
                        -1,
                        row_index,
                    ),
                    colors.HexColor(
                        "#EEF3F7"
                    ),
                )
            )

    table.setStyle(
        TableStyle(
            commands
        )
    )

    return table


def build_pdf_styles(
    reportlab: Mapping[str, Any],
) -> dict[str, Any]:
    ParagraphStyle = reportlab[
        "ParagraphStyle"
    ]

    getSampleStyleSheet = reportlab[
        "getSampleStyleSheet"
    ]

    colors = reportlab[
        "colors"
    ]

    TA_CENTER = reportlab[
        "TA_CENTER"
    ]

    sample = getSampleStyleSheet()

    return {
        "title": ParagraphStyle(
            "MonitoringTitle",
            parent=sample[
                "Title"
            ],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=27,
            textColor=colors.HexColor(
                "#071426"
            ),
            spaceAfter=8,
        ),
        "subtitle": ParagraphStyle(
            "MonitoringSubtitle",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor(
                "#405267"
            ),
            spaceAfter=8,
        ),
        "section": ParagraphStyle(
            "MonitoringSection",
            parent=sample[
                "Heading2"
            ],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=17,
            textColor=colors.HexColor(
                "#071426"
            ),
            spaceBefore=7,
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "MonitoringBody",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11.5,
            textColor=colors.HexColor(
                "#405267"
            ),
            spaceAfter=5,
        ),
        "metric_label": ParagraphStyle(
            "MonitoringMetricLabel",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica-Bold",
            fontSize=6.5,
            leading=8,
            textColor=colors.HexColor(
                "#708196"
            ),
        ),
        "metric_value": ParagraphStyle(
            "MonitoringMetricValue",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor(
                "#071426"
            ),
        ),
        "table_header": ParagraphStyle(
            "MonitoringTableHeader",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica-Bold",
            fontSize=6.5,
            leading=8,
            textColor=colors.white,
        ),
        "table_body": ParagraphStyle(
            "MonitoringTableBody",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica",
            fontSize=6.3,
            leading=8,
            textColor=colors.HexColor(
                "#405267"
            ),
        ),
        "footer": ParagraphStyle(
            "MonitoringFooter",
            parent=sample[
                "BodyText"
            ],
            fontName="Helvetica",
            fontSize=6,
            leading=7,
            textColor=colors.HexColor(
                "#708196"
            ),
            alignment=TA_CENTER,
        ),
    }


def export_monitoring_pdf(
    bundle: MonitoringBundle,
    figure_paths: Sequence[Path],
) -> Path:
    reportlab = require_reportlab()

    BaseDocTemplate = reportlab[
        "BaseDocTemplate"
    ]

    Frame = reportlab[
        "Frame"
    ]

    PageTemplate = reportlab[
        "PageTemplate"
    ]

    Image = reportlab[
        "Image"
    ]

    PageBreak = reportlab[
        "PageBreak"
    ]

    Spacer = reportlab[
        "Spacer"
    ]

    Table = reportlab[
        "Table"
    ]

    TableStyle = reportlab[
        "TableStyle"
    ]

    colors = reportlab[
        "colors"
    ]

    A4 = reportlab[
        "A4"
    ]

    mm = reportlab[
        "mm"
    ]

    styles = build_pdf_styles(
        reportlab
    )

    output_path = path_with_run_suffix(
        PDF_REPORT_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    document = BaseDocTemplate(
        str(
            output_path
        ),
        pagesize=A4,
        leftMargin=16
        * mm,
        rightMargin=16
        * mm,
        topMargin=15
        * mm,
        bottomMargin=17
        * mm,
        title=(
            "HeveMind Model Monitoring Report"
        ),
        author=(
            "HeveMind Monitoring Engine"
        ),
        subject=(
            "Semiconductor deployment drift and health assessment"
        ),
    )

    frame = Frame(
        document.leftMargin,
        document.bottomMargin,
        document.width,
        document.height,
        id="normal",
    )

    def draw_page(
        canvas: Any,
        doc: Any,
    ) -> None:
        canvas.saveState()

        page_width, _ = A4

        canvas.setStrokeColor(
            colors.HexColor(
                "#DCE3E9"
            )
        )

        canvas.line(
            16
            * mm,
            13
            * mm,
            page_width
            - 16
            * mm,
            13
            * mm,
        )

        canvas.setFont(
            "Helvetica",
            6.5,
        )

        canvas.setFillColor(
            colors.HexColor(
                "#708196"
            )
        )

        canvas.drawString(
            16
            * mm,
            8
            * mm,
            (
                f"Run {bundle.context.run_identity.run_id} | "
                f"Model modified: No"
            ),
        )

        canvas.drawRightString(
            page_width
            - 16
            * mm,
            8
            * mm,
            f"Page {doc.page}",
        )

        canvas.restoreState()

    document.addPageTemplates(
        [
            PageTemplate(
                id="monitoring",
                frames=[
                    frame
                ],
                onPage=draw_page,
            )
        ]
    )

    story: list[Any] = []

    story.append(
        pdf_paragraph(
            "HeveMind Model Monitoring Report",
            styles[
                "title"
            ],
            reportlab,
        )
    )

    story.append(
        pdf_paragraph(
            (
                "Industrial monitoring of sensor distributions, prediction "
                "outputs, decision patterns, confidence, uncertainty and "
                "data quality. This module does not change the deployed model."
            ),
            styles[
                "subtitle"
            ],
            reportlab,
        )
    )

    health = bundle.health_assessment

    metrics = [
        (
            "Overall health",
            f"{health.overall_health_score:.2f}/100",
        ),
        (
            "Health band",
            health.overall_health_band,
        ),
        (
            "Recommendation",
            health.recommendation,
        ),
        (
            "Critical alerts",
            str(
                health.critical_alerts
            ),
        ),
        (
            "Model modified",
            "No",
        ),
    ]

    metric_cells = []

    for label, value in metrics:
        metric_cells.append(
            Table(
                [
                    [
                        pdf_paragraph(
                            label.upper(),
                            styles[
                                "metric_label"
                            ],
                            reportlab,
                        )
                    ],
                    [
                        pdf_paragraph(
                            value,
                            styles[
                                "metric_value"
                            ],
                            reportlab,
                        )
                    ],
                ],
                colWidths=[
                    31
                    * mm
                ],
            )
        )

    metric_table = Table(
        [
            metric_cells
        ],
        colWidths=[
            34
            * mm
        ]
        * len(
            metric_cells
        ),
    )

    metric_table.setStyle(
        TableStyle(
            [
                (
                    "BOX",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    0.6,
                    colors.HexColor(
                        "#DCE3E9"
                    ),
                ),
                (
                    "INNERGRID",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    0.4,
                    colors.HexColor(
                        "#DCE3E9"
                    ),
                ),
                (
                    "BACKGROUND",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    colors.white,
                ),
                (
                    "VALIGN",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    5,
                ),
                (
                    "RIGHTPADDING",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    5,
                ),
                (
                    "TOPPADDING",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    5,
                ),
                (
                    "BOTTOMPADDING",
                    (
                        0,
                        0,
                    ),
                    (
                        -1,
                        -1,
                    ),
                    5,
                ),
            ]
        )
    )

    story.append(
        metric_table
    )

    story.append(
        Spacer(
            1,
            5
            * mm,
        )
    )

    story.append(
        pdf_paragraph(
            "Executive Assessment",
            styles[
                "section"
            ],
            reportlab,
        )
    )

    story.append(
        pdf_paragraph(
            health.summary_statement,
            styles[
                "body"
            ],
            reportlab,
        )
    )

    story.append(
        pdf_paragraph(
            (
                f"Production status: {health.production_monitoring_status}. "
                f"Reason: {health.recommendation_reason}"
            ),
            styles[
                "body"
            ],
            reportlab,
        )
    )

    component_rows = []

    for row in bundle.health_components.to_dict(
        orient="records"
    ):
        component_rows.append(
            [
                safe_text(
                    row.get(
                        "component_name"
                    )
                ).replace(
                    "_",
                    " ",
                ).title(),
                f"{safe_float(row.get('raw_score'), 0.0):.2f}",
                f"{safe_float(row.get('weight'), 0.0):.2f}",
                safe_text(
                    row.get(
                        "severity"
                    )
                ),
                safe_text(
                    row.get(
                        "explanation"
                    )
                ),
            ]
        )

    story.append(
        pdf_paragraph(
            "Health Components",
            styles[
                "section"
            ],
            reportlab,
        )
    )

    story.append(
        pdf_table(
            headers=[
                "Component",
                "Score",
                "Weight",
                "Severity",
                "Explanation",
            ],
            rows=component_rows,
            widths=[
                34
                * mm,
                20
                * mm,
                18
                * mm,
                22
                * mm,
                76
                * mm,
            ],
            styles=styles,
            reportlab=reportlab,
        )
    )

    for figure_path in figure_paths[
        :3
    ]:
        if figure_path.exists():
            story.append(
                Spacer(
                    1,
                    5
                    * mm,
                )
            )

            story.append(
                Image(
                    str(
                        figure_path
                    ),
                    width=165
                    * mm,
                    height=82
                    * mm,
                )
            )

    story.append(
        PageBreak()
    )

    story.append(
        pdf_paragraph(
            "Priority Monitoring Alerts",
            styles[
                "section"
            ],
            reportlab,
        )
    )

    if bundle.alerts.empty:
        story.append(
            pdf_paragraph(
                "No Watch, Warning or Critical alerts were generated.",
                styles[
                    "body"
                ],
                reportlab,
            )
        )
    else:
        alert_rows = []

        for row in bundle.alerts.head(
            30
        ).to_dict(
            orient="records"
        ):
            alert_rows.append(
                [
                    safe_text(
                        row.get(
                            "severity"
                        )
                    ),
                    safe_text(
                        row.get(
                            "category"
                        )
                    ),
                    safe_text(
                        row.get(
                            "resource_id"
                        )
                    ),
                    safe_text(
                        row.get(
                            "message"
                        )
                    ),
                    safe_text(
                        row.get(
                            "recommended_action"
                        )
                    ),
                ]
            )

        story.append(
            pdf_table(
                headers=[
                    "Severity",
                    "Category",
                    "Resource",
                    "Alert",
                    "Recommended action",
                ],
                rows=alert_rows,
                widths=[
                    18
                    * mm,
                    26
                    * mm,
                    27
                    * mm,
                    50
                    * mm,
                    49
                    * mm,
                ],
                styles=styles,
                reportlab=reportlab,
            )
        )

    story.append(
        pdf_paragraph(
            "Highest-Priority Feature Drift",
            styles[
                "section"
            ],
            reportlab,
        )
    )

    feature_rows = []

    for row in bundle.feature_drift.head(
        25
    ).to_dict(
        orient="records"
    ):
        feature_rows.append(
            [
                safe_text(
                    row.get(
                        "feature"
                    )
                ),
                safe_text(
                    row.get(
                        "overall_severity"
                    )
                ),
                f"{safe_float(row.get('feature_health_score'), 0.0):.2f}",
                (
                    f"{safe_float(row.get('psi'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "psi"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                (
                    f"{safe_float(row.get('ks_statistic'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "ks_statistic"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                safe_text(
                    row.get(
                        "notes"
                    )
                ),
            ]
        )

    story.append(
        pdf_table(
            headers=[
                "Feature",
                "Severity",
                "Health",
                "PSI",
                "KS",
                "Notes",
            ],
            rows=feature_rows,
            widths=[
                27
                * mm,
                20
                * mm,
                16
                * mm,
                17
                * mm,
                17
                * mm,
                73
                * mm,
            ],
            styles=styles,
            reportlab=reportlab,
        )
    )

    story.append(
        PageBreak()
    )

    story.append(
        pdf_paragraph(
            "Prediction and Decision Monitoring",
            styles[
                "section"
            ],
            reportlab,
        )
    )

    prediction_rows = []

    for row in bundle.prediction_drift.to_dict(
        orient="records"
    ):
        prediction_rows.append(
            [
                safe_text(
                    row.get(
                        "metric_name"
                    )
                ).replace(
                    "_",
                    " ",
                ).title(),
                safe_text(
                    row.get(
                        "severity"
                    )
                ),
                (
                    f"{safe_float(row.get('reference_mean'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "reference_mean"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                (
                    f"{safe_float(row.get('current_mean'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "current_mean"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                (
                    f"{safe_float(row.get('psi'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "psi"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                f"{safe_float(row.get('health_score'), 0.0):.2f}",
            ]
        )

    story.append(
        pdf_table(
            headers=[
                "Output",
                "Severity",
                "Reference mean",
                "Current mean",
                "PSI",
                "Health",
            ],
            rows=prediction_rows,
            widths=[
                43
                * mm,
                25
                * mm,
                27
                * mm,
                27
                * mm,
                22
                * mm,
                22
                * mm,
            ],
            styles=styles,
            reportlab=reportlab,
        )
    )

    story.append(
        pdf_paragraph(
            "Data Quality",
            styles[
                "section"
            ],
            reportlab,
        )
    )

    quality_rows = []

    for row in bundle.data_quality_table.to_dict(
        orient="records"
    ):
        quality_rows.append(
            [
                safe_text(
                    row.get(
                        "metric"
                    )
                ),
                (
                    f"{safe_float(row.get('value'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "value"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                (
                    f"{safe_float(row.get('reference_value'), 0.0):.4f}"
                    if safe_float(
                        row.get(
                            "reference_value"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                safe_text(
                    row.get(
                        "severity"
                    )
                ),
            ]
        )

    story.append(
        pdf_table(
            headers=[
                "Metric",
                "Current",
                "Reference",
                "Severity",
            ],
            rows=quality_rows,
            widths=[
                75
                * mm,
                30
                * mm,
                30
                * mm,
                30
                * mm,
            ],
            styles=styles,
            reportlab=reportlab,
        )
    )

    story.append(
        Spacer(
            1,
            6
            * mm,
        )
    )

    story.append(
        pdf_paragraph(
            (
                "Interpretation limitation: monitoring detects statistical "
                "changes and operational risk signals. It does not establish "
                "a physical failure mechanism and does not automatically "
                "replace or retrain the deployed model."
            ),
            styles[
                "body"
            ],
            reportlab,
        )
    )

    document.build(
        story
    )

    shutil.copy2(
        output_path,
        PDF_REPORT_PATH,
    )

    return output_path


def validate_report_outputs(
    *,
    excel_path: Path,
    pdf_path: Path | None,
    figure_paths: Sequence[Path],
) -> None:
    if not excel_path.exists():
        raise RuntimeError(
            "Monitoring Excel report was not generated."
        )

    if excel_path.stat().st_size <= 0:
        raise RuntimeError(
            "Monitoring Excel report is empty."
        )

    if pdf_path is not None:
        if not pdf_path.exists():
            raise RuntimeError(
                "Monitoring PDF report was not generated."
            )

        if pdf_path.stat().st_size <= 0:
            raise RuntimeError(
                "Monitoring PDF report is empty."
            )

    for path in figure_paths:
        if not path.exists():
            raise RuntimeError(
                f"Monitoring figure was not generated: {path}"
            )

        if path.stat().st_size <= 0:
            raise RuntimeError(
                f"Monitoring figure is empty: {path}"
            )


def file_sha256_or_none(
    path: Path | None,
) -> str | None:
    if path is None or not path.exists():
        return None

    return calculate_file_sha256(
        path
    )


def create_monitoring_report_artifacts(
    bundle: MonitoringBundle,
    *,
    generate_pdf: bool = True,
) -> MonitoringReportArtifacts:
    ensure_directories()

    (
        archival_summary_path,
        archival_health_path,
    ) = save_consolidated_json_reports(
        bundle
    )

    feature_drift_path = path_with_run_suffix(
        FEATURE_DRIFT_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    prediction_drift_path = path_with_run_suffix(
        PREDICTION_DRIFT_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    decision_drift_path = path_with_run_suffix(
        TABLES_DIR
        / "decision_drift.csv",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    data_quality_path = path_with_run_suffix(
        DATA_QUALITY_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    alerts_path = path_with_run_suffix(
        ALERTS_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    health_components_path = path_with_run_suffix(
        HEALTH_COMPONENTS_PATH,
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    bundle.feature_drift.to_csv(
        feature_drift_path,
        index=False,
    )

    bundle.prediction_drift.to_csv(
        prediction_drift_path,
        index=False,
    )

    bundle.decision_drift.to_csv(
        decision_drift_path,
        index=False,
    )

    bundle.data_quality_table.to_csv(
        data_quality_path,
        index=False,
    )

    bundle.alerts.to_csv(
        alerts_path,
        index=False,
    )

    bundle.health_components.to_csv(
        health_components_path,
        index=False,
    )

    shutil.copy2(
        feature_drift_path,
        FEATURE_DRIFT_PATH,
    )

    shutil.copy2(
        prediction_drift_path,
        PREDICTION_DRIFT_PATH,
    )

    shutil.copy2(
        decision_drift_path,
        TABLES_DIR
        / "decision_drift.csv",
    )

    shutil.copy2(
        data_quality_path,
        DATA_QUALITY_PATH,
    )

    shutil.copy2(
        alerts_path,
        ALERTS_PATH,
    )

    shutil.copy2(
        health_components_path,
        HEALTH_COMPONENTS_PATH,
    )

    figure_paths = create_all_monitoring_figures(
        bundle
    )

    excel_path = export_monitoring_excel(
        bundle
    )

    pdf_path: Path | None = None

    if generate_pdf:
        try:
            pdf_path = export_monitoring_pdf(
                bundle,
                figure_paths,
            )
        except RuntimeError as error:
            bundle.context.warnings.append(
                str(
                    error
                )
            )

    validate_report_outputs(
        excel_path=excel_path,
        pdf_path=pdf_path,
        figure_paths=figure_paths,
    )

    report_sha256 = file_sha256_or_none(
        pdf_path
        or excel_path
    )

    manifest_path = path_with_run_suffix(
        EXPORTS_DIR
        / "monitoring_report_manifest.json",
        bundle.context.run_identity.run_id,
        bundle.context.run_identity.generated_utc,
    )

    atomic_write_json(
        manifest_path,
        {
            "run_id": (
                bundle.context.run_identity.run_id
            ),
            "generated_utc": (
                bundle.context.run_identity.generated_utc
            ),
            "drift_summary_json": str(
                archival_summary_path
            ),
            "deployment_health_json": str(
                archival_health_path
            ),
            "feature_drift_csv": str(
                feature_drift_path
            ),
            "prediction_drift_csv": str(
                prediction_drift_path
            ),
            "decision_drift_csv": str(
                decision_drift_path
            ),
            "data_quality_csv": str(
                data_quality_path
            ),
            "alerts_csv": str(
                alerts_path
            ),
            "health_components_csv": str(
                health_components_path
            ),
            "monitoring_excel": str(
                excel_path
            ),
            "monitoring_pdf": (
                str(
                    pdf_path
                )
                if pdf_path is not None
                else None
            ),
            "figure_paths": [
                str(
                    path
                )
                for path in figure_paths
            ],
            "report_sha256": report_sha256,
            "model_modification_performed": False,
        },
    )

    record_monitoring_event(
        event_type="model_monitoring_report_generated",
        outcome="success",
        action="export_monitoring_report",
        resource_id=(
            bundle.context.run_identity.run_id
        ),
        details={
            "excel_path": str(
                excel_path
            ),
            "pdf_path": (
                str(
                    pdf_path
                )
                if pdf_path is not None
                else None
            ),
            "report_sha256": (
                report_sha256
            ),
            "figure_count": len(
                figure_paths
            ),
            "model_modification_performed": False,
        },
    )

    return MonitoringReportArtifacts(
        run_id=bundle.context.run_identity.run_id,
        generated_utc=(
            bundle.context.run_identity.generated_utc
        ),
        drift_summary_json=archival_summary_path,
        deployment_health_json=archival_health_path,
        feature_drift_csv=feature_drift_path,
        prediction_drift_csv=prediction_drift_path,
        decision_drift_csv=decision_drift_path,
        data_quality_csv=data_quality_path,
        alerts_csv=alerts_path,
        health_components_csv=health_components_path,
        monitoring_excel=excel_path,
        monitoring_pdf=pdf_path,
        figure_paths=figure_paths,
        report_sha256=report_sha256,
    )


def build_report_console_summary(
    artifacts: MonitoringReportArtifacts,
) -> str:
    lines = [
        "="
        * 118,
        "HEVEMIND MODEL MONITORING REPORTS",
        "="
        * 118,
        "",
        f"Run ID:                          {artifacts.run_id}",
        f"Drift summary JSON:              {artifacts.drift_summary_json}",
        f"Deployment health JSON:          {artifacts.deployment_health_json}",
        f"Feature drift CSV:               {artifacts.feature_drift_csv}",
        f"Prediction drift CSV:            {artifacts.prediction_drift_csv}",
        f"Decision drift CSV:              {artifacts.decision_drift_csv}",
        f"Data quality CSV:                {artifacts.data_quality_csv}",
        f"Alerts CSV:                      {artifacts.alerts_csv}",
        f"Health components CSV:           {artifacts.health_components_csv}",
        f"Excel report:                    {artifacts.monitoring_excel}",
        f"PDF report:                      {artifacts.monitoring_pdf}",
        f"Figures generated:               {len(artifacts.figure_paths)}",
        f"Report SHA-256:                  {artifacts.report_sha256}",
    ]

    return "\n".join(
        lines
    )


# End of Part 5.
# Append Part 6 immediately below this line.


# ============================================================
# PART 6 — STREAMLIT DASHBOARD, ORCHESTRATION, CLI AND MAIN
# ============================================================

import argparse
import shutil


MONITORING_DASHBOARD_CSS = """
<style>
    :root {
        --navy-950: #071426;
        --navy-900: #0b1f36;
        --navy-800: #123252;
        --blue-600: #1f6aa5;
        --blue-100: #e8f2f9;
        --slate-900: #172230;
        --slate-700: #405267;
        --slate-500: #708196;
        --slate-300: #c7d2dc;
        --slate-200: #dce3e9;
        --slate-100: #eef3f7;
        --white: #ffffff;
        --green: #1b7f5b;
        --green-bg: #e6f4ed;
        --amber: #a96b00;
        --amber-bg: #fff3d9;
        --red: #b43b3b;
        --red-bg: #fdeaea;
        --purple: #6f4f9c;
        --purple-bg: #f1ebf9;
    }

    .stApp {
        background:
            linear-gradient(
                180deg,
                #f5f8fb 0%,
                #edf3f7 100%
            );
    }

    .block-container {
        max-width: 1480px;
        padding-top: 1.5rem;
        padding-bottom: 3rem;
    }

    .monitoring-header {
        background:
            linear-gradient(
                110deg,
                var(--navy-950) 0%,
                var(--navy-800) 72%,
                var(--blue-600) 100%
            );
        border-radius: 18px;
        padding: 1.25rem 1.45rem;
        color: #ffffff;
        box-shadow:
            0 14px 34px
            rgba(7, 20, 38, 0.17);
        margin-bottom: 1rem;
    }

    .monitoring-title {
        font-size: 1.82rem;
        font-weight: 790;
        letter-spacing: -0.025em;
    }

    .monitoring-subtitle {
        color: #d8e8f4;
        font-size: 0.9rem;
        margin-top: 0.36rem;
        line-height: 1.5;
    }

    .monitoring-chip-row {
        display: flex;
        gap: 0.5rem;
        flex-wrap: wrap;
        margin-top: 0.9rem;
    }

    .monitoring-chip {
        border-radius: 999px;
        padding: 0.32rem 0.68rem;
        font-size: 0.72rem;
        font-weight: 720;
        background: rgba(255, 255, 255, 0.10);
        border: 1px solid rgba(255, 255, 255, 0.14);
        color: #ffffff;
    }

    .metric-card {
        min-height: 132px;
        height: 100%;
        background: #ffffff;
        border: 1px solid var(--slate-200);
        border-radius: 15px;
        padding: 0.95rem 1rem;
        box-shadow:
            0 7px 20px
            rgba(21, 48, 75, 0.06);
    }

    .metric-label {
        color: var(--slate-500);
        font-size: 0.70rem;
        font-weight: 760;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    .metric-value {
        color: var(--navy-950);
        font-size: 1.55rem;
        font-weight: 790;
        margin-top: 0.42rem;
    }

    .metric-note {
        color: var(--slate-700);
        font-size: 0.77rem;
        line-height: 1.4;
        margin-top: 0.36rem;
    }

    .status-panel {
        border-radius: 14px;
        padding: 0.95rem 1rem;
        background: #ffffff;
        border: 1px solid var(--slate-200);
        box-shadow:
            0 6px 18px
            rgba(21, 48, 75, 0.05);
        margin-bottom: 0.75rem;
    }

    .status-heading {
        font-size: 0.74rem;
        font-weight: 760;
        text-transform: uppercase;
        letter-spacing: 0.045em;
        color: var(--slate-500);
    }

    .status-value {
        margin-top: 0.35rem;
        font-size: 1.08rem;
        font-weight: 760;
        color: var(--navy-950);
    }

    .status-note {
        margin-top: 0.32rem;
        color: var(--slate-700);
        font-size: 0.78rem;
        line-height: 1.4;
    }

    .normal-callout {
        border-left: 4px solid var(--green);
        background: var(--green-bg);
        border-radius: 10px;
        padding: 0.8rem 0.95rem;
        color: #245a47;
        font-size: 0.82rem;
        line-height: 1.45;
        margin-top: 0.6rem;
    }

    .watch-callout {
        border-left: 4px solid var(--blue-600);
        background: var(--blue-100);
        border-radius: 10px;
        padding: 0.8rem 0.95rem;
        color: var(--slate-700);
        font-size: 0.82rem;
        line-height: 1.45;
        margin-top: 0.6rem;
    }

    .warning-callout {
        border-left: 4px solid var(--amber);
        background: var(--amber-bg);
        border-radius: 10px;
        padding: 0.8rem 0.95rem;
        color: #6b4c11;
        font-size: 0.82rem;
        line-height: 1.45;
        margin-top: 0.6rem;
    }

    .critical-callout {
        border-left: 4px solid var(--red);
        background: var(--red-bg);
        border-radius: 10px;
        padding: 0.8rem 0.95rem;
        color: #7a2f2f;
        font-size: 0.82rem;
        line-height: 1.45;
        margin-top: 0.6rem;
    }

    .section-note {
        border-left: 4px solid var(--blue-600);
        background: #eef7fc;
        border-radius: 10px;
        padding: 0.78rem 0.9rem;
        color: var(--slate-700);
        font-size: 0.79rem;
        line-height: 1.45;
        margin: 0.65rem 0;
    }

    .stTabs [aria-selected="true"] {
        background: #e8f2f9 !important;
        color: var(--navy-950) !important;
        border-bottom: 3px solid var(--blue-600) !important;
    }

    .stTabs [aria-selected="true"] p {
        color: var(--navy-950) !important;
        font-weight: 720 !important;
    }

    [data-testid="stDataFrame"] {
        border: 1px solid var(--slate-200);
        border-radius: 12px;
        overflow: hidden;
        background: #ffffff;
    }

    [data-testid="stSidebar"] {
        background:
            linear-gradient(
                180deg,
                var(--navy-950) 0%,
                var(--navy-900) 100%
            );
    }

    [data-testid="stSidebar"] * {
        color: #ffffff;
    }

    [data-testid="stSidebar"] input {
        color: var(--navy-950) !important;
        background: #ffffff !important;
    }

    footer {
        visibility: hidden;
    }
</style>
"""


# ============================================================
# COMPLETE MONITORING ORCHESTRATION
# ============================================================
def run_complete_monitoring_pipeline(
    *,
    reference_path: Path | None = None,
    current_path: Path | None = None,
    config_path: Path | None = None,
    save_output: bool = True,
    persist_history: bool = True,
    generate_reports: bool = True,
    generate_pdf: bool = True,
) -> tuple[
    MonitoringBundle,
    MonitoringReportArtifacts | None,
]:
    context = build_monitoring_context(
        reference_path=reference_path,
        current_path=current_path,
        config_path=config_path,
    )

    (
        feature_drift,
        feature_summary,
        unmatched_features,
    ) = run_feature_drift_engine(
        context,
        save_output=save_output,
    )

    (
        prediction_drift,
        decision_metric,
        decision_drift,
        data_quality_summary,
        data_quality_table,
        prediction_summary,
    ) = run_prediction_and_quality_engine(
        context,
        save_output=save_output,
    )

    (
        alerts,
        health_components,
        health_assessment,
    ) = run_health_and_alert_engine(
        context=context,
        feature_drift=feature_drift,
        feature_summary=feature_summary,
        unmatched_features=unmatched_features,
        prediction_drift=prediction_drift,
        decision_metric=decision_metric,
        data_quality_summary=data_quality_summary,
        data_quality_table=data_quality_table,
        save_output=save_output,
        persist_history=persist_history,
    )

    bundle = MonitoringBundle(
        context=context,
        feature_drift=feature_drift,
        feature_summary=feature_summary,
        unmatched_features=unmatched_features,
        prediction_drift=prediction_drift,
        decision_metric=decision_metric,
        decision_drift=decision_drift,
        data_quality_summary=data_quality_summary,
        data_quality_table=data_quality_table,
        prediction_summary=prediction_summary,
        alerts=alerts,
        health_components=health_components,
        health_assessment=health_assessment,
    )

    artifacts: MonitoringReportArtifacts | None = None

    if generate_reports:
        artifacts = create_monitoring_report_artifacts(
            bundle,
            generate_pdf=generate_pdf,
        )

    return (
        bundle,
        artifacts,
    )


def monitoring_pipeline_console_output(
    bundle: MonitoringBundle,
    artifacts: MonitoringReportArtifacts | None,
) -> str:
    sections = [
        build_feature_drift_console_summary(
            bundle.feature_summary
        ),
        build_prediction_monitoring_console_summary(
            bundle.prediction_summary,
            bundle.decision_metric,
            bundle.data_quality_summary,
        ),
        build_health_console_summary(
            bundle.health_assessment
        ),
    ]

    if artifacts is not None:
        sections.append(
            build_report_console_summary(
                artifacts
            )
        )

    return "\n\n".join(
        sections
    )


# ============================================================
# LOADING PREVIOUSLY GENERATED MONITORING OUTPUTS
# ============================================================
def read_csv_optional(
    path: Path,
    required_columns: Sequence[str] | None = None,
) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()

    dataframe = pd.read_csv(
        path,
        low_memory=False,
    )

    if required_columns:
        missing = [
            column
            for column in required_columns
            if column not in dataframe.columns
        ]

        if missing:
            raise RuntimeError(
                f"{path} is missing required columns: "
                + ", ".join(
                    missing
                )
            )

    return dataframe


def read_json_optional(
    path: Path,
) -> dict[str, Any]:
    if not path.exists():
        return {}

    return load_json_object(
        path
    )


def load_monitoring_history() -> pd.DataFrame:
    if HISTORY_CSV_PATH.exists():
        try:
            history = pd.read_csv(
                HISTORY_CSV_PATH,
                low_memory=False,
            )

            if (
                "generated_utc"
                in history.columns
            ):
                history[
                    "generated_utc"
                ] = pd.to_datetime(
                    history[
                        "generated_utc"
                    ],
                    errors="coerce",
                    utc=True,
                )

            return history

        except Exception:
            pass

    if not HISTORY_DATABASE_PATH.exists():
        return pd.DataFrame()

    try:
        with sqlite3.connect(
            HISTORY_DATABASE_PATH
        ) as connection:
            return pd.read_sql_query(
                """
                SELECT *
                FROM monitoring_runs
                ORDER BY generated_utc
                """,
                connection,
            )
    except Exception:
        return pd.DataFrame()


def load_latest_monitoring_outputs() -> dict[str, Any]:
    feature_drift = read_csv_optional(
        FEATURE_DRIFT_PATH,
        required_columns=[
            "feature",
            "overall_severity",
            "feature_health_score",
        ],
    )

    prediction_drift = read_csv_optional(
        PREDICTION_DRIFT_PATH,
        required_columns=[
            "metric_name",
            "severity",
            "health_score",
        ],
    )

    decision_drift = read_csv_optional(
        TABLES_DIR
        / "decision_drift.csv",
    )

    data_quality = read_csv_optional(
        DATA_QUALITY_PATH,
    )

    alerts = read_csv_optional(
        ALERTS_PATH,
    )

    health_components = read_csv_optional(
        HEALTH_COMPONENTS_PATH,
    )

    drift_summary = read_json_optional(
        DRIFT_SUMMARY_PATH
    )

    deployment_health = read_json_optional(
        DEPLOYMENT_HEALTH_PATH
    )

    history = load_monitoring_history()

    return {
        "feature_drift": feature_drift,
        "prediction_drift": prediction_drift,
        "decision_drift": decision_drift,
        "data_quality": data_quality,
        "alerts": alerts,
        "health_components": health_components,
        "drift_summary": drift_summary,
        "deployment_health": deployment_health,
        "history": history,
    }


def monitoring_outputs_available(
    outputs: Mapping[str, Any],
) -> bool:
    health_payload = outputs.get(
        "deployment_health",
        {},
    )

    return bool(
        isinstance(
            health_payload,
            Mapping,
        )
        and health_payload.get(
            "deployment_health"
        )
    )


# ============================================================
# DASHBOARD HELPERS
# ============================================================
def render_monitoring_metric(
    label: str,
    value: str,
    note: str,
) -> None:
    import streamlit as st

    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def severity_callout_class(
    severity: str,
) -> str:
    mapping = {
        "Normal": "normal-callout",
        "Watch": "watch-callout",
        "Warning": "warning-callout",
        "Critical": "critical-callout",
    }

    return mapping.get(
        severity,
        "watch-callout",
    )


def render_severity_callout(
    title: str,
    message: str,
    severity: str,
) -> None:
    import streamlit as st

    css_class = severity_callout_class(
        severity
    )

    st.markdown(
        f"""
        <div class="{css_class}">
            <strong>{title}</strong><br>
            {message}
        </div>
        """,
        unsafe_allow_html=True,
    )


def require_monitoring_access() -> dict[str, Any]:
    import streamlit as st

    user = st.session_state.get(
        "hevemind_user",
        {},
    )

    role = safe_text(
        user.get(
            "role"
        ),
        "",
    ).lower()

    standalone = (
        DEFAULT_STANDALONE_ADMIN_MODE
        or os.getenv(
            "HEVEMIND_MONITORING_STANDALONE_MODE",
            "false",
        )
        .strip()
        .lower()
        in {
            "1",
            "true",
            "yes",
            "on",
        }
    )

    if (
        role not in {
            "admin",
            "engineer",
        }
        and not standalone
    ):
        st.error(
            "Engineer or administrator access is required."
        )

        st.stop()

    return (
        user
        if isinstance(
            user,
            dict,
        )
        else {}
    )


def monitoring_health_payload(
    outputs: Mapping[str, Any],
) -> dict[str, Any]:
    payload = outputs.get(
        "deployment_health",
        {},
    )

    if not isinstance(
        payload,
        Mapping,
    ):
        return {}

    health = payload.get(
        "deployment_health",
        {},
    )

    return (
        dict(
            health
        )
        if isinstance(
            health,
            Mapping,
        )
        else {}
    )


def dataframe_download_button(
    *,
    label: str,
    dataframe: pd.DataFrame,
    filename: str,
    key: str,
) -> None:
    import streamlit as st

    st.download_button(
        label,
        data=dataframe.to_csv(
            index=False
        ).encode(
            "utf-8"
        ),
        file_name=filename,
        mime="text/csv",
        key=key,
        use_container_width=True,
    )


def dataframe_numeric_filter(
    dataframe: pd.DataFrame,
    column: str,
    minimum: float | None,
    maximum: float | None,
) -> pd.DataFrame:
    if column not in dataframe.columns:
        return dataframe

    numeric = pd.to_numeric(
        dataframe[
            column
        ],
        errors="coerce",
    )

    mask = pd.Series(
        True,
        index=dataframe.index,
    )

    if minimum is not None:
        mask &= (
            numeric
            >= minimum
        )

    if maximum is not None:
        mask &= (
            numeric
            <= maximum
        )

    return dataframe.loc[
        mask
    ].copy()


def available_figure(
    filename: str,
) -> Path | None:
    path = FIGURES_DIR / filename

    return (
        path
        if path.exists()
        else None
    )


# ============================================================
# DASHBOARD EXECUTIVE OVERVIEW
# ============================================================
def render_monitoring_overview(
    outputs: Mapping[str, Any],
) -> None:
    import streamlit as st

    health = monitoring_health_payload(
        outputs
    )

    feature_drift = outputs[
        "feature_drift"
    ]

    prediction_drift = outputs[
        "prediction_drift"
    ]

    alerts = outputs[
        "alerts"
    ]

    data_quality = outputs[
        "data_quality"
    ]

    overall_score = safe_float(
        health.get(
            "overall_health_score"
        ),
        0.0,
    ) or 0.0

    overall_band = safe_text(
        health.get(
            "overall_health_band"
        )
    )

    recommendation = safe_text(
        health.get(
            "recommendation"
        )
    )

    critical_alerts = safe_int(
        health.get(
            "critical_alerts"
        )
    )

    warning_alerts = safe_int(
        health.get(
            "warning_alerts"
        )
    )

    metric_columns = st.columns(
        5
    )

    metrics = [
        (
            "Deployment health",
            f"{overall_score:.1f}/100",
            f"Current health band: {overall_band}.",
        ),
        (
            "Recommendation",
            recommendation,
            "Monitoring recommendation only; no model change is executed.",
        ),
        (
            "Critical alerts",
            str(
                critical_alerts
            ),
            "Signals requiring immediate engineering or model-owner review.",
        ),
        (
            "Warning alerts",
            str(
                warning_alerts
            ),
            "Signals requiring enhanced monitoring and investigation.",
        ),
        (
            "Monitored sensors",
            str(
                len(
                    feature_drift
                )
            ),
            "Shared numerical features evaluated against the reference data.",
        ),
    ]

    for column, metric in zip(
        metric_columns,
        metrics,
    ):
        with column:
            render_monitoring_metric(
                metric[
                    0
                ],
                metric[
                    1
                ],
                metric[
                    2
                ],
            )

    st.write(
        ""
    )

    severity = safe_text(
        health.get(
            "overall_severity"
        ),
        "Watch",
    )

    render_severity_callout(
        "Deployment assessment",
        safe_text(
            health.get(
                "summary_statement"
            )
        ),
        severity,
    )

    st.subheader(
        "Health Components"
    )

    health_components = outputs[
        "health_components"
    ]

    if health_components.empty:
        st.info(
            "Health-component data is unavailable."
        )
    else:
        display = health_components.copy()

        display[
            "component_name"
        ] = display[
            "component_name"
        ].str.replace(
            "_",
            " ",
        ).str.title()

        st.dataframe(
            display,
            use_container_width=True,
            hide_index=True,
        )

    figure_columns = st.columns(
        2
    )

    health_figure = available_figure(
        "health_components.png"
    )

    alert_figure = available_figure(
        "alert_distribution.png"
    )

    with figure_columns[
        0
    ]:
        st.subheader(
            "Component Health"
        )

        if health_figure:
            st.image(
                str(
                    health_figure
                ),
                use_container_width=True,
            )
        else:
            st.info(
                "Run report generation to create the health-component figure."
            )

    with figure_columns[
        1
    ]:
        st.subheader(
            "Alert Distribution"
        )

        if alert_figure:
            st.image(
                str(
                    alert_figure
                ),
                use_container_width=True,
            )
        else:
            st.info(
                "Run report generation to create the alert-distribution figure."
            )

    st.subheader(
        "Operational Monitoring Snapshot"
    )

    snapshot_columns = st.columns(
        3
    )

    feature_dominant = (
        maximum_severity(
            feature_drift[
                "overall_severity"
            ].tolist()
        )
        if not feature_drift.empty
        else "Unavailable"
    )

    prediction_dominant = (
        maximum_severity(
            prediction_drift[
                "severity"
            ].tolist()
        )
        if not prediction_drift.empty
        else "Unavailable"
    )

    quality_dominant = (
        maximum_severity(
            data_quality[
                "severity"
            ].tolist()
        )
        if not data_quality.empty
        else "Unavailable"
    )

    for column, (
        heading,
        value,
        note,
    ) in zip(
        snapshot_columns,
        [
            (
                "Feature drift",
                feature_dominant,
                "Highest sensor-level drift severity.",
            ),
            (
                "Prediction drift",
                prediction_dominant,
                "Highest monitored model-output severity.",
            ),
            (
                "Data quality",
                quality_dominant,
                "Highest ingestion or schema-quality severity.",
            ),
        ],
    ):
        with column:
            st.markdown(
                f"""
                <div class="status-panel">
                    <div class="status-heading">{heading}</div>
                    <div class="status-value">{value}</div>
                    <div class="status-note">{note}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    if not alerts.empty:
        st.subheader(
            "Highest-Priority Active Alerts"
        )

        st.dataframe(
            alerts.head(
                12
            ),
            use_container_width=True,
            hide_index=True,
        )


# ============================================================
# DASHBOARD FEATURE-DRIFT PAGE
# ============================================================
def render_feature_drift_page(
    outputs: Mapping[str, Any],
) -> None:
    import streamlit as st

    dataframe = outputs[
        "feature_drift"
    ].copy()

    if dataframe.empty:
        st.warning(
            "Feature-drift results are unavailable."
        )

        return

    st.subheader(
        "Sensor Distribution Drift"
    )

    st.markdown(
        """
        <div class="section-note">
            Sensor drift indicates a statistical change between the reference
            process distribution and the current deployment window. It does
            not independently establish a physical root cause.
        </div>
        """,
        unsafe_allow_html=True,
    )

    filter_columns = st.columns(
        4
    )

    with filter_columns[
        0
    ]:
        severity_options = [
            "All",
            "Normal",
            "Watch",
            "Warning",
            "Critical",
        ]

        selected_severity = st.selectbox(
            "Severity",
            options=severity_options,
            key="monitoring_feature_severity",
        )

    with filter_columns[
        1
    ]:
        search_text = st.text_input(
            "Sensor search",
            key="monitoring_feature_search",
        )

    with filter_columns[
        2
    ]:
        maximum_health = st.slider(
            "Maximum health score",
            min_value=0,
            max_value=100,
            value=100,
            key="monitoring_feature_health",
        )

    with filter_columns[
        3
    ]:
        minimum_psi = st.number_input(
            "Minimum PSI",
            min_value=0.0,
            value=0.0,
            step=0.01,
            key="monitoring_feature_psi",
        )

    filtered = dataframe.copy()

    if selected_severity != "All":
        filtered = filtered.loc[
            filtered[
                "overall_severity"
            ]
            == selected_severity
        ]

    if search_text.strip():
        filtered = filtered.loc[
            filtered[
                "feature"
            ]
            .astype(
                str
            )
            .str.contains(
                search_text.strip(),
                case=False,
                regex=False,
            )
        ]

    filtered = dataframe_numeric_filter(
        filtered,
        "feature_health_score",
        None,
        float(
            maximum_health
        ),
    )

    filtered = dataframe_numeric_filter(
        filtered,
        "psi",
        float(
            minimum_psi
        ),
        None,
    )

    summary_columns = st.columns(
        5
    )

    severity_counts = dataframe[
        "overall_severity"
    ].value_counts()

    values = [
        (
            "Displayed sensors",
            str(
                len(
                    filtered
                )
            ),
            "Sensors matching the active filters.",
        ),
        (
            "Critical",
            str(
                int(
                    severity_counts.get(
                        "Critical",
                        0,
                    )
                )
            ),
            "Critical sensor distribution shifts.",
        ),
        (
            "Warning",
            str(
                int(
                    severity_counts.get(
                        "Warning",
                        0,
                    )
                )
            ),
            "Warning-level sensor distribution shifts.",
        ),
        (
            "Watch",
            str(
                int(
                    severity_counts.get(
                        "Watch",
                        0,
                    )
                )
            ),
            "Moderate shifts requiring observation.",
        ),
        (
            "Mean health",
            (
                f"{pd.to_numeric(dataframe['feature_health_score'], errors='coerce').mean():.1f}"
            ),
            "Mean feature-health score across monitored sensors.",
        ),
    ]

    for column, metric in zip(
        summary_columns,
        values,
    ):
        with column:
            render_monitoring_metric(
                metric[
                    0
                ],
                metric[
                    1
                ],
                metric[
                    2
                ],
            )

    st.dataframe(
        filtered,
        use_container_width=True,
        hide_index=True,
        height=600,
    )

    download_columns = st.columns(
        2
    )

    with download_columns[
        0
    ]:
        dataframe_download_button(
            label="Download filtered feature drift",
            dataframe=filtered,
            filename="hevemind_filtered_feature_drift.csv",
            key="download_filtered_feature_drift",
        )

    with download_columns[
        1
    ]:
        dataframe_download_button(
            label="Download complete feature drift",
            dataframe=dataframe,
            filename="hevemind_feature_drift.csv",
            key="download_complete_feature_drift",
        )

    figure_path = available_figure(
        "feature_drift_priority.png"
    )

    if figure_path:
        st.subheader(
            "Highest-Priority Sensor Shifts"
        )

        st.image(
            str(
                figure_path
            ),
            use_container_width=True,
        )

    st.subheader(
        "Sensor Investigation"
    )

    selected_sensor = st.selectbox(
        "Select sensor",
        options=dataframe[
            "feature"
        ].tolist(),
        key="monitoring_selected_sensor",
    )

    sensor_record = dataframe.loc[
        dataframe[
            "feature"
        ]
        == selected_sensor
    ]

    if not sensor_record.empty:
        record = sensor_record.iloc[
            0
        ].to_dict()

        sensor_columns = st.columns(
            4
        )

        sensor_metrics = [
            (
                "Severity",
                safe_text(
                    record.get(
                        "overall_severity"
                    )
                ),
                "Combined feature-drift classification.",
            ),
            (
                "Health",
                f"{safe_float(record.get('feature_health_score'), 0.0):.1f}",
                "Weighted sensor-health score.",
            ),
            (
                "PSI",
                (
                    f"{safe_float(record.get('psi'), 0.0):.4f}"
                    if safe_float(
                        record.get(
                            "psi"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                "Population Stability Index.",
            ),
            (
                "KS statistic",
                (
                    f"{safe_float(record.get('ks_statistic'), 0.0):.4f}"
                    if safe_float(
                        record.get(
                            "ks_statistic"
                        )
                    )
                    is not None
                    else "Unavailable"
                ),
                "Maximum difference between empirical CDFs.",
            ),
        ]

        for column, metric in zip(
            sensor_columns,
            sensor_metrics,
        ):
            with column:
                render_monitoring_metric(
                    metric[
                        0
                    ],
                    metric[
                        1
                    ],
                    metric[
                        2
                    ],
                )

        render_severity_callout(
            selected_sensor,
            safe_text(
                record.get(
                    "notes"
                )
            ),
            safe_text(
                record.get(
                    "overall_severity"
                ),
                "Watch",
            ),
        )


# ============================================================
# DASHBOARD PREDICTION AND DECISION PAGE
# ============================================================
def render_prediction_monitoring_page(
    outputs: Mapping[str, Any],
) -> None:
    import streamlit as st

    prediction_drift = outputs[
        "prediction_drift"
    ]

    decision_drift = outputs[
        "decision_drift"
    ]

    st.subheader(
        "Prediction, Confidence and Uncertainty Drift"
    )

    st.markdown(
        """
        <div class="section-note">
            This page monitors the distributions of calibrated failure
            probability, prediction confidence, combined uncertainty and
            available data-confidence outputs.
        </div>
        """,
        unsafe_allow_html=True,
    )

    if prediction_drift.empty:
        st.warning(
            "Prediction-drift results are unavailable."
        )
    else:
        st.dataframe(
            prediction_drift,
            use_container_width=True,
            hide_index=True,
        )

        dataframe_download_button(
            label="Download prediction drift",
            dataframe=prediction_drift,
            filename="hevemind_prediction_drift.csv",
            key="download_prediction_drift",
        )

    prediction_figure = available_figure(
        "prediction_shift.png"
    )

    if prediction_figure:
        st.image(
            str(
                prediction_figure
            ),
            use_container_width=True,
        )

    st.subheader(
        "Operational Decision Distribution"
    )

    if decision_drift.empty:
        st.info(
            "Decision-distribution results are unavailable."
        )
    else:
        st.dataframe(
            decision_drift,
            use_container_width=True,
            hide_index=True,
        )

        dataframe_download_button(
            label="Download decision drift",
            dataframe=decision_drift,
            filename="hevemind_decision_drift.csv",
            key="download_decision_drift",
        )

    decision_figure = available_figure(
        "decision_distribution.png"
    )

    if decision_figure:
        st.image(
            str(
                decision_figure
            ),
            use_container_width=True,
        )


# ============================================================
# DASHBOARD DATA-QUALITY AND ALERTS PAGE
# ============================================================
def render_quality_alerts_page(
    outputs: Mapping[str, Any],
) -> None:
    import streamlit as st

    data_quality = outputs[
        "data_quality"
    ]

    alerts = outputs[
        "alerts"
    ]

    st.subheader(
        "Data Quality Monitoring"
    )

    if data_quality.empty:
        st.warning(
            "Data-quality results are unavailable."
        )
    else:
        st.dataframe(
            data_quality,
            use_container_width=True,
            hide_index=True,
        )

        dataframe_download_button(
            label="Download data-quality monitoring",
            dataframe=data_quality,
            filename="hevemind_data_quality.csv",
            key="download_data_quality",
        )

    st.subheader(
        "Monitoring Alerts"
    )

    if alerts.empty:
        st.success(
            "No Watch, Warning or Critical alerts are active."
        )

        return

    filter_columns = st.columns(
        4
    )

    with filter_columns[
        0
    ]:
        severity = st.selectbox(
            "Alert severity",
            options=[
                "All",
                "Watch",
                "Warning",
                "Critical",
            ],
            key="monitoring_alert_severity",
        )

    with filter_columns[
        1
    ]:
        categories = [
            "All",
            *sorted(
                alerts[
                    "category"
                ]
                .dropna()
                .unique()
                .tolist()
            ),
        ]

        category = st.selectbox(
            "Alert category",
            options=categories,
            key="monitoring_alert_category",
        )

    with filter_columns[
        2
    ]:
        resource_search = st.text_input(
            "Resource search",
            key="monitoring_alert_resource",
        )

    with filter_columns[
        3
    ]:
        maximum_rows = st.number_input(
            "Maximum rows",
            min_value=10,
            max_value=1000,
            value=200,
            step=10,
            key="monitoring_alert_rows",
        )

    filtered = alerts.copy()

    if severity != "All":
        filtered = filtered.loc[
            filtered[
                "severity"
            ]
            == severity
        ]

    if category != "All":
        filtered = filtered.loc[
            filtered[
                "category"
            ]
            == category
        ]

    if resource_search.strip():
        mask = (
            filtered[
                "resource_id"
            ]
            .astype(
                str
            )
            .str.contains(
                resource_search.strip(),
                case=False,
                regex=False,
            )
            | filtered[
                "message"
            ]
            .astype(
                str
            )
            .str.contains(
                resource_search.strip(),
                case=False,
                regex=False,
            )
        )

        filtered = filtered.loc[
            mask
        ]

    filtered = filtered.head(
        int(
            maximum_rows
        )
    )

    alert_columns = st.columns(
        4
    )

    alert_metrics = [
        (
            "Displayed",
            str(
                len(
                    filtered
                )
            ),
            "Alerts matching active filters.",
        ),
        (
            "Critical",
            str(
                count_severity(
                    alerts,
                    "severity",
                    "Critical",
                )
            ),
            "Immediate engineering or model-owner attention.",
        ),
        (
            "Warning",
            str(
                count_severity(
                    alerts,
                    "severity",
                    "Warning",
                )
            ),
            "Enhanced investigation and monitoring.",
        ),
        (
            "Watch",
            str(
                count_severity(
                    alerts,
                    "severity",
                    "Watch",
                )
            ),
            "Continue deployment with observation.",
        ),
    ]

    for column, metric in zip(
        alert_columns,
        alert_metrics,
    ):
        with column:
            render_monitoring_metric(
                metric[
                    0
                ],
                metric[
                    1
                ],
                metric[
                    2
                ],
            )

    st.dataframe(
        filtered,
        use_container_width=True,
        hide_index=True,
        height=600,
    )

    dataframe_download_button(
        label="Download filtered alerts",
        dataframe=filtered,
        filename="hevemind_filtered_monitoring_alerts.csv",
        key="download_filtered_alerts",
    )

    st.subheader(
        "Alert Details"
    )

    selected_alert = st.selectbox(
        "Select alert",
        options=alerts[
            "alert_id"
        ].tolist(),
        key="monitoring_selected_alert",
    )

    selected = alerts.loc[
        alerts[
            "alert_id"
        ]
        == selected_alert
    ]

    if not selected.empty:
        record = selected.iloc[
            0
        ].to_dict()

        render_severity_callout(
            safe_text(
                record.get(
                    "resource_id"
                )
            ),
            safe_text(
                record.get(
                    "message"
                )
            ),
            safe_text(
                record.get(
                    "severity"
                ),
                "Watch",
            ),
        )

        st.markdown(
            f"""
            <div class="section-note">
                <strong>Recommended action:</strong><br>
                {safe_text(record.get("recommended_action"))}
            </div>
            """,
            unsafe_allow_html=True,
        )

        evidence_json = safe_text(
            record.get(
                "evidence_json"
            ),
            "{}",
        )

        try:
            st.json(
                json.loads(
                    evidence_json
                ),
                expanded=False,
            )
        except json.JSONDecodeError:
            st.code(
                evidence_json,
                language="text",
            )


# ============================================================
# DASHBOARD HISTORY AND REPORTS PAGE
# ============================================================
def render_history_reports_page(
    outputs: Mapping[str, Any],
) -> None:
    import streamlit as st

    history = outputs[
        "history"
    ]

    st.subheader(
        "Monitoring History"
    )

    if history.empty:
        st.info(
            "No monitoring history is available yet."
        )
    else:
        st.dataframe(
            history,
            use_container_width=True,
            hide_index=True,
        )

        health_history_figure = available_figure(
            "monitoring_history.png"
        )

        if health_history_figure:
            st.image(
                str(
                    health_history_figure
                ),
                use_container_width=True,
            )

        dataframe_download_button(
            label="Download monitoring history",
            dataframe=history,
            filename="hevemind_monitoring_history.csv",
            key="download_monitoring_history",
        )

    st.subheader(
        "Generated Monitoring Reports"
    )

    report_rows = []

    report_patterns = [
        (
            "Excel",
            EXPORTS_DIR.glob(
                "hevemind_model_monitoring_*.xlsx"
            ),
        ),
        (
            "PDF",
            EXPORTS_DIR.glob(
                "hevemind_model_monitoring_*.pdf"
            ),
        ),
        (
            "Manifest",
            EXPORTS_DIR.glob(
                "monitoring_report_manifest_*.json"
            ),
        ),
    ]

    for report_type, paths in report_patterns:
        for path in paths:
            report_rows.append(
                {
                    "type": report_type,
                    "filename": path.name,
                    "path": str(
                        path
                    ),
                    "size_bytes": (
                        path.stat().st_size
                    ),
                    "modified_utc": datetime.fromtimestamp(
                        path.stat().st_mtime,
                        tz=timezone.utc,
                    ).isoformat(),
                }
            )

    reports = pd.DataFrame(
        report_rows
    )

    if reports.empty:
        st.info(
            "No archived monitoring reports are available."
        )
    else:
        reports = reports.sort_values(
            "modified_utc",
            ascending=False,
        ).reset_index(
            drop=True
        )

        st.dataframe(
            reports,
            use_container_width=True,
            hide_index=True,
        )

        selected_report_path = st.selectbox(
            "Select report to download",
            options=reports[
                "path"
            ].tolist(),
            key="monitoring_report_download_selection",
        )

        selected_path = Path(
            selected_report_path
        )

        mime_type = {
            ".xlsx": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            ".pdf": "application/pdf",
            ".json": "application/json",
        }.get(
            selected_path.suffix.lower(),
            "application/octet-stream",
        )

        st.download_button(
            "Download selected report",
            data=selected_path.read_bytes(),
            file_name=selected_path.name,
            mime=mime_type,
            use_container_width=True,
            key="monitoring_selected_report_download",
        )


# ============================================================
# DASHBOARD RUN-CONTROL PAGE
# ============================================================
def render_monitoring_run_control(
    actor: Mapping[str, Any],
) -> None:
    import streamlit as st

    st.subheader(
        "Run Model Monitoring"
    )

    st.markdown(
        """
        <div class="section-note">
            Select a stable reference dataset and a current deployment batch.
            The monitoring engine reads these files and generates drift,
            quality, alert, health and reporting outputs. It does not train,
            overwrite or modify any deployed model.
        </div>
        """,
        unsafe_allow_html=True,
    )

    config = initialise_monitoring_config()

    default_reference = safe_text(
        config.get(
            "reference_data_path"
        ),
        "",
    )

    default_current = safe_text(
        config.get(
            "current_data_path"
        ),
        "",
    )

    with st.form(
        "monitoring_run_form"
    ):
        reference_text = st.text_input(
            "Reference dataset path",
            value=default_reference,
            placeholder=(
                "/Users/barkavi/Desktop/HeveMind/data/processed/development.csv"
            ),
        )

        current_text = st.text_input(
            "Current deployment dataset path",
            value=default_current,
            placeholder=(
                "/Users/barkavi/Desktop/HeveMind/data/processed/test.csv"
            ),
        )

        config_text = st.text_input(
            "Optional monitoring configuration path",
            value="",
            placeholder=str(
                CONFIG_PATH
            ),
        )

        option_columns = st.columns(
            3
        )

        with option_columns[
            0
        ]:
            persist_history = st.checkbox(
                "Persist monitoring history",
                value=True,
            )

        with option_columns[
            1
        ]:
            generate_reports = st.checkbox(
                "Generate Excel and figures",
                value=True,
            )

        with option_columns[
            2
        ]:
            generate_pdf = st.checkbox(
                "Generate PDF",
                value=True,
            )

        submitted = st.form_submit_button(
            "Run complete monitoring pipeline",
            use_container_width=True,
        )

    if submitted:
        reference_path = (
            Path(
                reference_text
            ).expanduser()
            if reference_text.strip()
            else None
        )

        current_path = (
            Path(
                current_text
            ).expanduser()
            if current_text.strip()
            else None
        )

        config_path = (
            Path(
                config_text
            ).expanduser()
            if config_text.strip()
            else None
        )

        with st.spinner(
            "Running HeveMind monitoring and generating outputs..."
        ):
            try:
                bundle, artifacts = run_complete_monitoring_pipeline(
                    reference_path=reference_path,
                    current_path=current_path,
                    config_path=config_path,
                    save_output=True,
                    persist_history=persist_history,
                    generate_reports=generate_reports,
                    generate_pdf=generate_pdf,
                )

                record_monitoring_event(
                    event_type="model_monitoring_run_completed",
                    outcome="success",
                    action="run_monitoring_pipeline",
                    resource_id=(
                        bundle.context.run_identity.run_id
                    ),
                    details={
                        "overall_health_score": (
                            bundle.health_assessment.overall_health_score
                        ),
                        "recommendation": (
                            bundle.health_assessment.recommendation
                        ),
                        "report_generated": (
                            artifacts is not None
                        ),
                        "model_modification_performed": False,
                    },
                    actor={
                        **dict(
                            actor
                        ),
                        "session_id": st.session_state.get(
                            "hevemind_audit_session_id"
                        ),
                    },
                )

                st.session_state[
                    "monitoring_last_run"
                ] = {
                    "run_id": (
                        bundle.context.run_identity.run_id
                    ),
                    "health_score": (
                        bundle.health_assessment.overall_health_score
                    ),
                    "recommendation": (
                        bundle.health_assessment.recommendation
                    ),
                    "artifacts": (
                        asdict(
                            artifacts
                        )
                        if artifacts is not None
                        else None
                    ),
                }

                st.success(
                    (
                        "Monitoring completed. Health score: "
                        f"{bundle.health_assessment.overall_health_score:.2f}/100. "
                        f"Recommendation: "
                        f"{bundle.health_assessment.recommendation}."
                    )
                )

                st.rerun()

            except Exception as error:
                record_monitoring_event(
                    event_type="model_monitoring_run_failed",
                    outcome="failure",
                    action="run_monitoring_pipeline",
                    resource_id=None,
                    details={
                        "error": str(
                            error
                        )
                    },
                    actor={
                        **dict(
                            actor
                        ),
                        "session_id": st.session_state.get(
                            "hevemind_audit_session_id"
                        ),
                    },
                )

                st.exception(
                    error
                )

    last_run = st.session_state.get(
        "monitoring_last_run"
    )

    if last_run:
        st.subheader(
            "Latest Run Result"
        )

        st.json(
            last_run,
            expanded=True,
        )

    st.subheader(
        "Monitoring Configuration"
    )

    with st.expander(
        "View active monitoring policy",
        expanded=False,
    ):
        st.json(
            config,
            expanded=False,
        )


# ============================================================
# COMPLETE STREAMLIT MONITORING DASHBOARD
# ============================================================
() -> None:
    import streamlit as st

    ensure_directories()

    st.markdown(
        MONITORING_DASHBOARD_CSS,
        unsafe_allow_html=True,
    )

    actor = require_monitoring_access()

    outputs = load_latest_monitoring_outputs()

    health = monitoring_health_payload(
        outputs
    )

    health_score = safe_float(
        health.get(
            "overall_health_score"
        )
    )

    health_band = safe_text(
        health.get(
            "overall_health_band"
        ),
        "Not assessed",
    )

    recommendation = safe_text(
        health.get(
            "recommendation"
        ),
        "Run monitoring",
    )

    st.markdown(
        f"""
        <div class="monitoring-header">
            <div class="monitoring-title">
                HeveMind Model Monitoring and Drift Detection
            </div>
            <div class="monitoring-subtitle">
                Semiconductor sensor drift, prediction stability,
                confidence, uncertainty, data quality, deployment health
                and retraining recommendation
            </div>
            <div class="monitoring-chip-row">
                <span class="monitoring-chip">
                    Version {APP_VERSION}
                </span>
                <span class="monitoring-chip">
                    Health {
                        f"{health_score:.1f}/100"
                        if health_score is not None
                        else "Not assessed"
                    }
                </span>
                <span class="monitoring-chip">
                    Band {health_band}
                </span>
                <span class="monitoring-chip">
                    {recommendation}
                </span>
                <span class="monitoring-chip">
                    Monitoring only
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    page = st.sidebar.radio(
        "Monitoring workspace",
        options=[
            "Executive Overview",
            "Feature Drift",
            "Prediction and Decisions",
            "Data Quality and Alerts",
            "History and Reports",
            "Run Monitoring",
        ],
    )

    st.sidebar.markdown(
        "---"
    )

    if monitoring_outputs_available(
        outputs
    ):
        st.sidebar.success(
            "Monitoring outputs available"
        )

        st.sidebar.caption(
            f"Latest health band: {health_band}"
        )
    else:
        st.sidebar.warning(
            "No complete monitoring run found"
        )

    if st.sidebar.button(
        "Refresh monitoring outputs",
        use_container_width=True,
    ):
        st.rerun()

    if (
        page
        != "Run Monitoring"
        and not monitoring_outputs_available(
            outputs
        )
    ):
        st.warning(
            "No complete monitoring outputs are available. Open Run Monitoring "
            "to execute the pipeline."
        )

        return

    if page == "Executive Overview":
        render_monitoring_overview(
            outputs
        )

    elif page == "Feature Drift":
        render_feature_drift_page(
            outputs
        )

    elif page == "Prediction and Decisions":
        render_prediction_monitoring_page(
            outputs
        )

    elif page == "Data Quality and Alerts":
        render_quality_alerts_page(
            outputs
        )

    elif page == "History and Reports":
        render_history_reports_page(
            outputs
        )

    else:
        render_monitoring_run_control(
            actor
        )


# ============================================================
# CLI CONFIGURATION AND VALIDATION
# ============================================================
def initialise_monitoring_environment() -> dict[str, Any]:
    config = initialise_monitoring_config()

    initialise_monitoring_history_database()

    return {
        "config_path": str(
            CONFIG_PATH
        ),
        "history_database": str(
            HISTORY_DATABASE_PATH
        ),
        "monitoring_directory": str(
            MONITORING_DIR
        ),
        "config": config,
    }


def monitoring_status_payload() -> dict[str, Any]:
    outputs = load_latest_monitoring_outputs()

    return {
        "application": APP_NAME,
        "version": APP_VERSION,
        "generated_utc": utc_now_iso(),
        "outputs_available": monitoring_outputs_available(
            outputs
        ),
        "paths": {
            "config": str(
                CONFIG_PATH
            ),
            "drift_summary": str(
                DRIFT_SUMMARY_PATH
            ),
            "deployment_health": str(
                DEPLOYMENT_HEALTH_PATH
            ),
            "feature_drift": str(
                FEATURE_DRIFT_PATH
            ),
            "prediction_drift": str(
                PREDICTION_DRIFT_PATH
            ),
            "alerts": str(
                ALERTS_PATH
            ),
            "history_database": str(
                HISTORY_DATABASE_PATH
            ),
            "excel_report": str(
                EXCEL_REPORT_PATH
            ),
            "pdf_report": str(
                PDF_REPORT_PATH
            ),
        },
        "deployment_health": monitoring_health_payload(
            outputs
        ),
        "row_counts": {
            "feature_drift": len(
                outputs[
                    "feature_drift"
                ]
            ),
            "prediction_drift": len(
                outputs[
                    "prediction_drift"
                ]
            ),
            "decision_drift": len(
                outputs[
                    "decision_drift"
                ]
            ),
            "data_quality": len(
                outputs[
                    "data_quality"
                ]
            ),
            "alerts": len(
                outputs[
                    "alerts"
                ]
            ),
            "history": len(
                outputs[
                    "history"
                ]
            ),
        },
    }


def validate_existing_monitoring_outputs() -> dict[str, Any]:
    outputs = load_latest_monitoring_outputs()

    checks: list[dict[str, Any]] = []

    path_checks = [
        (
            "Feature drift CSV",
            FEATURE_DRIFT_PATH,
            True,
        ),
        (
            "Prediction drift CSV",
            PREDICTION_DRIFT_PATH,
            True,
        ),
        (
            "Decision drift CSV",
            TABLES_DIR
            / "decision_drift.csv",
            True,
        ),
        (
            "Data quality CSV",
            DATA_QUALITY_PATH,
            True,
        ),
        (
            "Alerts CSV",
            ALERTS_PATH,
            True,
        ),
        (
            "Health components CSV",
            HEALTH_COMPONENTS_PATH,
            True,
        ),
        (
            "Drift summary JSON",
            DRIFT_SUMMARY_PATH,
            True,
        ),
        (
            "Deployment health JSON",
            DEPLOYMENT_HEALTH_PATH,
            True,
        ),
        (
            "History database",
            HISTORY_DATABASE_PATH,
            False,
        ),
        (
            "Latest Excel report",
            EXCEL_REPORT_PATH,
            False,
        ),
        (
            "Latest PDF report",
            PDF_REPORT_PATH,
            False,
        ),
    ]

    for label, path, critical in path_checks:
        exists = path.exists()

        checks.append(
            {
                "check": label,
                "critical": critical,
                "passed": exists,
                "message": (
                    str(
                        path
                    )
                    if exists
                    else f"Missing: {path}"
                ),
            }
        )

    if not outputs[
        "feature_drift"
    ].empty:
        try:
            validate_feature_drift_output(
                outputs[
                    "feature_drift"
                ],
                expected_feature_count=len(
                    outputs[
                        "feature_drift"
                    ]
                ),
            )

            checks.append(
                {
                    "check": (
                        "Feature-drift table validation"
                    ),
                    "critical": True,
                    "passed": True,
                    "message": (
                        "Feature-drift values and severity labels are valid."
                    ),
                }
            )

        except Exception as error:
            checks.append(
                {
                    "check": (
                        "Feature-drift table validation"
                    ),
                    "critical": True,
                    "passed": False,
                    "message": str(
                        error
                    ),
                }
            )

    if not outputs[
        "prediction_drift"
    ].empty:
        try:
            validate_prediction_drift_table(
                outputs[
                    "prediction_drift"
                ]
            )

            checks.append(
                {
                    "check": (
                        "Prediction-drift table validation"
                    ),
                    "critical": True,
                    "passed": True,
                    "message": (
                        "Prediction-drift values and severity labels are valid."
                    ),
                }
            )

        except Exception as error:
            checks.append(
                {
                    "check": (
                        "Prediction-drift table validation"
                    ),
                    "critical": True,
                    "passed": False,
                    "message": str(
                        error
                    ),
                }
            )

    failed_critical = sum(
        (
            not check[
                "passed"
            ]
            and check[
                "critical"
            ]
        )
        for check in checks
    )

    return {
        "generated_utc": utc_now_iso(),
        "valid": failed_critical == 0,
        "failed_critical_checks": (
            failed_critical
        ),
        "checks": checks,
    }


def parse_monitoring_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "HeveMind semiconductor model monitoring and drift detection."
        )
    )

    parser.add_argument(
        "--initialise",
        action="store_true",
        help=(
            "Create the monitoring configuration and history database."
        ),
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help=(
            "Run the complete monitoring pipeline."
        ),
    )

    parser.add_argument(
        "--reference",
        type=Path,
        help=(
            "def render_model_monitoring_dashboardReference or development dataset path."
        ),
    )

    parser.add_argument(
        "--current",
        type=Path,
        help=(
            "Current deployment dataset path."
        ),
    )

    parser.add_argument(
        "--config",
        type=Path,
        help=(
            "Optional monitoring configuration JSON."
        ),
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help=(
            "Print the latest monitoring status."
        ),
    )

    parser.add_argument(
        "--validate",
        action="store_true",
        help=(
            "Validate existing monitoring outputs."
        ),
    )

    parser.add_argument(
        "--no-history",
        action="store_true",
        help=(
            "Do not persist the monitoring run to history."
        ),
    )

    parser.add_argument(
        "--no-reports",
        action="store_true",
        help=(
            "Run monitoring without Excel, PDF or figure generation."
        ),
    )

    parser.add_argument(
        "--no-pdf",
        action="store_true",
        help=(
            "Generate the Excel report and figures without PDF."
        ),
    )

    parser.add_argument(
        "--json-output",
        type=Path,
        help=(
            "Optional path for a JSON summary of the completed run."
        ),
    )

    return parser.parse_args()


def monitoring_main() -> None:
    arguments = parse_monitoring_arguments()

    action_requested = False

    if arguments.initialise:
        payload = initialise_monitoring_environment()

        print(
            json.dumps(
                json_safe(
                    payload
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        action_requested = True

    if arguments.run:
        bundle, artifacts = run_complete_monitoring_pipeline(
            reference_path=arguments.reference,
            current_path=arguments.current,
            config_path=arguments.config,
            save_output=True,
            persist_history=(
                not arguments.no_history
            ),
            generate_reports=(
                not arguments.no_reports
            ),
            generate_pdf=(
                not arguments.no_pdf
            ),
        )

        console_output = monitoring_pipeline_console_output(
            bundle,
            artifacts,
        )

        print(
            console_output
        )

        summary = {
            "run_identity": asdict(
                bundle.context.run_identity
            ),
            "deployment_health": asdict(
                bundle.health_assessment
            ),
            "feature_drift_summary": asdict(
                bundle.feature_summary
            ),
            "prediction_monitoring_summary": asdict(
                bundle.prediction_summary
            ),
            "report_artifacts": (
                asdict(
                    artifacts
                )
                if artifacts is not None
                else None
            ),
            "model_modification_performed": False,
        }

        if arguments.json_output is not None:
            output_path = arguments.json_output.expanduser()

            if not output_path.is_absolute():
                output_path = (
                    ROOT_DIR
                    / output_path
                )

            atomic_write_json(
                output_path,
                summary,
            )

            print(
                f"\nJSON run summary:                {output_path}"
            )

        action_requested = True

    if arguments.status:
        print(
            json.dumps(
                json_safe(
                    monitoring_status_payload()
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        action_requested = True

    if arguments.validate:
        validation = validate_existing_monitoring_outputs()

        print(
            json.dumps(
                json_safe(
                    validation
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        if not validation[
            "valid"
        ]:
            raise SystemExit(
                1
            )

        action_requested = True

    if not action_requested:
        status = monitoring_status_payload()

        print(
            "\n"
            + "="
            * 118
        )

        print(
            "HEVEMIND MODEL MONITORING AND DRIFT DETECTION"
        )

        print(
            "="
            * 118
        )

        print(
            f"\nVersion:                         {APP_VERSION}"
        )

        print(
            f"Configuration:                   {CONFIG_PATH}"
        )

        print(
            f"Monitoring directory:            {MONITORING_DIR}"
        )

        print(
            f"Outputs available:               {status['outputs_available']}"
        )

        health = status.get(
            "deployment_health",
            {},
        )

        if health:
            print(
                "Latest health score:             "
                f"{health.get('overall_health_score')}"
            )

            print(
                "Latest recommendation:           "
                f"{health.get('recommendation')}"
            )

        print(
            "\nRun with --help to view monitoring commands."
        )


def streamlit_context_active() -> bool:
    try:
        from streamlit.runtime.scriptrunner import (
            get_script_run_ctx,
        )

        return (
            get_script_run_ctx()
            is not None
        )
    except Exception:
        return False



# ============================================================
# PART 7 — SCHEDULER SUPPORT, LOCKING, ORCHESTRATION AND MAIN
# ============================================================

import atexit
import signal
import time


SCHEDULER_DIR = MONITORING_DIR / "scheduler"
SCHEDULER_LOG_PATH = SCHEDULER_DIR / "monitoring_scheduler.log"
SCHEDULER_PID_PATH = SCHEDULER_DIR / "monitoring_scheduler.pid"
SCHEDULER_LOCK_PATH = SCHEDULER_DIR / "monitoring_run.lock"
SCHEDULER_STATE_PATH = SCHEDULER_DIR / "monitoring_scheduler_state.json"


@dataclass(frozen=True)
class SchedulerConfiguration:
    interval_minutes: int
    reference_path: Path | None
    current_path: Path | None
    config_path: Path | None
    generate_reports: bool
    generate_pdf: bool
    persist_history: bool
    run_immediately: bool


@dataclass(frozen=True)
class SchedulerState:
    pid: int
    started_utc: str
    last_run_started_utc: str | None
    last_run_completed_utc: str | None
    last_run_id: str | None
    last_health_score: float | None
    last_recommendation: str | None
    last_error: str | None
    successful_runs: int
    failed_runs: int
    next_run_utc: str | None
    status: str


class MonitoringRunLock:
    def __init__(
        self,
        path: Path = SCHEDULER_LOCK_PATH,
    ) -> None:
        self.path = path
        self.acquired = False

    def acquire(self) -> None:
        ensure_directories()
        SCHEDULER_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        if self.path.exists():
            try:
                payload = load_json_object(
                    self.path
                )

                existing_pid = int(
                    payload.get(
                        "pid",
                        0,
                    )
                )

            except Exception:
                existing_pid = 0

            if (
                existing_pid
                and process_is_running(
                    existing_pid
                )
            ):
                raise RuntimeError(
                    "Another HeveMind monitoring run is active "
                    f"under PID {existing_pid}."
                )

            self.path.unlink(
                missing_ok=True
            )

        payload = {
            "pid": os.getpid(),
            "acquired_utc": utc_now_iso(),
        }

        atomic_write_json(
            self.path,
            payload,
        )

        self.acquired = True

    def release(self) -> None:
        if self.acquired:
            self.path.unlink(
                missing_ok=True
            )

            self.acquired = False

    def __enter__(self) -> "MonitoringRunLock":
        self.acquire()
        return self

    def __exit__(
        self,
        exc_type: Any,
        exc_value: Any,
        traceback: Any,
    ) -> None:
        self.release()


def ensure_scheduler_directories() -> None:
    ensure_directories()

    SCHEDULER_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )


def append_scheduler_log(
    message: str,
) -> None:
    ensure_scheduler_directories()

    line = (
        f"{utc_now_iso()} | {message}\n"
    )

    with SCHEDULER_LOG_PATH.open(
        "a",
        encoding="utf-8",
    ) as file:
        file.write(
            line
        )


def process_is_running(
    pid: int | None,
) -> bool:
    if pid is None or pid <= 0:
        return False

    try:
        os.kill(
            pid,
            0,
        )

        return True

    except (
        ProcessLookupError,
        PermissionError,
        OSError,
    ):
        return False


def read_scheduler_pid() -> int | None:
    if not SCHEDULER_PID_PATH.exists():
        return None

    try:
        return int(
            SCHEDULER_PID_PATH.read_text(
                encoding="utf-8"
            ).strip()
        )

    except Exception:
        return None


def write_scheduler_pid(
    pid: int,
) -> None:
    ensure_scheduler_directories()

    SCHEDULER_PID_PATH.write_text(
        str(
            pid
        ),
        encoding="utf-8",
    )


def scheduler_is_running() -> bool:
    pid = read_scheduler_pid()

    if process_is_running(
        pid
    ):
        return True

    SCHEDULER_PID_PATH.unlink(
        missing_ok=True
    )

    return False


def scheduler_state_payload(
    **overrides: Any,
) -> dict[str, Any]:
    current: dict[str, Any] = {}

    if SCHEDULER_STATE_PATH.exists():
        try:
            current = load_json_object(
                SCHEDULER_STATE_PATH
            )

        except Exception:
            current = {}

    payload = {
        "pid": os.getpid(),
        "started_utc": current.get(
            "started_utc",
            utc_now_iso(),
        ),
        "last_run_started_utc": current.get(
            "last_run_started_utc"
        ),
        "last_run_completed_utc": current.get(
            "last_run_completed_utc"
        ),
        "last_run_id": current.get(
            "last_run_id"
        ),
        "last_health_score": current.get(
            "last_health_score"
        ),
        "last_recommendation": current.get(
            "last_recommendation"
        ),
        "last_error": current.get(
            "last_error"
        ),
        "successful_runs": int(
            current.get(
                "successful_runs",
                0,
            )
        ),
        "failed_runs": int(
            current.get(
                "failed_runs",
                0,
            )
        ),
        "next_run_utc": current.get(
            "next_run_utc"
        ),
        "status": current.get(
            "status",
            "Initialising",
        ),
    }

    payload.update(
        json_safe(
            overrides
        )
    )

    atomic_write_json(
        SCHEDULER_STATE_PATH,
        payload,
    )

    return payload


def load_scheduler_state() -> dict[str, Any]:
    if not SCHEDULER_STATE_PATH.exists():
        return {
            "running": False,
            "pid": None,
            "status": "Not started",
        }

    payload = load_json_object(
        SCHEDULER_STATE_PATH
    )

    pid = safe_int(
        payload.get(
            "pid"
        ),
        0,
    )

    payload[
        "running"
    ] = process_is_running(
        pid
    )

    return payload


def scheduled_monitoring_run(
    configuration: SchedulerConfiguration,
) -> tuple[
    MonitoringBundle,
    MonitoringReportArtifacts | None,
]:
    with MonitoringRunLock():
        append_scheduler_log(
            "Starting scheduled monitoring run."
        )

        state = load_scheduler_state()

        scheduler_state_payload(
            status="Running monitoring",
            last_run_started_utc=utc_now_iso(),
            last_error=None,
        )

        try:
            bundle, artifacts = run_complete_monitoring_pipeline(
                reference_path=configuration.reference_path,
                current_path=configuration.current_path,
                config_path=configuration.config_path,
                save_output=True,
                persist_history=configuration.persist_history,
                generate_reports=configuration.generate_reports,
                generate_pdf=configuration.generate_pdf,
            )

            scheduler_state_payload(
                status="Waiting",
                last_run_completed_utc=utc_now_iso(),
                last_run_id=(
                    bundle.context.run_identity.run_id
                ),
                last_health_score=(
                    bundle.health_assessment.overall_health_score
                ),
                last_recommendation=(
                    bundle.health_assessment.recommendation
                ),
                last_error=None,
                successful_runs=(
                    int(
                        state.get(
                            "successful_runs",
                            0,
                        )
                    )
                    + 1
                ),
            )

            append_scheduler_log(
                (
                    "Scheduled monitoring completed successfully. "
                    f"Run ID={bundle.context.run_identity.run_id}; "
                    f"health={bundle.health_assessment.overall_health_score:.2f}; "
                    f"recommendation={bundle.health_assessment.recommendation}."
                )
            )

            return (
                bundle,
                artifacts,
            )

        except Exception as error:
            scheduler_state_payload(
                status="Run failed",
                last_run_completed_utc=utc_now_iso(),
                last_error=str(
                    error
                ),
                failed_runs=(
                    int(
                        state.get(
                            "failed_runs",
                            0,
                        )
                    )
                    + 1
                ),
            )

            append_scheduler_log(
                f"Scheduled monitoring failed: {error}"
            )

            raise


def run_monitoring_scheduler(
    configuration: SchedulerConfiguration,
) -> None:
    if configuration.interval_minutes < 1:
        raise ValueError(
            "Scheduler interval must be at least one minute."
        )

    if scheduler_is_running():
        pid = read_scheduler_pid()

        if pid != os.getpid():
            raise RuntimeError(
                f"Monitoring scheduler is already running under PID {pid}."
            )

    ensure_scheduler_directories()

    write_scheduler_pid(
        os.getpid()
    )

    stop_requested = False

    def request_stop(
        signal_number: int,
        frame: Any,
    ) -> None:
        nonlocal stop_requested

        stop_requested = True

        append_scheduler_log(
            f"Stop requested by signal {signal_number}."
        )

        scheduler_state_payload(
            status="Stopping",
        )

    signal.signal(
        signal.SIGTERM,
        request_stop,
    )

    signal.signal(
        signal.SIGINT,
        request_stop,
    )

    def cleanup() -> None:
        SCHEDULER_PID_PATH.unlink(
            missing_ok=True
        )

        SCHEDULER_LOCK_PATH.unlink(
            missing_ok=True
        )

        scheduler_state_payload(
            status="Stopped",
            next_run_utc=None,
        )

        append_scheduler_log(
            "Monitoring scheduler stopped."
        )

    atexit.register(
        cleanup
    )

    scheduler_state_payload(
        pid=os.getpid(),
        started_utc=utc_now_iso(),
        status="Starting",
        next_run_utc=None,
    )

    append_scheduler_log(
        (
            "Monitoring scheduler started. "
            f"Interval={configuration.interval_minutes} minutes."
        )
    )

    interval_seconds = (
        configuration.interval_minutes
        * 60
    )

    next_run_epoch = (
        time.time()
        if configuration.run_immediately
        else time.time()
        + interval_seconds
    )

    while not stop_requested:
        current_time = time.time()

        if current_time >= next_run_epoch:
            try:
                scheduled_monitoring_run(
                    configuration
                )

            except Exception:
                pass

            next_run_epoch = (
                time.time()
                + interval_seconds
            )

        next_run_utc = datetime.fromtimestamp(
            next_run_epoch,
            tz=timezone.utc,
        ).isoformat()

        scheduler_state_payload(
            status="Waiting",
            next_run_utc=next_run_utc,
        )

        remaining = max(
            next_run_epoch
            - time.time(),
            0.0,
        )

        sleep_duration = min(
            remaining,
            5.0,
        )

        if sleep_duration <= 0:
            sleep_duration = 0.2

        time.sleep(
            sleep_duration
        )


def stop_monitoring_scheduler(
    timeout_seconds: int = 15,
) -> dict[str, Any]:
    pid = read_scheduler_pid()

    if pid is None:
        return {
            "stopped": True,
            "message": (
                "No scheduler PID is registered."
            ),
        }

    if not process_is_running(
        pid
    ):
        SCHEDULER_PID_PATH.unlink(
            missing_ok=True
        )

        scheduler_state_payload(
            status="Stopped",
            next_run_utc=None,
        )

        return {
            "stopped": True,
            "pid": pid,
            "message": (
                "Scheduler process was not active; stale PID removed."
            ),
        }

    os.kill(
        pid,
        signal.SIGTERM,
    )

    deadline = (
        time.time()
        + timeout_seconds
    )

    while (
        time.time()
        < deadline
    ):
        if not process_is_running(
            pid
        ):
            SCHEDULER_PID_PATH.unlink(
                missing_ok=True
            )

            scheduler_state_payload(
                status="Stopped",
                next_run_utc=None,
            )

            return {
                "stopped": True,
                "pid": pid,
                "message": (
                    "Scheduler stopped gracefully."
                ),
            }

        time.sleep(
            0.25
        )

    return {
        "stopped": False,
        "pid": pid,
        "message": (
            "Scheduler did not stop within the timeout."
        ),
    }


def build_scheduler_configuration(
    arguments: argparse.Namespace,
) -> SchedulerConfiguration:
    return SchedulerConfiguration(
        interval_minutes=int(
            arguments.schedule_interval_minutes
        ),
        reference_path=arguments.reference,
        current_path=arguments.current,
        config_path=arguments.config,
        generate_reports=(
            not arguments.no_reports
        ),
        generate_pdf=(
            not arguments.no_pdf
        ),
        persist_history=(
            not arguments.no_history
        ),
        run_immediately=(
            not arguments.schedule_wait_first
        ),
    )


def parse_complete_monitoring_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "HeveMind semiconductor model monitoring, drift detection "
            "and scheduled production monitoring."
        )
    )

    parser.add_argument(
        "--initialise",
        action="store_true",
    )

    parser.add_argument(
        "--run",
        action="store_true",
    )

    parser.add_argument(
        "--reference",
        type=Path,
    )

    parser.add_argument(
        "--current",
        type=Path,
    )

    parser.add_argument(
        "--config",
        type=Path,
    )

    parser.add_argument(
        "--status",
        action="store_true",
    )

    parser.add_argument(
        "--validate",
        action="store_true",
    )

    parser.add_argument(
        "--no-history",
        action="store_true",
    )

    parser.add_argument(
        "--no-reports",
        action="store_true",
    )

    parser.add_argument(
        "--no-pdf",
        action="store_true",
    )

    parser.add_argument(
        "--json-output",
        type=Path,
    )

    parser.add_argument(
        "--schedule",
        action="store_true",
        help=(
            "Run monitoring repeatedly at the configured interval."
        ),
    )

    parser.add_argument(
        "--schedule-once",
        action="store_true",
        help=(
            "Execute one lock-protected scheduled monitoring iteration."
        ),
    )

    parser.add_argument(
        "--schedule-interval-minutes",
        type=int,
        default=1440,
        help=(
            "Scheduler interval in minutes. Default: 1440 (daily)."
        ),
    )

    parser.add_argument(
        "--schedule-wait-first",
        action="store_true",
        help=(
            "Wait for the first interval before the initial run."
        ),
    )

    parser.add_argument(
        "--scheduler-status",
        action="store_true",
    )

    parser.add_argument(
        "--stop-scheduler",
        action="store_true",
    )

    return parser.parse_args()


def complete_monitoring_main() -> None:
    arguments = parse_complete_monitoring_arguments()

    action_requested = False

    if arguments.initialise:
        payload = initialise_monitoring_environment()

        ensure_scheduler_directories()

        payload[
            "scheduler_directory"
        ] = str(
            SCHEDULER_DIR
        )

        print(
            json.dumps(
                json_safe(
                    payload
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        action_requested = True

    if arguments.run:
        bundle, artifacts = run_complete_monitoring_pipeline(
            reference_path=arguments.reference,
            current_path=arguments.current,
            config_path=arguments.config,
            save_output=True,
            persist_history=(
                not arguments.no_history
            ),
            generate_reports=(
                not arguments.no_reports
            ),
            generate_pdf=(
                not arguments.no_pdf
            ),
        )

        print(
            monitoring_pipeline_console_output(
                bundle,
                artifacts,
            )
        )

        summary = {
            "run_identity": asdict(
                bundle.context.run_identity
            ),
            "deployment_health": asdict(
                bundle.health_assessment
            ),
            "feature_drift_summary": asdict(
                bundle.feature_summary
            ),
            "prediction_monitoring_summary": asdict(
                bundle.prediction_summary
            ),
            "report_artifacts": (
                asdict(
                    artifacts
                )
                if artifacts is not None
                else None
            ),
            "model_modification_performed": False,
        }

        if arguments.json_output is not None:
            output_path = (
                arguments.json_output.expanduser()
            )

            if not output_path.is_absolute():
                output_path = (
                    ROOT_DIR
                    / output_path
                )

            atomic_write_json(
                output_path,
                summary,
            )

            print(
                f"\nJSON run summary: {output_path}"
            )

        action_requested = True

    if arguments.schedule_once:
        configuration = build_scheduler_configuration(
            arguments
        )

        bundle, artifacts = scheduled_monitoring_run(
            configuration
        )

        print(
            monitoring_pipeline_console_output(
                bundle,
                artifacts,
            )
        )

        action_requested = True

    if arguments.schedule:
        configuration = build_scheduler_configuration(
            arguments
        )

        run_monitoring_scheduler(
            configuration
        )

        action_requested = True

    if arguments.scheduler_status:
        print(
            json.dumps(
                json_safe(
                    load_scheduler_state()
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        action_requested = True

    if arguments.stop_scheduler:
        print(
            json.dumps(
                stop_monitoring_scheduler(),
                indent=2,
                ensure_ascii=False,
            )
        )

        action_requested = True

    if arguments.status:
        payload = monitoring_status_payload()

        payload[
            "scheduler"
        ] = load_scheduler_state()

        print(
            json.dumps(
                json_safe(
                    payload
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        action_requested = True

    if arguments.validate:
        validation = validate_existing_monitoring_outputs()

        print(
            json.dumps(
                json_safe(
                    validation
                ),
                indent=2,
                ensure_ascii=False,
            )
        )

        if not validation[
            "valid"
        ]:
            raise SystemExit(
                1
            )

        action_requested = True

    if not action_requested:
        status = monitoring_status_payload()

        scheduler = load_scheduler_state()

        print(
            "\n"
            + "="
            * 118
        )

        print(
            "HEVEMIND MODEL MONITORING AND DRIFT DETECTION"
        )

        print(
            "="
            * 118
        )

        print(
            f"\nVersion:                  {APP_VERSION}"
        )

        print(
            f"Outputs available:        {status['outputs_available']}"
        )

        print(
            f"Scheduler status:         {scheduler.get('status')}"
        )

        print(
            f"Scheduler running:        {scheduler.get('running')}"
        )

        health = status.get(
            "deployment_health",
            {},
        )

        if health:
            print(
                "Latest health score:     "
                f"{health.get('overall_health_score')}"
            )

            print(
                "Latest recommendation:   "
                f"{health.get('recommendation')}"
            )

        print(
            "\nRun with --help to view monitoring commands."
        )


if __name__ == "__main__":
    if streamlit_context_active():
        import streamlit as st

        st.set_page_config(
            page_title=(
                "HeveMind Model Monitoring"
            ),
            page_icon=None,
            layout="wide",
            initial_sidebar_state="expanded",
        )

        render_model_monitoring_dashboard()

    else:
        complete_monitoring_main()
